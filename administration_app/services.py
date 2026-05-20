'''
# Использование в любом месте проекта
from templates.services import TemplateService

# Генерация договора
context = {
    'client_name': 'ООО "Ромашка"',
    'contract_number': '123/2024',
    'contract_date': '2024-01-15',
    'total_amount': '1 000 000 руб.'
}

# Создание документа
doc_buffer = TemplateService.render_word_template('contract_template', context)

# Сохранение или отправка пользователю
with open('generated_contract.docx', 'wb') as f:
    f.write(doc_buffer.getvalue())
'''

# services.py
from django.core.files.storage import default_storage
from django.conf import settings
from django.utils import timezone

from .models import TemplateDocument
import os
from docxtpl import DocxTemplate
from openpyxl import load_workbook
import io


class TemplateService:
    """
    Сервис для работы с шаблонами документов
    """

    @staticmethod
    def get_template_path(unique_code):
        """
        Получить путь к актуальному шаблону
        """
        template = TemplateDocument.get_active_template(unique_code)
        if not template:
            raise ValueError(f"Активный шаблон с кодом '{unique_code}' не найден")

        return template.template_file.path

    @staticmethod
    def render_word_template(unique_code, context):
        """
        Заполнить Word шаблон данными
        """
        template_path = TemplateService.get_template_path(unique_code)

        # Загружаем шаблон
        doc = DocxTemplate(template_path)

        # Заполняем контекстом
        doc.render(context)

        # Сохраняем в буфер
        buffer = io.BytesIO()
        doc.save(buffer)
        buffer.seek(0)

        return buffer

    @staticmethod
    def render_excel_template(unique_code, context):
        """
        Заполнить Excel шаблон данными
        """
        template_path = TemplateService.get_template_path(unique_code)

        # Загружаем шаблон
        wb = load_workbook(template_path)

        # Пример заполнения (настройте под свои нужды)
        for sheet_name, data in context.get('sheets', {}).items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell, value in data.items():
                    ws[cell] = value

        # Сохраняем в буфер
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    @staticmethod
    def get_template_info(unique_code):
        """
        Получить информацию о шаблоне
        """
        template = TemplateDocument.get_active_template(unique_code)
        if not template:
            return None

        return {
            'name': template.name,
            'version': template.version,
            'type': template.template_type,
            'valid_until': template.end_date,
            'file_size': template.template_file.size,
        }

    @staticmethod
    def create_new_version(unique_code, new_file, created_by=None, **kwargs):
        """
        Создать новую версию шаблона
        """
        old_template = TemplateDocument.get_active_template(unique_code)

        new_template = TemplateDocument(
            name=kwargs.get('name', old_template.name if old_template else unique_code),
            unique_code=unique_code,
            template_type=kwargs.get('template_type',
                                     old_template.template_type if old_template else 'word'),
            template_file=new_file,
            start_date=kwargs.get('start_date', timezone.now()),
            end_date=kwargs.get('end_date'),
            description=kwargs.get('description', ''),
            created_by=created_by,
            is_active=True  # Новая версия становится активной
        )

        # Если старая версия существует, деактивируем её
        if old_template:
            old_template.is_active = False
            old_template.save()

        new_template.save()
        return new_template
