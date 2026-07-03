import io
from datetime import date, timedelta
from typing import Any
from django.db.models import Sum, Q, QuerySet
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

from finance_app.models import (
    FinancialContract,
    FinancialObligation,
    PaymentSchedule,
    PaymentFact,
    CreditAgreement,
    CreditPaymentSchedule,
    CreditPaymentFact,
)
from finance_app.services.finance_selector import FinanceSelector

User = get_user_model()

# Регистрация кириллического шрифта DejaVuSans для reportlab
FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
FONT_BOLD_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

if os.path.exists(FONT_PATH):
    pdfmetrics.registerFont(TTFont("DejaVuSans", FONT_PATH))
if os.path.exists(FONT_BOLD_PATH):
    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", FONT_BOLD_PATH))


class ReportService:
    """
    Сервис для сбора данных и генерации финансовых отчетов в XLSX и PDF форматах.
    """

    @classmethod
    def get_report_data(cls, report_type: str, user: User, year: int | None = None, month: int | None = None) -> dict[str, Any]:
        """
        Собирает данные для отчета на основе его типа и прав доступа пользователя.

        Args:
            report_type: Тип отчета.
            user: Пользователь системы.

        Returns:
            Словарь с заголовками столбцов и списком строк данных.
        """
        today = date.today()
        selected_year = year or today.year
        
        # Применяем фильтр видимости договоров и кредитов
        contracts = FinanceSelector.apply_visibility_filter(user, FinancialContract.objects.all())
        credits_qs = FinanceSelector.apply_visibility_filter(user, CreditAgreement.objects.all())
        
        if year:
            contracts = contracts.filter(date_conclusion__year=selected_year)
            credits_qs = credits_qs.filter(contract_date__year=selected_year)
        if month:
            contracts = contracts.filter(date_conclusion__month=month)
            credits_qs = credits_qs.filter(contract_date__month=month)
            
        obligations = FinancialObligation.objects.filter(contract__in=contracts)

        headers: list[str] = []
        rows: list[list[Any]] = []
        title = ""

        if report_type == "contracts_registry":
            title = "Реестр финансовых договоров"
            headers = ["Номер договора", "Дата", "Контрагент", "Организация", "Сумма договора", "Валюта", "Начало", "Окончание", "Ответственный", "Статус"]
            for c in contracts.select_related("counteragent", "organization", "employee"):
                rows.append([
                    c.contract_number,
                    c.date_conclusion.strftime("%d.%m.%Y") if c.date_conclusion else "",
                    c.counteragent.short_name,
                    c.organization.name,
                    float(c.cost),
                    c.currency,
                    c.start_date.strftime("%d.%m.%Y") if c.start_date else "",
                    c.end_date.strftime("%d.%m.%Y") if c.end_date else "Не ограничен",
                    c.employee.get_full_name() or c.employee.username if c.employee else "",
                    c.get_status_display()
                ])

        elif report_type == "credits_registry":
            title = "Реестр кредитов и займов"
            headers = ["Банк", "Номер договора", "Дата договора", "Сумма кредита", "Ставка (%)", "Срок (мес)", "Остаток долга", "Ответственный"]
            for cr in credits_qs.select_related("bank", "employee"):
                rows.append([
                    cr.bank.short_name,
                    cr.contract_number,
                    cr.contract_date.strftime("%d.%m.%Y") if cr.contract_date else "",
                    float(cr.amount),
                    float(cr.interest_rate),
                    cr.term_months,
                    float(cr.remaining_debt),
                    cr.employee.get_full_name() or cr.employee.username if cr.employee else ""
                ])

        elif report_type == "payment_calendar":
            title = "Платежный календарь"
            headers = ["Дата платежа", "Тип обязательства", "Договор", "Контрагент", "Ожидаемое поступление", "Плановый платеж", "Статус"]
            
            # Собираем плановые платежи и поступления
            schedules = PaymentSchedule.objects.filter(obligation__in=obligations)
            if year:
                schedules = schedules.filter(payment_date__year=selected_year)
            if month:
                schedules = schedules.filter(payment_date__month=month)
                
            schedules = schedules.select_related(
                "obligation__contract__counteragent", "obligation__obligation_type", "obligation__contract"
            )
            for s in schedules:
                is_inflow = s.obligation.obligation_type.code == "customer"
                rows.append([
                    s.payment_date.strftime("%d.%m.%Y"),
                    s.obligation.obligation_type.name,
                    s.obligation.contract.contract_number,
                    s.obligation.contract.counteragent.short_name,
                    float(s.amount) if is_inflow else 0.00,
                    float(s.amount) if not is_inflow else 0.00,
                    s.get_status_display()
                ])

            # Кредиты
            credit_schedules = CreditPaymentSchedule.objects.filter(credit_agreement__in=credits_qs)
            if year:
                credit_schedules = credit_schedules.filter(payment_date__year=selected_year)
            if month:
                credit_schedules = credit_schedules.filter(payment_date__month=month)
                
            credit_schedules = credit_schedules.select_related(
                "credit_agreement__bank"
            )
            for cs in credit_schedules:
                rows.append([
                    cs.payment_date.strftime("%d.%m.%Y"),
                    "Кредитный платеж",
                    cs.credit_agreement.contract_number,
                    cs.credit_agreement.bank.short_name,
                    0.00,
                    float(cs.total_amount),
                    cs.get_status_display()
                ])
            rows.sort(key=lambda x: x[0])

        elif report_type == "creditor_debts":
            title = "Кредиторская задолженность"
            headers = ["Договор", "Контрагент", "Сумма обязательств", "Оплачено", "Остаток задолженности", "Срок оплаты"]
            # Обязательства перед поставщиками (исключаем покупателей)
            creditor_obs = obligations.exclude(obligation_type__code="customer").select_related("contract__counteragent")
            for ob in creditor_obs:
                paid = PaymentFact.objects.filter(obligation=ob).aggregate(total=Sum("amount"))["total"] or 0.00
                debt = max(0.00, float(ob.cost) - float(paid))
                rows.append([
                    ob.contract.contract_number,
                    ob.contract.counteragent.short_name,
                    float(ob.cost),
                    float(paid),
                    debt,
                    ob.date_execution.strftime("%d.%m.%Y") if ob.date_execution else ""
                ])

        elif report_type == "debtor_debts":
            title = "Дебиторская задолженность"
            headers = ["Договор", "Контрагент", "Сумма договора", "Поступило платежей", "Остаток задолженности", "Дата ожидания"]
            # Договоры с покупателями
            debtor_obs = obligations.filter(obligation_type__code="customer").select_related("contract__counteragent")
            for ob in debtor_obs:
                paid = PaymentFact.objects.filter(obligation=ob).aggregate(total=Sum("amount"))["total"] or 0.00
                debt = max(0.00, float(ob.cost) - float(paid))
                rows.append([
                    ob.contract.contract_number,
                    ob.contract.counteragent.short_name,
                    float(ob.cost),
                    float(paid),
                    debt,
                    ob.date_execution.strftime("%d.%m.%Y") if ob.date_execution else ""
                ])

        elif report_type == "overdue_obligations":
            title = "Просроченные финансовые обязательства"
            headers = ["Договор", "Контрагент", "Тип обязательства", "Сумма", "Срок оплаты", "Дней просрочки"]
            
            # Обычные просроченные плановые платежи
            overdue_scheds = PaymentSchedule.objects.filter(
                obligation__in=obligations, status="overdue"
            ).select_related("obligation__contract__counteragent", "obligation__obligation_type", "obligation__contract")
            for s in overdue_scheds:
                rows.append([
                    s.obligation.contract.contract_number,
                    s.obligation.contract.counteragent.short_name,
                    s.obligation.obligation_type.name,
                    float(s.amount),
                    s.payment_date.strftime("%d.%m.%Y"),
                    (today - s.payment_date).days
                ])

            # Просроченные кредиты
            overdue_credits = CreditPaymentSchedule.objects.filter(
                credit_agreement__in=credits_qs, status="overdue"
            ).select_related("credit_agreement__bank")
            for cs in overdue_credits:
                rows.append([
                    cs.credit_agreement.contract_number,
                    cs.credit_agreement.bank.short_name,
                    "Кредитный платеж",
                    float(cs.total_amount),
                    cs.payment_date.strftime("%d.%m.%Y"),
                    (today - cs.payment_date).days
                ])
            rows.sort(key=lambda x: -x[5])

        elif report_type == "payment_plan_fact":
            title = "План-факт оплат (Расходы)"
            headers = ["Срок оплаты", "Договор", "Контрагент", "Запланировано", "Оплачено фактически", "Отклонение"]
            
            schedules = PaymentSchedule.objects.filter(
                obligation__in=obligations.exclude(obligation_type__code="customer")
            ).select_related("obligation__contract__counteragent", "obligation__contract")
            
            for s in schedules:
                # Находим факты оплаты по этому обязательству до даты платежа + допуск
                facts = PaymentFact.objects.filter(
                    obligation=s.obligation,
                    payment_date__lte=s.payment_date + timedelta(days=5)
                ).aggregate(total=Sum("amount"))["total"] or 0.00
                deviation = float(s.amount) - float(facts)
                rows.append([
                    s.payment_date.strftime("%d.%m.%Y"),
                    s.obligation.contract.contract_number,
                    s.obligation.contract.counteragent.short_name,
                    float(s.amount),
                    float(facts),
                    deviation
                ])

        elif report_type == "inflow_plan_fact":
            title = "План-факт поступлений (Доходы)"
            headers = ["Срок ожидания", "Договор", "Контрагент", "Запланировано", "Поступило фактически", "Отклонение"]
            
            schedules = PaymentSchedule.objects.filter(
                obligation__in=obligations.filter(obligation_type__code="customer")
            ).select_related("obligation__contract__counteragent", "obligation__contract")
            
            for s in schedules:
                facts = PaymentFact.objects.filter(
                    obligation=s.obligation,
                    payment_date__lte=s.payment_date + timedelta(days=5)
                ).aggregate(total=Sum("amount"))["total"] or 0.00
                deviation = float(s.amount) - float(facts)
                rows.append([
                    s.payment_date.strftime("%d.%m.%Y"),
                    s.obligation.contract.contract_number,
                    s.obligation.contract.counteragent.short_name,
                    float(s.amount),
                    float(facts),
                    deviation
                ])

        elif report_type == "cash_flow":
            title = "Отчет о движении денежных средств"
            headers = ["Дата", "Документ №", "Договор", "Контрагент", "Тип обязательства", "Приход (Принято)", "Расход (Списано)"]
            
            # Фактические платежи по обычным договорам
            facts = PaymentFact.objects.filter(obligation__in=obligations)
            if year:
                facts = facts.filter(payment_date__year=year)
            if month:
                facts = facts.filter(payment_date__month=month)
                
            facts = facts.select_related(
                "obligation__contract__counteragent", "obligation__obligation_type", "obligation__contract"
            )
            for f in facts:
                is_inflow = f.obligation.obligation_type.code == "customer"
                rows.append([
                    f.payment_date.strftime("%d.%m.%Y"),
                    f.payment_doc_number,
                    f.obligation.contract.contract_number,
                    f.obligation.contract.counteragent.short_name,
                    f.obligation.obligation_type.name,
                    float(f.amount) if is_inflow else 0.00,
                    float(f.amount) if not is_inflow else 0.00
                ])

            # Фактические платежи по кредитам
            credit_facts = CreditPaymentFact.objects.filter(credit_agreement__in=credits_qs)
            if year:
                credit_facts = credit_facts.filter(payment_date__year=year)
            if month:
                credit_facts = credit_facts.filter(payment_date__month=month)
                
            credit_facts = credit_facts.select_related(
                "credit_agreement__bank"
            )
            for cf in credit_facts:
                rows.append([
                    cf.payment_date.strftime("%d.%m.%Y"),
                    cf.payment_doc_number,
                    cf.credit_agreement.contract_number,
                    cf.credit_agreement.bank.short_name,
                    "Кредит",
                    0.00,
                    float(cf.amount)
                ])
            rows.sort(key=lambda x: x[0])

        return {
            "title": title,
            "headers": headers,
            "rows": rows
        }

    @classmethod
    def generate_xlsx(cls, report_type: str, user: User, year: int | None = None, month: int | None = None) -> bytes:
        """
        Формирует отчет в формате XLSX.

        Args:
            report_type: Тип отчета.
            user: Пользователь системы.
            year: Год.
            month: Месяц.

        Returns:
            Бинарные данные XLSX-файла.
        """
        data = cls.get_report_data(report_type, user, year, month)
        wb = Workbook()
        ws = wb.active
        ws.title = "Финансовый отчет"

        # Стили
        font_title = Font(name="Calibri", size=16, bold=True, color="1F497D")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11)
        font_total = Font(name="Calibri", size=11, bold=True)
        
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_total = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        border_thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )
        border_total = Border(
            top=Side(style="thin", color="1F497D"),
            bottom=Side(style="double", color="1F497D")
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Заголовок отчета
        ws.append([])
        ws.cell(row=2, column=1, value=data["title"]).font = font_title
        ws.append([])

        # Заголовки колонок
        ws.append(data["headers"])
        header_row_idx = 4
        for col_idx in range(1, len(data["headers"]) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_thin

        # Данные
        start_row = 5
        for row_data in data["rows"]:
            ws.append(row_data)
            current_row_idx = ws.max_row
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=current_row_idx, column=col_idx)
                cell.font = font_data
                cell.border = border_thin
                
                # Форматирование ячеек
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
                    cell.alignment = align_right
                elif isinstance(val, int):
                    cell.number_format = "#,##0"
                    cell.alignment = align_right
                elif col_idx in [1, 2, 7, 8]:  # Даты, номера
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # Строка итогов для числовых отчетов
        numeric_cols: list[int] = []
        if report_type in ["contracts_registry", "credits_registry"]:
            numeric_cols = [5] if report_type == "contracts_registry" else [4, 7]
        elif report_type == "payment_calendar":
            numeric_cols = [5, 6]
        elif report_type in ["creditor_debts", "debtor_debts"]:
            numeric_cols = [3, 4, 5]
        elif report_type == "overdue_obligations":
            numeric_cols = [4]
        elif report_type in ["payment_plan_fact", "inflow_plan_fact"]:
            numeric_cols = [4, 5, 6]
        elif report_type == "cash_flow":
            numeric_cols = [6, 7]

        if numeric_cols and data["rows"]:
            total_row_idx = ws.max_row + 1
            ws.cell(row=total_row_idx, column=1, value="ИТОГО:").font = font_total
            ws.cell(row=total_row_idx, column=1).alignment = align_left
            ws.cell(row=total_row_idx, column=1).fill = fill_total
            ws.cell(row=total_row_idx, column=1).border = border_total

            for col_idx in range(2, len(data["headers"]) + 1):
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell.fill = fill_total
                cell.border = border_total
                
                if col_idx in numeric_cols:
                    col_letter = cell.column_letter
                    cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{total_row_idx-1})"
                    cell.font = font_total
                    cell.number_format = "#,##0.00"
                    cell.alignment = align_right

        # Авто-подгон ширины колонок
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if cell.number_format == "#,##0.00" and isinstance(cell.value, (int, float)):
                    val_str = f"{cell.value:,.2f}"
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws.row_dimensions[header_row_idx].height = 25

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @classmethod
    def generate_pdf(cls, report_type: str, user: User, year: int | None = None, month: int | None = None) -> bytes:
        """
        Формирует отчет в формате PDF.

        Args:
            report_type: Тип отчета.
            user: Пользователь системы.
            year: Год.
            month: Месяц.

        Returns:
            Бинарные данные PDF-файла.
        """
        data = cls.get_report_data(report_type, user, year, month)
        buffer = io.BytesIO()

        # Размеры полей и альбомная ориентация
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        elements: list[Any] = []
        styles = getSampleStyleSheet()

        # Добавим свои стили с использованием зарегистрированного шрифта DejaVuSans
        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Heading1"],
            fontName="DejaVuSans-Bold" if pdfmetrics.getRegisteredFont("DejaVuSans-Bold") else "Helvetica-Bold",
            fontSize=16,
            textColor="black",
            alignment=0, # Слева
            spaceAfter=15
        )
        
        cell_style = ParagraphStyle(
            name="ReportCell",
            parent=styles["Normal"],
            fontName="DejaVuSans" if pdfmetrics.getRegisteredFont("DejaVuSans") else "Helvetica",
            fontSize=8,
            leading=10
        )

        cell_bold_style = ParagraphStyle(
            name="ReportCellBold",
            parent=cell_style,
            fontName="DejaVuSans-Bold" if pdfmetrics.getRegisteredFont("DejaVuSans-Bold") else "Helvetica-Bold"
        )

        elements.append(Paragraph(data["title"], title_style))
        elements.append(Spacer(1, 10))

        # Формируем таблицу для PDF
        pdf_table_data: list[list[Paragraph]] = []
        
        # Шапка таблицы
        headers_p = [Paragraph(f"<b>{h}</b>", cell_bold_style) for h in data["headers"]]
        pdf_table_data.append(headers_p)

        # Тело таблицы
        for r in data["rows"]:
            row_p = []
            for val in r:
                if isinstance(val, float):
                    val_str = f"{val:,.2f}"
                elif isinstance(val, int):
                    val_str = f"{val:,}"
                else:
                    val_str = str(val)
                row_p.append(Paragraph(val_str, cell_style))
            pdf_table_data.append(row_p)

        # Добавление строки Итогов для PDF
        numeric_cols = []
        if report_type in ["contracts_registry", "credits_registry"]:
            numeric_cols = [4] if report_type == "contracts_registry" else [3, 6]
        elif report_type == "payment_calendar":
            numeric_cols = [4, 5]
        elif report_type in ["creditor_debts", "debtor_debts"]:
            numeric_cols = [2, 3, 4]
        elif report_type == "overdue_obligations":
            numeric_cols = [3]
        elif report_type in ["payment_plan_fact", "inflow_plan_fact"]:
            numeric_cols = [3, 4, 5]
        elif report_type == "cash_flow":
            numeric_cols = [5, 6]

        if numeric_cols and data["rows"]:
            totals = ["ИТОГО:"] + [""] * (len(data["headers"]) - 1)
            for col_idx in numeric_cols:
                col_sum = 0.00
                for r in data["rows"]:
                    try:
                        col_sum += float(r[col_idx])
                    except (ValueError, TypeError):
                        pass
                totals[col_idx] = f"{col_sum:,.2f}"

            totals_p = [Paragraph(f"<b>{t}</b>", cell_bold_style) for t in totals]
            pdf_table_data.append(totals_p)

        # Вычисляем примерную ширину колонок для A4 альбомного (доступная ширина ~750 единиц)
        col_widths = []
        num_cols = len(data["headers"])
        base_width = 750 / num_cols
        for i in range(num_cols):
            # Можно сделать более интеллектуальный подбор, но пока сделаем равные
            col_widths.append(base_width)

        t = Table(pdf_table_data, colWidths=col_widths, repeatRows=1)
        
        # Стилизация таблицы
        t_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), "#1F497D"),
            ("TEXTCOLOR", (0, 0), (-1, 0), "#FFFFFF"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, "#D9D9D9"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), ["#FFFFFF", "#F2F5F9"]),
        ])

        # Стилизуем строку Итогов, если она добавлена
        if numeric_cols and data["rows"]:
            t_style.add("BACKGROUND", (0, -1), (-1, -1), "#DCE6F1")
            t_style.add("LINEABOVE", (0, -1), (-1, -1), 1.5, "#1F497D")
            t_style.add("LINEBELOW", (0, -1), (-1, -1), 1.5, "#1F497D")

        t.setStyle(t_style)
        elements.append(t)

        doc.build(elements)
        return buffer.getvalue()
