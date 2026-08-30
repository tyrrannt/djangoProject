import datetime
import re

from decouple import config
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Q
from django.http import QueryDict, JsonResponse, HttpResponse
from django.shortcuts import HttpResponseRedirect, redirect, render
from django.views.generic import DetailView, UpdateView, ListView, CreateView, DeleteView
from dadata import Dadata
from administration_app.models import PortalProperty
from administration_app.utils import int_validate, ajax_search
from contracts_app.models import Contract, Posts, TypeContract, TypeProperty, TypeDocuments, Estate
from contracts_app.forms import ContractsAddForm, ContractsPostAddForm, ContractsUpdateForm, TypeDocumentsUpdateForm, \
    TypeDocumentsAddForm, TypeContractsAddForm, TypeContractsUpdateForm, TypePropertysUpdateForm, TypePropertysAddForm, \
    EstateAddForm, EstateUpdateForm
from django.urls import reverse, reverse_lazy
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from customers_app.models import DataBaseUser, CounteragentDocuments

from core import logger


class ContractList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """
    Отображение списка договоров
    """
    model = Contract
    permission_required = 'contracts_app.view_contract'

    def get_context_data(self, **kwargs):
        context = super(ContractList, self).get_context_data(**kwargs)
        context['title'] = f'База договоров'
        return context

    def get_queryset(self):
        access = self.request.user.user_access
        return Contract.objects.filter(Q(allowed_placed=True) & Q(access_id__gte=access))

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        # if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        #     contract_list = Contract.objects.filter(type_of_document__type_document='Договор').order_by('pk').reverse()
        #     data = [contract_item.get_data() for contract_item in contract_list]
        #     response = {'data': data}
        #     # report_card_separator()
        #     return JsonResponse(response)
        # return super().get(request, *args, **kwargs)

        access = self.request.user.user_access
        query = (Q(parent_category__isnull=True)) & Q(access_id__gte=access)
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            search_list = ['actuality', 'contract_number', 'date_conclusion', 'type_of_document__type_document',
                           'type_of_contract__type_contract', 'subject_contract',
                           'contract_counteragent__short_name', ]
            try:
                context = ajax_search(request, self, search_list, Contract, query, triger=1)
            except Exception as e:
                context = ajax_search(request, self, search_list, Contract, query, triger=1)
                logger.error(e)
            return JsonResponse(context, safe=False)
        return super().get(request, *args, **kwargs)


class ContractListAdmin(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """
    Отображение списка договоров
    """
    model = Contract
    template_name = 'contracts_app/contract_list_admin.html'
    permission_required = 'contracts_app.view_contract'

    def get_context_data(self, **kwargs):
        context = super(ContractListAdmin, self).get_context_data(**kwargs)
        context['title'] = f'База договоров'
        return context

    def get_queryset(self):
        access = self.request.user.user_access
        return Contract.objects.filter(
            Q(allowed_placed=True) &
            Q(access_id__gte=access) &
            ~Q(doc_file__endswith='.pdf')
        )

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        # if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        #     contract_list = Contract.objects.filter(type_of_document__type_document='Договор').order_by('pk').reverse()
        #     data = [contract_item.get_data() for contract_item in contract_list]
        #     response = {'data': data}
        #     # report_card_separator()
        #     return JsonResponse(response)
        # return super().get(request, *args, **kwargs)
        access = self.request.user.user_access
        query = (Q(type_of_document__type_document='Договор') & Q(access_id__gte=access) &
                 ~Q(doc_file__iendswith='.pdf') & ~Q(doc_file__iendswith='.docx'))
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            search_list = ['actuality', 'contract_number', 'date_conclusion',
                           'type_of_contract__type_contract', 'subject_contract',
                           'contract_counteragent__short_name', ]
            context = ajax_search(request, self, search_list, Contract, query)
            return JsonResponse(context, safe=False)
        return super().get(request, *args, **kwargs)


class ContractSearch(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """
    Поиск договоров в базе
    ToDo: Не работает пагинация при прямом открытии списка. Разобраться почему!!! После нажатия кнопки поиска, все норм.
    """
    template_name_suffix = '_search'
    context_object_name = 'object'
    object_list = None
    paginate_by = 6
    permission_required = 'contracts_app.view_contract'

    def post(self, request):  # ***** this method required! ******
        self.object_list = self.get_queryset()
        return HttpResponseRedirect(reverse('contracts_app:search'))

    # Работает с GET запросом
    def get_queryset(self):
        query = Q()
        query &= Q(allowed_placed=True)
        # query &= Q(access__pk__gte=DataBaseUser.objects.get(
        #     pk=self.request.user.pk).access_level.contracts_access_view.level)
        qs = Contract.objects.filter(query).order_by('pk')
        if self.request.GET:
            dv = self.request.GET.get('dv')
            ca = self.request.GET.get('ca')
            tc = self.request.GET.get('tc')
            tp = self.request.GET.get('tp')
            cn = self.request.GET.get('cn')
            sn = self.request.GET.get('sn')
            """Формируем запрос на лету, в зависимости от полученных параметров, создаем Q объект,
               и добавляем к нему запросы, в зависимости от значений передаваемых параметров.
            """
            if dv != '0':
                query &= Q(divisions=int_validate(dv))
            if ca != '0':
                query &= Q(contract_counteragent=int_validate(ca))
            if tc != '0':
                query &= Q(type_of_contract=int_validate(tc))
            if tp != '0':
                query &= Q(type_property=int_validate(tp))
            if cn:
                query &= Q(contract_number__contains=cn)
            if sn:
                query &= Q(subject_contract__contains=sn)
            qs = Contract.objects.filter(query).order_by('pk')
        return qs

    ##ToDo: Доработать передачу поискового запроса через POST
    ## Работает с POST запросом
    # def get_queryset(self):
    #     qs = Contract.objects.all()
    #     if self.request.POST:
    #         print(self.request.POST)
    #         return Contract.objects.filter(divisions=int(self.request.POST.get('division')))
    #     return qs

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        # Формируем строку GET запроса при пагинации
        get_request_string = f"dv={self.request.GET.get('dv')}&ca={self.request.GET.get('ca')}" \
                             f"&tc={self.request.GET.get('tc')}&tp={self.request.GET.get('tp')}" \
                             f"&cn={self.request.GET.get('cn')}&sn={self.request.GET.get('sn')}&"
        if get_request_string == 'dv=None&ca=None&tc=None&tp=None&cn=None&sn=None&':
            context['s'] = ''
        else:
            context['s'] = get_request_string

        context['title'] = f'Поиск по базе договоров'
        return context


class ContractAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    """
    Создание нового договора
    """
    model = Contract
    form_class = ContractsAddForm
    # success_url = reverse_lazy('contracts_app:index')
    permission_required = 'contracts_app.add_contract'

    def form_valid(self, form):
        # Сохраняем QueryDict в переменную content для возможности его редактирования
        # content = QueryDict.copy(self.request.POST)
        # Проверяем на корректность ввода головного документа, если головной документ не указан, то вырезаем его
        # if content['parent_category']:
        # print(content['parent_category'])
        # content.setlist('parent_category', '')
        # Проверяем подразделения, если пришел список с 0 значением, то удаляем его
        refreshed_form = form.save(commit=False)
        # if refreshed_form.parent_category:
        #     refreshed_form.parent_category = Contract.objects.get(pk=refreshed_form.parent_category)
        refreshed_form.official_information = refreshed_form.doc_file
        filename = str(refreshed_form.doc_file)
        if refreshed_form.parent_category:
            refreshed_form.comment = filename.split('/')[-1]
        else:
            if refreshed_form.comment == '':
                refreshed_form.comment = filename.split('/')[-1]

        refreshed_form.save()

        return super().form_valid(form)

    def form_invalid(self, form):
        # print('Invalid form', form)
        # print(form['parent_category'])
        return super().form_invalid(form)

    def get_success_url(self):
        obj = self.object
        if obj.parent_category:
            return reverse('contracts_app:detail', kwargs={'pk': obj.parent_category.pk})
        else:
            return reverse('contracts_app:index')

    def post(self, request, *args, **kwargs):
        # Сохраняем QueryDict в переменную content для возможности его редактирования
        content = QueryDict.copy(self.request.POST)
        # if content['parent_category'] == content['contract_counteragent']:
        #     print(content['parent_category'], content['contract_counteragent'])
        # Проверяем на корректность ввода головного документа, если головной документ не указан, то вырезаем его
        # if content['parent_category'] == 'none':
        #     content.setlist('parent_category', '')
        # Проверяем подразделения, если пришел список с 0 значением, то удаляем его из списка, генерируя новый список
        division = list(k for k in content.getlist('divisions') if k != '0')
        type_propertyes = list(k for k in content.getlist('type_property') if k != '0')
        content.setlist('divisions', division)
        content.setlist('type_property', type_propertyes)
        # Возвращаем измененный QueryDict обратно в запрос
        self.request.POST = content
        return super(ContractAdd, self).post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(ContractAdd, self).get_context_data(**kwargs)
        if self.request.GET.get('parent'):
            context['parent'] = self.request.GET.get('parent')
        context['title'] = f'Добавить новый договор'
        return context

    def get_form_kwargs(self):
        """
        Передаем в форму текущего пользователя. В форме переопределяем метод __init__
        :return: PK текущего пользователя
        """
        parent = self.request.GET.get('parent', None)
        kwargs = super().get_form_kwargs()
        kwargs.update({'parent': parent})
        kwargs.update({'executor': self.request.user.pk})
        return kwargs

    def get(self, request, *args, **kwargs):

        return super(ContractAdd, self).get(request, *args, **kwargs)


class ContractDetail(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    """
    Просмотр договора.
    """
    model = Contract
    permission_required = 'contracts_app.view_contract'

    def dispatch(self, request, *args, **kwargs):
        try:
            if request.user.is_anonymous:
                return redirect(reverse('customers_app:login'))
            contract_object = self.get_object()
            # if request.user.user_access.pk <= contract_object.access.pk or request.user.is_superuser:
            #     return super(ContractDetail, self).dispatch(request, *args, **kwargs)
            # Используем суффикс _id вместо .access.pk
            if request.user.user_access_id <= contract_object.access_id or request.user.is_superuser:
                return super().dispatch(request, *args, **kwargs)
            else:
                logger.warning(f'Пользователь {request.user} хотел получить доступ к договору {contract_object}')
                raise PermissionDenied
        except PermissionDenied:
            return render(request, "library_app/403.html")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # if context.get('contract').access.level < int(self.request.user.access_level.contracts_access_view):
        # print(context.get('contract').pk)
        # Выбираем из таблицы Posts все записи относящиеся к текущему договору
        # post = Posts.objects.filter(contract_number=self.object.pk)
        # slaves = Contract.objects.filter(Q(parent_category=self.object.pk))
        slaves = Contract.objects.filter(parent_category_id=self.object.pk)
        # Формируем заголовок страницы и передаем в контекст
        if self.object.contract_number:
            cn = self.object.contract_number
        else:
            cn = '(без номера)'
        context[
            'title'] = title = f'{PortalProperty.objects.all().last().portal_name} // Просмотр договора №' + cn + ' от ' + str(
            self.object.date_conclusion)
        # Передаем найденные записи в контекст
        if not self.object.parent_category:
            context['not_parent'] = True
        context['posts'] = Posts.objects.filter(contract_number=self.object).select_related('responsible_person')
        context['slaves'] = slaves
        context['counteragent_docs'] = CounteragentDocuments.objects.filter(package=self.object.contract_counteragent)
        return context

    def get_queryset(self):
        return super().get_queryset().select_related(
            'access',
            'type_of_document',
            'type_of_contract',
            'contract_counteragent'
        ).prefetch_related(
            'type_property',
            'divisions',
            'employee__user_work_profile__job',  # предзагрузка сотрудников и их должностей
            # 'slaves' # Если у вас есть related_name='slaves' для дочерних договоров, раскомментируйте это
        )

    def get_object(self, queryset=None):
        # Кэшируем объект, чтобы не дергать базу дважды (в dispatch и get)
        if not hasattr(self, '_cached_object'):
            self._cached_object = super().get_object(queryset)
        return self._cached_object


class ContractUpdate(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Contract
    form_class = ContractsUpdateForm
    template_name = 'contracts_app/contract_form_update.html'
    permission_required = 'contracts_app.change_contract'

    def dispatch(self, request, *args, **kwargs):
        try:
            if request.user.is_anonymous:
                return redirect(reverse('customers_app:login'))
            contract_object = self.get_object()
            if request.user.user_access_id <= contract_object.access_id:
                return super(ContractUpdate, self).dispatch(request, *args, **kwargs)
            else:
                logger.warning(f'Пользователь {request.user} хотел получить доступ к договору {contract_object}')
                raise PermissionDenied
        except PermissionDenied:
            return render(request, "library_app/403.html")

    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super().get_form(form_class=self.form_class)
        # form.fields['contract_counteragent'].queryset = Counteragent.objects.filter(
        #     pk=self.object.contract_counteragent.pk)
        if self.object.parent_category:
            form.fields['parent_category'].queryset = Contract.objects.filter(
                parent_category=self.object.parent_category.pk)
        else:
            form.fields['parent_category'].queryset = Contract.objects.none()
        return form

    def form_valid(self, form):
        """
        Проверяем корректность переданной формы
        :param form: Передаваемая форма с сайта
        :return: Редирект обратно на страницу с обновленными данными
        """

        if form.is_valid():

            # 1. Достаем старое состояние из БД (оставляем как у вас, это нужно!)
            old_instance = Contract.objects.get(pk=self.object.pk).__dict__

            # 2. Подготавливаем новые данные и сохраняем их
            refreshed_form = form.save(commit=False)
            if refreshed_form.official_information == '':
                refreshed_form.official_information = refreshed_form.doc_file
            filename = str(refreshed_form.doc_file)

            if refreshed_form.parent_category or refreshed_form.comment == '':
                refreshed_form.comment = filename.split('/')[-1]

            # Сохраняем в БД
            refreshed_form.save()

            # в new_instance сохраняем новые значения записи
            new_instance = refreshed_form.__dict__

            # создаем генератор списка
            diffkeys = [k for k in old_instance if old_instance[k] != new_instance[k]]
            message = '<b>Запись внесена автоматически!</b> <u>Внесены изменения</u>:\n'
            for k in diffkeys:
                if k != '_state':
                    message += f'{Contract._meta.get_field(k).verbose_name}: <strike>{old_instance[k]}</strike> -> {new_instance[k]}\n'

            # post_record = Posts(contract_number=Contract.objects.get(pk=self.object.pk), post_description=message,
            #                     responsible_person=DataBaseUser.objects.get(pk=self.request.user.pk))

            # 5. Сохраняем пост без дополнительных обращений к базе (через суффикс _id)
            post_record = Posts(
                contract_number_id=self.object.pk,  # Вместо contract_number=Contract.objects.get(...)
                post_description=message,
                responsible_person_id=self.request.user.pk
                # Вместо responsible_person=DataBaseUser.objects.get(...)
            )
            post_record.save()

            return HttpResponseRedirect(reverse('contracts_app:detail', args=[self.object.pk]))
        else:
            print(f'Что то не то')

    def get(self, request, *args, **kwargs):
        # """
        # Проверка прав доступа на изменение записи. Если прав нет, то пользователь перенаправляется в общую базу.
        # """
        # pk = int(self.request.user.pk)
        # try:
        #     if DataBaseUser.objects.get(pk=pk).access_level.contracts_access_edit:
        #         return super(ContractUpdate, self).get(request, *args, **kwargs)
        #     else:
        #         url_match = reverse_lazy('contracts_app:index')
        #         return redirect(url_match)
        # except Exception as _ex:
        #     url_match = reverse_lazy('contracts_app:index')
        #     return redirect(url_match)
        # contragent = request.GET.get("contragent", None)
        # print(contragent)
        # if contragent:
        #     print(contragent)
        return super(ContractUpdate, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):

        context = super(ContractUpdate, self).get_context_data(**kwargs)
        # Формируем заголовок страницы и передаем в контекст
        if self.object.contract_number:
            cn = self.object.contract_number
        else:
            cn = '(без номера)'
        context[
            'title'] = title = f'{PortalProperty.objects.all().last().portal_name} // Изменить договор №' + cn + ' от ' + str(
            self.object.date_conclusion)

        return context

    def get_success_url(self):
        return reverse_lazy('contracts_app:detail', {'pk': self.object.pk})

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_object(self, queryset=None):
        # Кэшируем объект, чтобы не дергать базу дважды (в dispatch и get)
        if not hasattr(self, '_cached_object'):
            self._cached_object = super().get_object(queryset)
        return self._cached_object


class ContractDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    model = Contract
    permission_required = "contracts_app.delete_contract"
    success_url = reverse_lazy('contracts_app:index')


class ContractPostAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    """
    Добавление записи к договору.
    """
    model = Posts
    form_class = ContractsPostAddForm
    permission_required = 'hrdepartment_app.add_posts'

    def get_success_url(self):
        """
        Переопределяется метод 'get_success_url', для получения номера договора 'pk',
        к которому добавляется запись, для того чтоб вернуться на страницу договора
        :return: Возвращается URL на договор
        """
        pk = self.object.contract_number.pk
        return reverse("contracts_app:detail", kwargs={"pk": pk})


class ContractPostList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    """
    Вывод списка записей, относящихся к конкретному договору
    """
    model = Posts
    permission_required = 'hrdepartment_app.view_posts'

    def get_queryset(self):
        """
        Переопределен метод получения QuerySet. Записи фильтруются исходя из GET запроса, в котором передается
        параметр contract_number.
        :return: Отфильтрованный QuerySet если задан параметр GET, иначе выводит полный список записей  модели Post
        """
        qs = self.model.objects.all()
        search = self.request.GET.get('cn')
        if search:
            qs = qs.filter(contract_number=search)
        return qs


class ContractPostDelete(PermissionRequiredMixin, LoginRequiredMixin, DeleteView):
    """Представление для удаления служебной заметки к договору."""

    model = Posts
    permission_required = 'hrdepartment_app.delete_posts'

    def get_success_url(self) -> str:
        """Возвращает URL перенаправления после успешного удаления заметки.

        Returns:
            str: URL страницы детального просмотра родительского договора либо реестра.
        """
        if self.object.contract_number:
            return reverse('contracts_app:detail', kwargs={'pk': self.object.contract_number.pk})
        return reverse('contracts_app:index')



"""
Типы документов: Список, Добавление, Детализация, Обновление
"""


class TypeDocumentsList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = TypeDocuments
    template_name = 'contracts_app/typedocuments_list.html'
    permission_required = 'hrdepartment_app.view_typedocuments'

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            type_documents_list = TypeDocuments.objects.all()
            data = [type_documents_item.get_data() for type_documents_item in type_documents_list]
            response = {'data': data}
            return JsonResponse(response)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Типы документов'
        return context


class TypeDocumentsAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = TypeDocuments
    form_class = TypeDocumentsAddForm
    template_name = 'contracts_app/typedocuments_add.html'
    permission_required = 'hrdepartment_app.add_typedocuments'

    def get(self, request, *args, **kwargs):
        return super(TypeDocumentsAdd, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typedocuments_list'))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Добавить тип документа'
        return context


class TypeDocumentsDetail(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = TypeDocuments
    template_name = 'contracts_app/typedocuments_detail.html'
    permission_required = 'hrdepartment_app.view_typedocuments'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


class TypeDocumentsUpdate(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = TypeDocuments
    template_name = 'contracts_app/typedocuments_update.html'
    form_class = TypeDocumentsUpdateForm
    permission_required = 'hrdepartment_app.change_typedocuments'

    def get(self, request, *args, **kwargs):
        return super(TypeDocumentsUpdate, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typedocuments', args=[self.object.pk]))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


"""
Типы договоров: Список, Добавление, Детализация, Обновление
"""


class TypeContractsList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = TypeContract
    template_name = 'contracts_app/typecontracts_list.html'
    permission_required = 'hrdepartment_app.view_typecontract'

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            type_contracts_list = TypeContract.objects.all()
            data = [type_contracts_item.get_data() for type_contracts_item in type_contracts_list]
            response = {'data': data}
            return JsonResponse(response)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Типы договоров'
        return context


class TypeContractsAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = TypeContract
    form_class = TypeContractsAddForm
    template_name = 'contracts_app/typecontracts_add.html'
    permission_required = 'hrdepartment_app.add_typecontract'

    def get(self, request, *args, **kwargs):
        return super(TypeContractsAdd, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typecontracts_list'))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Добавить тип договора'
        return context


class TypeContractsDetail(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = TypeContract
    template_name = 'contracts_app/typecontracts_detail.html'
    permission_required = 'hrdepartment_app.view_typecontract'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


class TypeContractsUpdate(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = TypeContract
    template_name = 'contracts_app/typecontracts_update.html'
    form_class = TypeContractsUpdateForm
    permission_required = 'hrdepartment_app.change_typecontract'

    def get(self, request, *args, **kwargs):
        return super(TypeContractsUpdate, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typecontracts', args=[self.object.pk]))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


"""
Типы имущества: Список, Добавление, Детализация, Обновление
"""


class TypePropertysList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = TypeProperty
    template_name = 'contracts_app/typepropertys_list.html'
    permission_required = 'hrdepartment_app.view_typeproperty'

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            type_property_list = TypeProperty.objects.all()
            data = [type_property_item.get_data() for type_property_item in type_property_list]
            response = {'data': data}
            return JsonResponse(response)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Типы имущества'
        return context


class TypePropertysAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = TypeProperty
    form_class = TypePropertysAddForm
    template_name = 'contracts_app/typepropertys_add.html'
    permission_required = 'hrdepartment_app.add_typeproperty'

    def get(self, request, *args, **kwargs):
        return super(TypePropertysAdd, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typepropertys_list'))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Добавить тип имущества'
        return context


class TypePropertysDetail(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = TypeProperty
    template_name = 'contracts_app/typepropertys_detail.html'
    permission_required = 'hrdepartment_app.view_typeproperty'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


class TypePropertysUpdate(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = TypeProperty
    template_name = 'contracts_app/typepropertys_update.html'
    form_class = TypePropertysUpdateForm
    permission_required = 'hrdepartment_app.change_typeproperty'

    def get(self, request, *args, **kwargs):
        return super(TypePropertysUpdate, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:typepropertys', args=[self.object.pk]))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


"""
Имущества: Список, Добавление, Детализация, Обновление
"""


class EstateList(PermissionRequiredMixin, LoginRequiredMixin, ListView):
    model = Estate
    permission_required = 'hrdepartment_app.view_estate'

    def get(self, request, *args, **kwargs):
        # Определяем, пришел ли запрос как JSON? Если да, то возвращаем JSON ответ
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            type_property_list = Estate.objects.all()
            data = [type_property_item.get_data() for type_property_item in type_property_list]
            response = {'data': data}
            return JsonResponse(response)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Имущества'
        return context


class EstateAdd(PermissionRequiredMixin, LoginRequiredMixin, CreateView):
    model = Estate
    form_class = EstateAddForm
    permission_required = 'hrdepartment_app.add_estate'

    def get(self, request, *args, **kwargs):
        return super(EstateAdd, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:estate_list'))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'Добавить имущества'
        return context


class EstateDetail(PermissionRequiredMixin, LoginRequiredMixin, DetailView):
    model = Estate
    permission_required = 'hrdepartment_app.view_estate'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


class EstateUpdate(PermissionRequiredMixin, LoginRequiredMixin, UpdateView):
    model = Estate
    form_class = EstateUpdateForm
    permission_required = 'hrdepartment_app.change_estate'

    def get(self, request, *args, **kwargs):
        return super(EstateUpdate, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        if form.is_valid():
            form.save()
        return HttpResponseRedirect(reverse('contracts_app:estate_list'))

    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=None, **kwargs)
        context['title'] = f'{self.get_object()}'
        return context


def counteragent_check(request):
    if request.method == 'POST':
        data = request.POST
        if data.get('counteragent') == '' and data.get('counteragent_name') == '':
            return HttpResponseRedirect(reverse('contracts_app:counteragent_check'))
        else:
            token = config('FNS')
            ddata = Dadata(token)
            inn = str(data.get('counteragent'))
            kpp = str(data.get('counteragent_kpp'))
            name = str(data.get('counteragent_name')).strip()
            if kpp:
                res = ddata.find_by_id("party", inn, kpp=kpp)
            else:
                if inn:
                    res = ddata.find_by_id("party", inn)
                else:
                    res = ddata.suggest("party", name)

            data = {'query': res}
            return render(request, 'contracts_app/counteragent_check.html', context=data)
    else:
        return render(request, 'contracts_app/counteragent_check.html')


def update_contract_dates_from_comment():
    """
    Обновляет поле date_conclusion из поля comment для всех контрактов,
    кроме тех, у кого тип документа — не 'Договор'.
    """
    pattern = r'от\s+(\d{2}\.\d{2}\.(?:\d{2}|\d{4}))'

    contracts = Contract.objects.exclude(type_of_document__type_document__iexact='Договор')

    updated_count = 0
    for contract in contracts:
        comment = contract.comment or ''
        match = re.search(pattern, comment)

        if match:
            date_str = match.group(1)

            # Обработка сокращённых дат типа 24 -> 2024
            try:
                day, month, year = date_str.split('.')

                if len(year) == 2:
                    year = int(year)
                    # Если год больше 30 — считаем, что это 1900-е, иначе 2000-е
                    year += 1900 if year > 30 else 2000

                date_obj = datetime.datetime(int(year), int(month), int(day)).date()
            except ValueError:
                logger.warning(f'ValueError.')
                continue  # если дата некорректна — пропускаем

            contract.date_conclusion = date_obj
            contract.save(update_fields=["date_conclusion"])
            updated_count += 1
    logger.warning(f'Обновлено {updated_count} записей.')


def parse_date_param(date_str: str | None) -> datetime.date | None:
    """Парсит строковое представление даты в объект datetime.date.

    Поддерживает распространенные форматы (%Y-%m-%d, %d.%m.%Y, %d-%m-%Y, %Y.%m.%d).

    Args:
        date_str (str | None): Входная строка с датой.

    Returns:
        datetime.date | None: Распознанный объект даты или None при невозможности парсинга.
    """
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y.%m.%d"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def generate_contracts_excel_report(contracts, start_date: datetime.date, end_date: datetime.date, user=None):
    """Генерирует стилизованный и структурированный файл отчета по договорам в формате Microsoft Excel (.xlsx).

    Формирует официальную книгу Excel со стилизованной шапкой компании ООО «Авиакомпания «БАРКОЛ»»,
    информационным блоком ключевых показателей (KPI), таблицей с группировкой по родительским
    договорам и подчиненным дополнительным соглашениям, выделением актуальности, форматированием
    дат и денежных сумм, а также итоговой строкой и автоподбором ширины колонок.

    Args:
        contracts (Iterable[Contract]): Выборка экземпляров модели Contract за указанный период.
        start_date (datetime.date): Начальная дата анализируемого периода.
        end_date (datetime.date): Конечная дата анализируемого периода.
        user (DataBaseUser | None): Пользователь, запросивший выгрузку отчета.

    Returns:
        openpyxl.Workbook: Сформированная и оформленная рабочая книга openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт по договорам"
    ws.views.sheetView[0].showGridLines = True

    # Цветовая палитра и типографика
    company_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
    dept_font = Font(name="Calibri", size=10, italic=True, color="475569")
    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    meta_font = Font(name="Calibri", size=9, italic=True, color="64748B")

    kpi_title_font = Font(name="Calibri", size=9, bold=True, color="475569")
    kpi_val_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    kpi_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    tbl_hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    tbl_hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    group_hdr_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    group_hdr_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    data_font = Font(name="Calibri", size=9.5, color="1E293B")
    data_font_child = Font(name="Calibri", size=9.5, color="334155")
    total_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    status_active_font = Font(name="Calibri", size=9.5, bold=True, color="166534")
    status_expired_font = Font(name="Calibri", size=9.5, color="991B1B")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )
    header_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=Side(border_style="medium", color="0F172A"),
    )
    total_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=Side(border_style="thin", color="475569"),
        bottom=Side(border_style="double", color="0F172A"),
    )

    # 1. Шапка документа
    company_name = "ООО «Авиакомпания «БАРКОЛ»"
    ws.merge_cells("A1:K1")
    ws["A1"] = company_name
    ws["A1"].font = company_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"] = "Учет договоров и контрагентов // Корпоративный портал"
    ws["A2"].font = dept_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # Заголовок отчета
    ws.merge_cells("A4:K4")
    start_str = start_date.strftime("%d.%m.%Y") if start_date else ""
    end_str = end_date.strftime("%d.%m.%Y") if end_date else ""
    ws["A4"] = f"ОТЧЁТ ПО ДОГОВОРАМ ЗА ПЕРИОД С {start_str} ПО {end_str}"
    ws["A4"].font = title_font
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 26

    # Мета-информация
    now_str = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    user_str = str(user) if user and user.is_authenticated else "Система"
    ws.merge_cells("A5:K5")
    ws["A5"] = f"Сформирован: {now_str} | Пользователь: {user_str}"
    ws["A5"].font = meta_font
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 18

    # 2. Блок ключевых показателей (KPI)
    total_docs = len(contracts)
    main_docs = sum(1 for c in contracts if c.parent_category_id is None)
    supp_docs = total_docs - main_docs
    counteragents_count = len(
        set(c.contract_counteragent_id for c in contracts if c.contract_counteragent_id)
    )
    total_cost = sum(c.cost or 0.0 for c in contracts)

    kpi_items = [
        ("Всего документов", str(total_docs)),
        ("Основных договоров", str(main_docs)),
        ("ДС и приложений", str(supp_docs)),
        ("Контрагентов", str(counteragents_count)),
        ("Общая сумма (руб.)", f"{total_cost:,.2f}".replace(",", " ")),
    ]

    ws.merge_cells("A7:B7")
    ws.merge_cells("A8:B8")
    ws["A7"] = kpi_items[0][0]
    ws["A8"] = kpi_items[0][1]

    ws.merge_cells("C7:D7")
    ws.merge_cells("C8:D8")
    ws["C7"] = kpi_items[1][0]
    ws["C8"] = kpi_items[1][1]

    ws.merge_cells("E7:F7")
    ws.merge_cells("E8:F8")
    ws["E7"] = kpi_items[2][0]
    ws["E8"] = kpi_items[2][1]

    ws.merge_cells("G7:H7")
    ws.merge_cells("G8:H8")
    ws["G7"] = kpi_items[3][0]
    ws["G8"] = kpi_items[3][1]

    ws.merge_cells("I7:K7")
    ws.merge_cells("I8:K8")
    ws["I7"] = kpi_items[4][0]
    ws["I8"] = kpi_items[4][1]

    for col in range(1, 12):
        c7 = ws.cell(row=7, column=col)
        c7.font = kpi_title_font
        c7.fill = kpi_fill
        c7.alignment = Alignment(horizontal="center", vertical="center")
        c7.border = cell_border

        c8 = ws.cell(row=8, column=col)
        c8.font = kpi_val_font
        c8.fill = kpi_fill
        c8.alignment = Alignment(horizontal="center", vertical="center")
        c8.border = cell_border

    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 22

    # 3. Заголовки таблицы
    headers = [
        "№ п/п",
        "Контрагент",
        "Номер документа",
        "Тип документа",
        "Тип договора",
        "Дата заключения",
        "Срок действия",
        "Стоимость, руб.",
        "Предмет договора",
        "Статус",
        "Примечание",
    ]
    tbl_start_row = 10
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=tbl_start_row, column=col_idx, value=h)
        cell.font = tbl_hdr_font
        cell.fill = tbl_hdr_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[tbl_start_row].height = 28

    # 4. Группировка и вывод данных
    grouped = {}
    for c in contracts:
        parent = c.parent_category or c
        grouped.setdefault(parent, []).append(c)

    current_row = tbl_start_row + 1
    row_num = 1

    for parent, children in grouped.items():
        is_grouped = len(children) > 1 or (len(children) == 1 and children[0].pk != parent.pk)

        # Если родительский договор отсутствует в выборке детей, но является группирующим — выводим строку группы
        if is_grouped and parent not in children:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=11,
            )
            group_cell = ws.cell(
                row=current_row,
                column=1,
                value=f"📁 Основной договор: № {parent.contract_number or '(без номера)'} — {parent.contract_counteragent or 'Контрагент не указан'}",
            )
            group_cell.font = group_hdr_font
            group_cell.fill = group_hdr_fill
            group_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c_idx in range(1, 12):
                ws.cell(row=current_row, column=c_idx).border = cell_border
                ws.cell(row=current_row, column=c_idx).fill = group_hdr_fill
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        for child in children:
            is_child = child.parent_category_id is not None
            prefix = "  └─ " if is_child else ""
            doc_num_text = f"{prefix}{child.contract_number or '(без номера)'}"

            status_text = "Действующий" if getattr(child, "is_past_due", False) else "Срок истёк"
            status_fnt = (
                status_active_font
                if getattr(child, "is_past_due", False)
                else status_expired_font
            )

            date_concl = (
                child.date_conclusion.strftime("%d.%m.%Y") if child.date_conclusion else ""
            )
            date_close = (
                child.closing_date.strftime("%d.%m.%Y") if child.closing_date else "-"
            )

            counteragent_name = (
                str(child.contract_counteragent) if child.contract_counteragent else "-"
            )
            type_doc = str(child.type_of_document) if child.type_of_document else "-"
            type_contr = str(child.type_of_contract) if child.type_of_contract else "-"
            subject = child.subject_contract or "-"
            comment = child.comment or ""

            cost_val = float(child.cost) if child.cost else 0.0

            row_values = [
                (row_num, Alignment(horizontal="center", vertical="center"), data_font, "@"),
                (counteragent_name, Alignment(horizontal="left", vertical="center"), data_font, None),
                (doc_num_text, Alignment(horizontal="left" if is_child else "center", vertical="center"), data_font_child if is_child else data_font, "@"),
                (type_doc, Alignment(horizontal="center", vertical="center"), data_font, None),
                (type_contr, Alignment(horizontal="left", vertical="center"), data_font, None),
                (date_concl, Alignment(horizontal="center", vertical="center"), data_font, "DD.MM.YYYY"),
                (date_close, Alignment(horizontal="center", vertical="center"), data_font, "DD.MM.YYYY"),
                (cost_val if cost_val > 0 else "-", Alignment(horizontal="right", vertical="center"), data_font, "#,##0.00" if cost_val > 0 else None),
                (subject, Alignment(horizontal="left", vertical="center", wrap_text=True), data_font, None),
                (status_text, Alignment(horizontal="center", vertical="center"), status_fnt, None),
                (comment, Alignment(horizontal="left", vertical="center", wrap_text=True), data_font, None),
            ]

            ws.row_dimensions[current_row].height = 22

            # Зебра-заливка (четные строки с легким серым фоном)
            row_fill = (
                PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                if row_num % 2 == 0
                else PatternFill(fill_type=None)
            )

            for col_idx, (val, align, font, num_format) in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = align
                cell.font = font
                cell.border = cell_border
                if row_fill.fill_type:
                    cell.fill = row_fill
                if num_format and isinstance(val, (int, float)):
                    cell.number_format = num_format

            row_num += 1
            current_row += 1

    # 5. Итоговая строка
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    ws.cell(row=current_row, column=1, value="ИТОГО:").alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(
        row=current_row, column=8, value=total_cost if total_cost > 0 else "-"
    ).alignment = Alignment(horizontal="right", vertical="center")
    if total_cost > 0:
        ws.cell(row=current_row, column=8).number_format = "#,##0.00"
    else:
        ws.cell(row=current_row, column=8).number_format = "@"

    ws.merge_cells(start_row=current_row, start_column=9, end_row=current_row, end_column=11)

    for c_idx in range(1, 12):
        c = ws.cell(row=current_row, column=c_idx)
        c.font = total_font
        c.fill = total_fill
        c.border = total_border

    ws.row_dimensions[current_row].height = 24

    # 6. Настройка ширины колонок
    col_widths = {
        1: 8,    # № п/п
        2: 32,   # Контрагент
        3: 24,   # Номер документа
        4: 20,   # Тип документа
        5: 24,   # Тип договора
        6: 16,   # Дата заключения
        7: 16,   # Срок действия
        8: 18,   # Сумма
        9: 40,   # Предмет
        10: 16,  # Статус
        11: 30,  # Примечание
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Закрепляем шапку таблицы (строки выше 11)
    ws.freeze_panes = "A11"

    return wb


def contract_report_view(request):
    """Формирует аналитический веб-отчёт по договорам за указанный период с группировкой по родительским договорам.

    Отображает интерактивную страницу со сводными карточками показателей (KPI),
    фильтрами по датам и быстрым переключением пресетов (текущий месяц, с начала года),
    а также иерархическим списком договоров и подчиненных соглашений.

    Требует обязательной аутентификации и наличия права 'contracts_app.view_contract'
    (или прав суперпользователя). Строки выборки дополнительно фильтруются с учетом
    уровня доступа пользователя (user_access) и признака публикации (allowed_placed).

    Args:
        request (HttpRequest): Объект HTTP-запроса с GET-параметрами 'start' и 'end'.

    Returns:
        HttpResponse: Отрендеренная HTML-страница отчета с контекстом данных либо 403 Forbidden.
    """
    if not request.user.is_authenticated:
        return redirect(reverse("customers_app:login"))

    if not (request.user.has_perm("contracts_app.view_contract") or request.user.is_superuser):
        logger.warning(
            f"Пользователь {request.user} попытался получить несанкционированный доступ к отчёту по договорам."
        )
        return render(request, "library_app/403.html", {"title": "Доступ ограничен"}, status=403)

    raw_start = request.GET.get("start")
    raw_end = request.GET.get("end")

    today = datetime.date.today()
    error_msg = None

    if raw_start or raw_end:
        start_date = parse_date_param(raw_start)
        end_date = parse_date_param(raw_end)
        if not start_date or not end_date:
            error_msg = "Некорректный формат даты. Выберите период из календаря."
            start_date = start_date or datetime.date(today.year, 1, 1)
            end_date = end_date or today
    else:
        # Период по умолчанию: с начала текущего года по текущую дату
        start_date = datetime.date(today.year, 1, 1)
        end_date = today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    # Учет прав доступа пользователя и публикации
    query = Q(date_conclusion__range=(start_date, end_date))
    if not request.user.is_superuser:
        access_val = getattr(request.user, "user_access_id", None)
        if access_val is None and hasattr(request.user, "user_access") and request.user.user_access:
            access_val = (
                request.user.user_access.pk
                if hasattr(request.user.user_access, "pk")
                else request.user.user_access
            )
        if access_val is None:
            access_val = 5

        query &= Q(allowed_placed=True) & Q(access_id__gte=access_val)

    contracts = (
        Contract.objects.filter(query)
        .select_related(
            "parent_category",
            "parent_category__contract_counteragent",
            "parent_category__type_of_document",
            "contract_counteragent",
            "type_of_document",
            "type_of_contract",
            "executor",
        )
        .order_by("contract_counteragent__short_name", "date_conclusion", "pk")
    )

    grouped_contracts = {}
    total_docs = 0
    main_docs = 0
    supp_docs = 0
    counteragents = set()
    total_cost = 0.0

    for contract in contracts:
        total_docs += 1
        if contract.cost:
            total_cost += float(contract.cost)
        if contract.contract_counteragent_id:
            counteragents.add(contract.contract_counteragent_id)

        if contract.parent_category_id is None:
            main_docs += 1
        else:
            supp_docs += 1

        parent = contract.parent_category or contract
        grouped_contracts.setdefault(parent, []).append(contract)

    kpi = {
        "total_docs": total_docs,
        "main_docs": main_docs,
        "supp_docs": supp_docs,
        "counteragents_count": len(counteragents),
        "total_cost": total_cost,
    }

    current_year_start = datetime.date(today.year, 1, 1)
    current_month_start = datetime.date(today.year, today.month, 1)

    context = {
        "title": "Отчёт по договорам",
        "breadcrumbs": [
            {"name": "База договоров", "url": reverse("contracts_app:index")},
            {"name": "Отчёт по договорам", "url": ""},
        ],
        "grouped_contracts": grouped_contracts,
        "start_date": start_date,
        "end_date": end_date,
        "today": today,
        "current_year_start": current_year_start,
        "current_month_start": current_month_start,
        "kpi": kpi,
        "error": error_msg,
    }
    return render(request, "contracts_app/contract_report.html", context)


def export_contracts_excel(request):
    """Экспортирует отчёт по договорам за выбранный период в профессионально оформленный файл Excel (.xlsx).

    Выполняет строгую проверку прав доступа ('contracts_app.view_contract' или superuser),
    фильтрацию записей по уровню доступа (user_access) и публикации (allowed_placed),
    вызывает генератор книги Excel с корпоративным стилем и возвращает файл для скачивания.

    Args:
        request (HttpRequest): Объект HTTP-запроса с GET-параметрами 'start' и 'end'.

    Returns:
        HttpResponse: Поток данных xlsx-файла с заголовком Content-Disposition либо 403 Forbidden.
    """
    if not request.user.is_authenticated:
        return redirect(reverse("customers_app:login"))

    if not (request.user.has_perm("contracts_app.view_contract") or request.user.is_superuser):
        logger.warning(
            f"Пользователь {request.user} попытался несанкционированно выгрузить Excel-отчёт по договорам."
        )
        return render(request, "library_app/403.html", {"title": "Доступ ограничен"}, status=403)

    raw_start = request.GET.get("start")
    raw_end = request.GET.get("end")

    today = datetime.date.today()
    start_date = parse_date_param(raw_start) or datetime.date(today.year, 1, 1)
    end_date = parse_date_param(raw_end) or today

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    query = Q(date_conclusion__range=(start_date, end_date))
    if not request.user.is_superuser:
        access_val = getattr(request.user, "user_access_id", None)
        if access_val is None and hasattr(request.user, "user_access") and request.user.user_access:
            access_val = (
                request.user.user_access.pk
                if hasattr(request.user.user_access, "pk")
                else request.user.user_access
            )
        if access_val is None:
            access_val = 5

        query &= Q(allowed_placed=True) & Q(access_id__gte=access_val)

    contracts = (
        Contract.objects.filter(query)
        .select_related(
            "parent_category",
            "parent_category__contract_counteragent",
            "parent_category__type_of_document",
            "contract_counteragent",
            "type_of_document",
            "type_of_contract",
            "executor",
        )
        .order_by("contract_counteragent__short_name", "date_conclusion", "pk")
    )

    wb = generate_contracts_excel_report(contracts, start_date, end_date, user=request.user)

    filename = f"contracts_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
