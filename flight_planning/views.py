# flight_planning/views.py
import json
from datetime import datetime, timedelta, date
from typing import List, Dict, Any, Optional, Tuple, Union

from django.db.models import Q
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction
from django.contrib.auth.decorators import login_required

from customers_app.models import DataBaseUser
from hrdepartment_app.models import PlaceProductionActivity
from contracts_app.models import Estate, TypeProperty
from .models import (
    PilotAssignment, AircraftMovement, FlightCrew, CrewMember,
    FlightCrewNote, FlightPlanningDocument, PeriodicCheckType, PeriodicCheckRecord,
    EmployeeRequiredCheck, EmployeeStatusType, EmployeeStatusRecord,
    CREW_ROLES, FLIGHT_TYPES, DOCUMENT_STATUSES, CHECK_APPLIES_TO, EMPLOYEE_STATUS_CODES
)
from .forms import (
    AircraftMovementForm, PeriodicCheckRecordForm, PeriodicCheckTypeForm,
    EmployeeStatusRecordForm, EmployeeStatusTypeForm
)
from .permissions import (
    is_flight_planner,
    can_view_flight_reports,
    can_view_flight_planning,
    flight_planner_required,
    flight_reports_required,
    flight_planning_view_required
)
from .selectors import (
    get_pilot_assignments_for_month,
    get_active_aircraft,
    get_latest_aircraft_locations,
    get_aircraft_movement_history,
    get_mpd_aircraft_map,
    get_mpd_aircraft_intervals_map,
    get_all_aircraft_intervals,
    get_crews_for_month,
    get_mpd_crew_map,
    get_available_aircraft_for_mpd,
    get_personnel_utilization_report_data,
    get_aircraft_basing_report_data
)
from .services import (
    get_grouped_pilot_schedule,
    record_aircraft_movement,
    handle_aircraft_movement_crew_fallback,
    validate_crew_composition,
    check_crew_member_conflicts,
    create_or_update_flight_crew_range,
    update_flight_crew,
    delete_flight_crew,
    batch_swap_aircraft,
    get_month_name_ru,
    build_flight_planning_snapshot,
    calculate_flight_planning_diff,
    get_next_document_number,
    create_planning_document,
    approve_planning_document,
    get_latest_approved_document,
    get_pending_document,
    check_snapshot_matches_live,
    calculate_check_end_date,
    get_pilot_periodic_check_status,
    get_month_pilots_check_status_map,
    get_employee_check_assignments,
    save_employee_check_assignments,
    get_batch_employee_check_assignments,
    get_pilot_schedule_from_snapshot,
    get_month_employee_statuses_map,
    get_pilot_employee_statuses,
    get_allowed_staff_queryset,
    get_user_personnel_scope,
    FLIGHT_CREW_JOB_NAMES,
    ENGINEERING_STAFF_JOB_NAMES,
    ALL_STAFF_JOB_NAMES,
    format_short_job
)
from contracts_app.templatetags.custom import FIO_format

# Должности летного состава для планирования
ALLOWED_JOBS = ['командир', 'пилот', 'бортмеханик', 'Командир', 'Бортмеханик', 'инструктор', 'Бортовой']


def get_pilot_allowed_roles(job_name: str) -> list:
    """Возвращает список допустимых ролей в экипаже на основе должности сотрудника.

    Args:
        job_name (str): Наименование должности сотрудника.

    Returns:
        list: Список кодов допустимых ролей ('commander', 'copilot', 'pilot_instructor',
            'flight_engineer', 'flight_engineer_instructor').
    """
    j = (job_name or "").lower()
    roles = []

    # 1. КВС
    if any(k in j for k in ['командир воздушного судна', 'командир летного', 'заместитель командира летного', 'пилот-инструктор', 'пилот-инспектор', 'летный директор']) and not any(x in j for x in ['бортмеханик', 'врач', 'штаб', 'водитель']):
        roles.append('commander')

    # 2. Второй пилот
    if any(k in j for k in ['второй пилот', 'командир воздушного судна', 'командир летного', 'заместитель командира летного', 'пилот-инструктор', 'пилот-инспектор', 'летный директор']) and not any(x in j for x in ['бортмеханик', 'врач', 'штаб', 'водитель']):
        roles.append('copilot')

    # 3. Пилот-инструктор
    if any(k in j for k in ['пилот-инструктор', 'пилот-инспектор', 'командир летного', 'заместитель командира летного', 'летный директор']) and not any(x in j for x in ['бортмеханик', 'врач', 'штаб', 'водитель']):
        roles.append('pilot_instructor')

    # 4. Бортмеханик
    if any(k in j for k in ['бортмеханик', 'бортовой механик', 'старший бортмеханик']):
        roles.append('flight_engineer')

    # 5. Бортмеханик-инструктор
    if any(k in j for k in ['бортмеханик-инструктор', 'старший бортмеханик']):
        roles.append('flight_engineer_instructor')

    return roles if roles else ['commander', 'copilot', 'flight_engineer']


def populate_template_maps_from_snapshot(snapshot_data: dict) -> Tuple[dict, dict]:
    """Преобразует снимок документа в форматы assignment_map и crew_map для шаблона table.html.

    Args:
        snapshot_data (dict): JSON-снимок документа.

    Returns:
        Tuple[dict, dict]: (assignment_map, crew_map).
    """
    grid = snapshot_data.get('grid', {})
    assignment_map = {}
    crew_map = {}

    today = timezone.now().date()
    min_editable_date = today - timedelta(days=1)
    max_editable_date = today + timedelta(days=1)

    for mpd_key, dates_dict in grid.items():
        mpd_id = int(mpd_key) if mpd_key.isdigit() else mpd_key
        if mpd_id not in assignment_map:
            assignment_map[mpd_id] = {}
        if mpd_id not in crew_map:
            crew_map[mpd_id] = {}

        for date_str, crews_list in dates_dict.items():
            if date_str not in assignment_map[mpd_id]:
                assignment_map[mpd_id][date_str] = []
            if date_str not in crew_map[mpd_id]:
                crew_map[mpd_id][date_str] = []

            try:
                cell_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                can_add_note = (min_editable_date <= cell_date <= max_editable_date)
            except Exception:
                can_add_note = False

            for c in crews_list:
                crew_id = c.get('crew_id') or c.get('id')
                flight_type = c.get('flight_type', 'standard')
                flight_type_label = c.get('flight_type_label') or c.get('flight_type_display') or flight_type
                ac_number = c.get('aircraft_number') or c.get('aircraft_reg') or c.get('aircraft_registration') or 'Резерв'
                members = c.get('members', [])
                notes = c.get('notes', [])
                latest_note = notes[0] if notes else None

                members_list = []
                for m in members:
                    p_name = m.get('name') or m.get('pilot_name') or m.get('member_name') or ''
                    r_label = m.get('role_label') or m.get('role_display') or m.get('role') or ''
                    p_id = m.get('pilot_id') or m.get('member_id') or m.get('id')
                    members_list.append({
                        'id': p_id,
                        'member_id': p_id,
                        'pilot_id': p_id,
                        'name': p_name,
                        'pilot_name': p_name,
                        'role': m.get('role', ''),
                        'role_label': r_label,
                        'role_display': r_label,
                        'job': m.get('job', ''),
                    })

                crew_map[mpd_id][date_str].append({
                    'id': crew_id,
                    'aircraft_id': c.get('aircraft_id'),
                    'aircraft_number': ac_number,
                    'aircraft_registration': ac_number,
                    'aircraft_type': c.get('aircraft_type', ''),
                    'flight_type': flight_type,
                    'flight_type_label': flight_type_label,
                    'flight_type_display': flight_type_label,
                    'name': c.get('name', ''),
                    'comment': c.get('comment', ''),
                    'members': members_list,
                    'notes': notes,
                    'notes_count': len(notes),
                    'latest_note': latest_note,
                    'can_add_note': can_add_note
                })

                for m in members_list:
                    assignment_map[mpd_id][date_str].append({
                        'pilot_id': m['pilot_id'],
                        'pilot_name': m['name'],
                        'pilot_job': m.get('job', ''),
                        'is_commander': m['role'] == 'commander',
                        'is_instructor': 'instructor' in (m['role'] or ''),
                        'assignment_id': None,
                        'crew_id': crew_id,
                        'role_in_crew': m['role'],
                        'role_in_crew_label': m['role_label'],
                        'aircraft_number': ac_number,
                        'flight_type': flight_type
                    })

    return assignment_map, crew_map


@login_required
def my_schedule_view(request):
    """Отображает страницу личного графика для текущего пилота или бортмеханика.

    Для летного состава график строится на основании официально утвержденного документа.
    Для диспетчеров планирования отображаются назначения из текущей рабочей базы.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Отрендеренная страница личного графика.
    """
    year = request.GET.get('year', timezone.now().year)
    month = request.GET.get('month', timezone.now().month)

    try:
        year = int(year)
        month = int(month)
    except ValueError:
        year = timezone.now().year
        month = timezone.now().month

    is_planner = is_flight_planner(request.user)
    latest_approved_doc = get_latest_approved_document(year, month)

    if not is_planner and latest_approved_doc:
        # Для летного состава отображаем утвержденный снимок
        grouped_schedule = get_pilot_schedule_from_snapshot(
            latest_approved_doc.snapshot_data,
            request.user.id,
            year,
            month
        )
        is_official = True
        doc_info = latest_approved_doc
    else:
        # Для планировщиков или при отсутствии утвержденного документа
        assignments = get_pilot_assignments_for_month(
            pilot_id=request.user.id,
            year=year,
            month=month
        )
        grouped_schedule = get_grouped_pilot_schedule(list(assignments), year, month)
        is_official = bool(latest_approved_doc)
        doc_info = latest_approved_doc

    # Генерируем даты выбранного месяца для навигации
    first_day = datetime(year, month, 1).date()
    prev_month_date = first_day - timedelta(days=1)
    if month == 12:
        next_month_date = datetime(year + 1, 1, 1).date()
    else:
        next_month_date = datetime(year, month + 1, 1).date()

    context = {
        'grouped_schedule': grouped_schedule,
        'year': year,
        'month': month,
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
        'month_name': first_day.strftime('%B %Y'),
        'is_planner': is_planner,
        'can_view_reports': can_view_flight_reports(request.user),
        'latest_approved_doc': doc_info,
        'is_official': is_official,
        'no_approved_doc': (not is_planner and not latest_approved_doc),
    }

    return render(request, 'flight_planning/my_schedule.html', context)


@login_required
@flight_planning_view_required
def planning_table(request):
    """Главная страница с интерактивной таблицей-шахматкой планирования полетов.

    Планировщикам доступен интерактивный рабочий черновик с возможностью фиксации
    состояния и формирования документа расстановки.
    Летному составу и руководству отображается официально утвержденный снимок плана.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Отрендеренная страница таблицы планирования.
    """
    year = request.GET.get('year', timezone.now().year)
    month = request.GET.get('month', timezone.now().month)
    view_doc_id = request.GET.get('view_doc')

    try:
        year = int(year)
        month = int(month)
    except ValueError:
        year = timezone.now().year
        month = timezone.now().month

    is_planner = is_flight_planner(request.user)
    can_reports = can_view_flight_reports(request.user)
    latest_approved_doc = get_latest_approved_document(year, month)
    pending_doc = get_pending_document(year, month)

    viewing_doc = None
    if view_doc_id:
        try:
            viewing_doc = FlightPlanningDocument.objects.get(pk=int(view_doc_id))
        except (FlightPlanningDocument.DoesNotExist, ValueError):
            pass

    # Получаем все МПД
    mpds = PlaceProductionActivity.objects.filter(in_planning=True).order_by('name')

    # Создаём динамические условия Q для пилотов
    q_conditions = Q()
    for keyword in ALLOWED_JOBS:
        q_conditions |= Q(user_work_profile__job__name__icontains=keyword)
    pilots = DataBaseUser.objects.filter(
        is_active=True,
        user_work_profile__isnull=False
    ).filter(q_conditions).select_related('user_work_profile__job').order_by('last_name', 'first_name').distinct()

    # Генерируем даты выбранного месяца
    first_day = datetime(year, month, 1).date()
    if month == 12:
        last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)

    dates = []
    current = first_day
    while current <= last_day:
        dates.append(current)
        current += timedelta(days=1)

    # Режим отображения данных:
    # 1) Если запрошен просмотр конкретного документа (view_doc)
    # 2) Если пользователь НЕ диспетчер и есть утвержденный документ -> берем снимок утвержденного
    # 3) Иначе -> берем живую рабочую базу данных
    is_snapshot_view = False
    active_display_doc = None

    if viewing_doc and viewing_doc.snapshot_data:
        is_snapshot_view = True
        active_display_doc = viewing_doc
        assignment_map, crew_map = populate_template_maps_from_snapshot(viewing_doc.snapshot_data)
    elif not is_planner and latest_approved_doc and latest_approved_doc.snapshot_data:
        is_snapshot_view = True
        active_display_doc = latest_approved_doc
        assignment_map, crew_map = populate_template_maps_from_snapshot(latest_approved_doc.snapshot_data)
    else:
        # Живая рабочая база
        assignments = PilotAssignment.objects.filter(
            date__year=year,
            date__month=month
        ).select_related('pilot', 'mpd', 'crew', 'crew__aircraft', 'crew__aircraft__type_property')

        assignment_map = {}
        for a in assignments:
            mpd_id = a.mpd_id
            date_str = a.date.isoformat()

            if mpd_id not in assignment_map:
                assignment_map[mpd_id] = {}
            if date_str not in assignment_map[mpd_id]:
                assignment_map[mpd_id][date_str] = []

            job_name = None
            is_commander = False
            is_instructor = False
            try:
                if hasattr(a.pilot, 'user_work_profile') and a.pilot.user_work_profile:
                    if hasattr(a.pilot.user_work_profile, 'job') and a.pilot.user_work_profile.job:
                        job_name = a.pilot.user_work_profile.job.name
                        if job_name:
                            job_lower = job_name.lower()
                            is_commander = 'командир' in job_lower
                            is_instructor = 'инструктор' in job_lower
            except Exception:
                pass

            assignment_map[mpd_id][date_str].append({
                'pilot_id': a.pilot_id,
                'pilot_name': FIO_format(a.pilot.title or a.pilot.username),
                'pilot_job': format_short_job(job_name) if job_name else 'Должность не указана',
                'pilot_job_full': job_name or 'Должность не указана',
                'is_commander': is_commander,
                'is_instructor': is_instructor,
                'assignment_id': a.id,
                'crew_id': a.crew_id,
                'role_in_crew': a.role_in_crew,
                'role_in_crew_label': a.get_role_in_crew_display() if a.role_in_crew else '',
                'aircraft_number': a.crew.aircraft.registration_number if (a.crew and a.crew.aircraft) else ('Резерв' if a.crew else ''),
                'flight_type': a.crew.flight_type if a.crew else ''
            })

        crew_map = get_mpd_crew_map(year, month)

    # Проверяем наличие расхождений между живым черновиком и утвержденным документом
    has_pending_changes = False
    live_diff_list = []
    if is_planner:
        if latest_approved_doc:
            is_match, live_diff_list = check_snapshot_matches_live(year, month, latest_approved_doc)
            has_pending_changes = not is_match
        else:
            # Если утвержденного документа еще нет, но есть созданные экипажи/назначения
            has_pending_changes = bool(FlightCrew.objects.filter(date__year=year, date__month=month).exists())

    # Карты ВС
    mpd_aircraft_map = get_mpd_aircraft_map()
    mpd_aircraft_intervals_map = get_mpd_aircraft_intervals_map(year, month)
    all_aircraft_intervals = get_all_aircraft_intervals()

    pilots_js_list = []
    for p in pilots:
        job_title = p.user_work_profile.job.name if (hasattr(p, 'user_work_profile') and p.user_work_profile and p.user_work_profile.job) else ''
        job_lower = job_title.lower()
        allowed_roles = get_pilot_allowed_roles(job_title)
        suggested_role = 'copilot'
        if 'commander' in allowed_roles:
            suggested_role = 'commander'
        elif 'flight_engineer_instructor' in allowed_roles:
            suggested_role = 'flight_engineer_instructor'
        elif 'flight_engineer' in allowed_roles:
            suggested_role = 'flight_engineer'
        elif 'pilot_instructor' in allowed_roles:
            suggested_role = 'pilot_instructor'

        pilots_js_list.append({
            'id': p.id,
            'name': FIO_format(p.title or p.username),
            'full_name': p.title or p.username,
            'job': format_short_job(job_title),
            'full_job': job_title,
            'is_commander': 'командир' in job_lower,
            'is_instructor': 'инструктор' in job_lower,
            'suggested_role': suggested_role,
            'allowed_roles': allowed_roles
        })

    prev_month_date = first_day - timedelta(days=1)
    next_month_date = last_day + timedelta(days=1)

    all_active_aircraft = get_active_aircraft(first_day)
    all_aircraft_list = [
        {
            'id': ac.id,
            'reg': ac.registration_number,
            'type': ac.type_property.type_property if ac.type_property else ''
        }
        for ac in all_active_aircraft
    ]

    today = timezone.now().date()
    pilots_check_status_map = get_month_pilots_check_status_map([p.id for p in pilots], year, month)
    employee_statuses_map = get_month_employee_statuses_map([p.id for p in pilots], year, month)

    context = {
        'mpds': mpds,
        'pilots': pilots,
        'dates': dates,
        'today': today,
        'is_current_month': (year == today.year and month == today.month),
        'year': year,
        'month': month,
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
        'month_name': first_day.strftime('%B %Y'),
        'assignment_map': assignment_map,
        'crew_map': crew_map,
        'crew_map_json': json.dumps(crew_map),
        'mpd_aircraft_json': json.dumps(mpd_aircraft_map),
        'mpd_aircraft_intervals_json': json.dumps(mpd_aircraft_intervals_map),
        'all_aircraft_intervals_json': json.dumps(all_aircraft_intervals),
        'pilots_js_json': json.dumps(pilots_js_list),
        'all_aircraft_json': json.dumps(all_aircraft_list),
        'pilots_check_status_map': pilots_check_status_map,
        'pilots_check_status_json': json.dumps(pilots_check_status_map),
        'employee_statuses_map': employee_statuses_map,
        'employee_statuses_json': json.dumps(employee_statuses_map),
        'is_planner': is_planner,
        'can_view_reports': can_reports,
        'latest_approved_doc': latest_approved_doc,
        'pending_doc': pending_doc,
        'has_pending_changes': has_pending_changes,
        'live_diff_count': len(live_diff_list),
        'live_diff_list': live_diff_list[:8],
        'is_snapshot_view': is_snapshot_view,
        'active_display_doc': active_display_doc,
        'no_approved_doc': (not is_planner and not latest_approved_doc and not viewing_doc),
    }

    return render(request, 'flight_planning/table.html', context)


# ========================================================
# ПРЕДСТАВЛЕНИЯ ДОКУМЕНТОВ РАССТАНОВКИ ЭКИПАЖЕЙ (ВЕРСИИ)
# ========================================================

@login_required
@flight_planning_view_required
def document_list_view(request):
    """Отображает журнал документов расстановки экипажей по месяцам и редакциям.

    Args:
        request (HttpRequest): Объект HTTP-запроса с фильтрами year, month, status.

    Returns:
        HttpResponse: Отрендеренная страница журнала документов.
    """
    year = request.GET.get('year')
    month = request.GET.get('month')
    status = request.GET.get('status')

    qs = FlightPlanningDocument.objects.all().select_related('author', 'approved_by', 'previous_document')

    if year:
        try:
            qs = qs.filter(year=int(year))
        except ValueError:
            pass

    if month:
        try:
            qs = qs.filter(month=int(month))
        except ValueError:
            pass

    if status:
        qs = qs.filter(status=status)

    years_list = sorted(list(set(FlightPlanningDocument.objects.values_list('year', flat=True))), reverse=True)
    if not years_list:
        years_list = [timezone.now().year]

    context = {
        'title': 'Журнал документов расстановки экипажей',
        'documents': qs,
        'selected_year': int(year) if year and year.isdigit() else None,
        'selected_month': int(month) if month and month.isdigit() else None,
        'selected_status': status,
        'years_list': years_list,
        'months_list': [(i, get_month_name_ru(i)) for i in range(1, 13)],
        'status_choices': DOCUMENT_STATUSES,
        'is_planner': is_flight_planner(request.user),
        'can_approve': can_view_flight_reports(request.user) or request.user.is_superuser,
    }
    return render(request, 'flight_planning/document_list.html', context)


@login_required
@flight_planning_view_required
def document_detail_view(request, pk: int):
    """Отображает детальную карточку документа расстановки с сеткой и реестром изменений.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Первичный ключ документа.

    Returns:
        HttpResponse: Отрендеренная страница документа.
    """
    document = get_object_or_404(
        FlightPlanningDocument.objects.select_related('author', 'approved_by', 'previous_document'),
        pk=pk
    )

    context = {
        'title': f"Документ № {document.number}",
        'document': document,
        'snapshot': document.snapshot_data,
        'diff_list': document.diff_data,
        'is_planner': is_flight_planner(request.user),
        'can_approve': (can_view_flight_reports(request.user) or request.user.is_superuser) and document.is_pending,
    }
    return render(request, 'flight_planning/document_detail.html', context)


@login_required
@flight_planning_view_required
def document_print_view(request, pk: int):
    """Отображает официальную печатную форму документа расстановки экипажей (формат А4).

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Первичный ключ документа.

    Returns:
        HttpResponse: Отрендеренная печатная страница.
    """
    document = get_object_or_404(
        FlightPlanningDocument.objects.select_related(
            'author', 'author__user_work_profile__job',
            'approved_by', 'approved_by__user_work_profile__job',
            'previous_document'
        ),
        pk=pk
    )

    author_job = ""
    if hasattr(document.author, 'user_work_profile') and document.author and document.author.user_work_profile and document.author.user_work_profile.job:
        author_job = document.author.user_work_profile.job.name

    approver_job = ""
    if hasattr(document.approved_by, 'user_work_profile') and document.approved_by and document.approved_by.user_work_profile and document.approved_by.user_work_profile.job:
        approver_job = document.approved_by.user_work_profile.job.name

    context = {
        'title': f"Печатная форма — Документ № {document.number}",
        'document': document,
        'snapshot': document.snapshot_data,
        'diff_list': document.diff_data,
        'author_job': author_job or "Диспетчер по планированию полетов",
        'approver_job': approver_job or "Летный директор / Командир летного отряда",
        'generation_date': timezone.now().strftime('%d.%m.%Y %H:%M'),
    }
    return render(request, 'flight_planning/document_print.html', context)


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def document_create_view(request):
    """Создает новый документ расстановки экипажей на основе текущего рабочего плана.

    Args:
        request (HttpRequest): POST-запрос с параметрами year, month, reason.

    Returns:
        JsonResponse: Результат создания документа с ID и номером.
    """
    year = request.POST.get('year')
    month = request.POST.get('month')
    reason = request.POST.get('reason', '').strip()

    if not year or not month:
        return JsonResponse({'error': 'Параметры года и месяца обязательны.'}, status=400)

    try:
        year = int(year)
        month = int(month)
    except ValueError:
        return JsonResponse({'error': 'Некорректный формат года или месяца.'}, status=400)

    # Проверяем, есть ли уже утвержденный документ
    latest_approved = get_latest_approved_document(year, month)
    if latest_approved and not reason:
        return JsonResponse({
            'error': 'Для внесения изменений (повторной редакции плана) обязательно укажите причину / основание!'
        }, status=400)

    # Проверяем, нет ли уже документа в статусе pending
    pending_doc = get_pending_document(year, month)
    if pending_doc:
        return JsonResponse({
            'error': f'Документ № {pending_doc.number} уже сформирован и ожидает утверждения руководством. Дождитесь его утверждения или отклонения.'
        }, status=400)

    document = create_planning_document(
        year=year,
        month=month,
        author=request.user,
        reason=reason
    )

    return JsonResponse({
        'status': 'success',
        'document_id': document.id,
        'number': document.number,
        'message': f"Документ № {document.number} успешно сформирован и отправлен на утверждение руководству!",
        'redirect_url': reverse('flight_planning:document_detail', args=[document.id])
    })


@login_required
@flight_reports_required
@require_http_methods(["POST"])
def document_approve_view(request, pk: int):
    """Утверждает документ расстановки экипажей (доступно руководству).

    Args:
        request (HttpRequest): POST-запрос.
        pk (int): Первичный ключ документа.

    Returns:
        HttpResponseRedirect | JsonResponse: Перенаправление на карточку документа или JSON.
    """
    document = get_object_or_404(FlightPlanningDocument, pk=pk)

    if document.status != 'pending':
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Утвердить можно только документ в статусе «На утверждении».'}, status=400)
        messages.error(request, 'Утвердить можно только документ в статусе «На утверждении».')
        return redirect('flight_planning:document_detail', pk=document.id)

    approve_planning_document(document, approver=request.user)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'message': f"Документ № {document.number} успешно утвержден и вступил в силу!",
            'redirect_url': reverse('flight_planning:document_detail', args=[document.id])
        })

    messages.success(request, f"Документ № {document.number} успешно утвержден и вступил в силу!")
    return redirect('flight_planning:document_detail', pk=document.id)



@login_required
@require_http_methods(["GET"])
def get_assignments_api(request):
    """Получить назначения за месяц в формате JSON.

    Args:
        request (HttpRequest): Объект HTTP-запроса с параметрами year и month.

    Returns:
        JsonResponse: Список назначений в формате JSON.
    """
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not month:
        return JsonResponse({'error': 'year and month required'}, status=400)

    assignments = PilotAssignment.objects.filter(
        date__year=year,
        date__month=month
    ).values('id', 'pilot_id', 'mpd_id', 'date')

    return JsonResponse({'assignments': list(assignments)}, safe=False)


@login_required
@require_http_methods(["GET"])
def get_my_assignments_api(request):
    """Возвращает назначения текущего пользователя за указанный месяц в формате JSON.

    Args:
        request (HttpRequest): Объект HttpRequest с параметрами year и month.

    Returns:
        JsonResponse: JSON-ответ со списком назначений текущего пользователя.
    """
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not month:
        now = timezone.now()
        year = now.year
        month = now.month

    try:
        year = int(year)
        month = int(month)
    except ValueError:
        return JsonResponse({'error': 'Invalid year or month'}, status=400)

    assignments = get_pilot_assignments_for_month(
        pilot_id=request.user.id,
        year=year,
        month=month
    )

    data = [
        {
            'id': a.id,
            'date': a.date.isoformat(),
            'mpd_id': a.mpd_id,
            'mpd_name': a.mpd.name
        }
        for a in assignments
    ]

    return JsonResponse({'assignments': data})


@login_required
@flight_planner_required
@csrf_exempt
@require_http_methods(["POST"])
def assign_pilot_api(request):
    """Назначает пилота на диапазон дат для МПД с проверкой конфликтов.

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом (pilot_id, mpd_id, start_date, end_date).

    Returns:
        JsonResponse: Результат назначения или список конфликтов при их наличии.
    """
    try:
        data = json.loads(request.body)
        pilot_id = data.get('pilot_id')
        mpd_id = data.get('mpd_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        # Валидация
        if not all([pilot_id, mpd_id, start_date, end_date]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        pilot = get_object_or_404(DataBaseUser, id=pilot_id, is_active=True)
        mpd = get_object_or_404(PlaceProductionActivity, id=mpd_id)

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        if start > end:
            return JsonResponse({'error': 'Start date must be before end date'}, status=400)

        # Проверяем конфликты
        conflicts = []
        current = start
        while current <= end:
            existing = PilotAssignment.objects.filter(
                pilot=pilot,
                date=current
            ).select_related('mpd').first()

            if existing and existing.mpd_id != mpd_id:
                conflicts.append({
                    'date': current.isoformat(),
                    'old_mpd_id': existing.mpd_id,
                    'old_mpd_name': existing.mpd.name,
                    'assignment_id': existing.id
                })
            current += timedelta(days=1)

        # Получаем информацию о должности пилота
        job_name = None
        is_commander = False
        is_instructor = False
        try:
            if hasattr(pilot, 'user_work_profile') and pilot.user_work_profile:
                if hasattr(pilot.user_work_profile, 'job') and pilot.user_work_profile.job:
                    job_name = pilot.user_work_profile.job.name
                    if job_name:
                        job_lower = job_name.lower()
                        is_commander = 'командир' in job_lower
                        is_instructor = 'пилот-инструктор' in job_lower
        except Exception:
            pass

        # Если есть конфликты, возвращаем их для подтверждения
        if conflicts:
            return JsonResponse({
                'status': 'conflict',
                'conflicts': conflicts,
                'pilot_id': pilot_id,
                'mpd_id': mpd_id,
                'start_date': start_date,
                'end_date': end_date
            }, status=409)

        # Нет конфликтов — создаём назначения
        assignments_created = []
        with transaction.atomic():
            current = start
            while current <= end:
                assignment, created = PilotAssignment.objects.get_or_create(
                    pilot=pilot,
                    date=current,
                    defaults={'mpd': mpd, 'created_by': request.user}
                )
                if not created and assignment.mpd_id != mpd_id:
                    assignment.mpd = mpd
                    assignment.save()
                assignments_created.append({
                    'date': current.isoformat(),
                    'assignment_id': assignment.id
                })
                current += timedelta(days=1)

        return JsonResponse({
            'status': 'success',
            'assignments': assignments_created,
            'mpd_id': mpd_id,
            'pilot_id': pilot_id,
            'pilot_name': FIO_format(pilot.title or pilot.username),
            'pilot_job': format_short_job(job_name) if job_name else 'Должность не указана',
            'pilot_job_full': job_name or 'Должность не указана',
            'is_commander': is_commander,
            'is_instructor': is_instructor
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_planner_required
@csrf_exempt
@require_http_methods(["POST"])
def resolve_conflict_api(request):
    """Разрешает конфликт назначений — удаляет старые конфликтующие назначения и создает новые.

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом (conflicts, pilot_id, mpd_id, start_date, end_date).

    Returns:
        JsonResponse: Результат создания обновленных назначений.
    """
    try:
        data = json.loads(request.body)
        conflicts = data.get('conflicts', [])
        pilot_id = data.get('pilot_id')
        mpd_id = data.get('mpd_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')

        if not all([pilot_id, mpd_id, start_date, end_date]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        pilot = get_object_or_404(DataBaseUser, id=pilot_id)
        mpd = get_object_or_404(PlaceProductionActivity, id=mpd_id)
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        with transaction.atomic():
            # Удаляем конфликтующие назначения
            conflict_ids = [c['assignment_id'] for c in conflicts if c.get('assignment_id')]
            if conflict_ids:
                PilotAssignment.objects.filter(id__in=conflict_ids).delete()

            # Создаём новые назначения
            assignments_created = []
            current = start
            while current <= end:
                # Проверяем, не создали ли уже (на случай частичного конфликта)
                existing = PilotAssignment.objects.filter(pilot=pilot, date=current).first()
                if existing:
                    existing.mpd = mpd
                    existing.save()
                    assignments_created.append({
                        'date': current.isoformat(),
                        'assignment_id': existing.id
                    })
                else:
                    assignment = PilotAssignment.objects.create(
                        pilot=pilot,
                        mpd=mpd,
                        date=current,
                        created_by=request.user
                    )
                    assignments_created.append({
                        'date': current.isoformat(),
                        'assignment_id': assignment.id
                    })
                current += timedelta(days=1)

        return JsonResponse({
            'status': 'success',
            'assignments': assignments_created
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_planner_required
@csrf_exempt
@require_http_methods(["POST"])
def remove_assignments_api(request):
    """Удаляет назначения пилотов по списку идентификаторов.

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом (assignment_ids).

    Returns:
        JsonResponse: Количество удаленных записей.
    """
    try:
        data = json.loads(request.body)
        assignment_ids = data.get('assignment_ids', [])

        if not assignment_ids:
            return JsonResponse({'error': 'No assignment IDs provided'}, status=400)

        deleted_count, _ = PilotAssignment.objects.filter(
            id__in=assignment_ids
        ).delete()

        return JsonResponse({
            'status': 'success',
            'deleted': deleted_count
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_pilot_job_info(request):
    """
    Получить должность пилота и статус командира/инструктора
    """
    pilot_id = request.GET.get('pilot_id')
    if not pilot_id:
        return JsonResponse({'error': 'pilot_id required'}, status=400)

    try:
        pilot = get_object_or_404(DataBaseUser, id=pilot_id)
        job_name = None
        is_commander = False
        is_instructor = False

        if hasattr(pilot, 'user_work_profile') and pilot.user_work_profile:
            if pilot.user_work_profile.job:
                job_name = pilot.user_work_profile.job.name
                if job_name:
                    job_lower = job_name.lower()
                    is_commander = 'командир' in job_lower
                    is_instructor = 'пилот-инструктор' in job_lower

        return JsonResponse({
            'job_name': format_short_job(job_name) if job_name else 'Должность не указана',
            'full_job_name': job_name or 'Должность не указана',
            'is_commander': is_commander,
            'is_instructor': is_instructor
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_planning_view_required
def aircraft_movement_list_view(request):
    """Отображает страницу журнала перемещений воздушных судов (ВС) по МПД.

    Включает сводную информацию о текущей дислокации бортов и фильтрацию истории.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Отрендеренная страница журнала перемещений или JsonResponse для Ajax.
    """
    aircraft_id = request.GET.get('aircraft')
    mpd_id = request.GET.get('mpd')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    movements_qs = AircraftMovement.objects.select_related(
        'aircraft', 'aircraft__type_property', 'mpd', 'created_by'
    )

    if aircraft_id:
        movements_qs = movements_qs.filter(aircraft_id=aircraft_id)
    if mpd_id:
        movements_qs = movements_qs.filter(mpd_id=mpd_id)
    if date_from:
        try:
            d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            movements_qs = movements_qs.filter(date__gte=d_from)
        except ValueError:
            pass
    if date_to:
        try:
            d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            movements_qs = movements_qs.filter(date__lte=d_to)
        except ValueError:
            pass

    movements = movements_qs.order_by('-date', '-created_at')

    # Текущее распределение активных бортов по МПД
    mpd_aircraft_map = get_mpd_aircraft_map()
    all_mpds_in_planning = PlaceProductionActivity.objects.filter(in_planning=True).order_by('name')

    # Формируем структуру сводки по всем МПД в планировании
    mpd_summary = []
    for mpd_obj in all_mpds_in_planning:
        mpd_summary.append({
            'mpd': mpd_obj,
            'aircrafts': mpd_aircraft_map.get(mpd_obj.id, [])
        })

    # Справочники для фильтров формы
    aircrafts = Estate.objects.select_related('type_property').order_by('registration_number')
    mpds = all_mpds_in_planning

    # Если запрос через Ajax/DataTable
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and request.GET.get('format') == 'json':
        data = [
            {
                'id': m.id,
                'aircraft_number': m.aircraft.registration_number,
                'aircraft_type': m.aircraft.type_property.type_property if m.aircraft.type_property else '',
                'is_decommissioned': m.aircraft.is_decommissioned,
                'mpd_name': m.mpd.name,
                'mpd_id': m.mpd_id,
                'date': m.date.strftime('%d.%m.%Y'),
                'raw_date': m.date.isoformat(),
                'comment': m.comment,
                'created_by': m.created_by.title or m.created_by.username if m.created_by else 'Система',
                'created_at': m.created_at.strftime('%d.%m.%Y %H:%M'),
            }
            for m in movements
        ]
        return JsonResponse({'data': data})

    is_planner = is_flight_planner(request.user)
    can_reports = can_view_flight_reports(request.user)

    context = {
        'title': 'Журнал перемещения воздушных судов по МПД',
        'movements': movements,
        'mpd_summary': mpd_summary,
        'aircrafts': aircrafts,
        'mpds': mpds,
        'selected_aircraft_id': int(aircraft_id) if aircraft_id and aircraft_id.isdigit() else None,
        'selected_mpd_id': int(mpd_id) if mpd_id and mpd_id.isdigit() else None,
        'selected_date_from': date_from or '',
        'selected_date_to': date_to or '',
        'is_planner': is_planner,
        'can_view_reports': can_reports,
    }
    return render(request, 'flight_planning/aircraft_movement_list.html', context)


@login_required
@flight_planner_required
def aircraft_movement_create_view(request):
    """Создание новой записи в журнале перемещения ВС.

    Автоматически переводит будущие экипажи на старых МПД с этим бортом в Резерв.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Перенаправление на список или страница формы.
    """
    if request.method == 'POST':
        form = AircraftMovementForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            movement.created_by = request.user
            movement.save()

            # Сброс экипажей на старых МПД в статус Резерв
            fallback_count, affected_mpds = handle_aircraft_movement_crew_fallback(
                aircraft_id=movement.aircraft_id,
                new_mpd_id=movement.mpd_id,
                movement_date=movement.date
            )

            msg = f"Перемещение борта {movement.aircraft.registration_number} на МПД '{movement.mpd.name}' успешно зарегистрировано."
            if fallback_count > 0:
                mpds_str = ", ".join(affected_mpds)
                msg += f" {fallback_count} запланированных экипажей на МПД [{mpds_str}] с {movement.date.strftime('%d.%m.%Y')} переведены в Резерв."

            messages.success(request, msg)
            return redirect('flight_planning:aircraft_movement_list')
    else:
        initial = {}
        if request.GET.get('aircraft_id'):
            initial['aircraft'] = request.GET.get('aircraft_id')
        if request.GET.get('mpd_id'):
            initial['mpd'] = request.GET.get('mpd_id')
        form = AircraftMovementForm(initial=initial)

    context = {
        'title': 'Добавить перемещение воздушного судна',
        'form': form,
        'is_edit': False,
        'is_planner': True,
        'can_view_reports': can_view_flight_reports(request.user),
    }
    return render(request, 'flight_planning/aircraft_movement_form.html', context)


@login_required
@flight_planner_required
def aircraft_movement_update_view(request, pk):
    """Редактирование существующей записи журнала перемещения ВС.

    Автоматически переводит будущие экипажи на старых МПД с этим бортом в Резерв.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор записи перемещения.

    Returns:
        HttpResponse: Перенаправление на список или страница формы.
    """
    movement = get_object_or_404(AircraftMovement, pk=pk)

    if request.method == 'POST':
        form = AircraftMovementForm(request.POST, instance=movement)
        if form.is_valid():
            movement = form.save()

            # Сброс экипажей на старых МПД в статус Резерв
            fallback_count, affected_mpds = handle_aircraft_movement_crew_fallback(
                aircraft_id=movement.aircraft_id,
                new_mpd_id=movement.mpd_id,
                movement_date=movement.date
            )

            msg = f"Запись о перемещении борта {movement.aircraft.registration_number} успешно обновлена."
            if fallback_count > 0:
                mpds_str = ", ".join(affected_mpds)
                msg += f" {fallback_count} запланированных экипажей на МПД [{mpds_str}] с {movement.date.strftime('%d.%m.%Y')} переведены в Резерв."

            messages.success(request, msg)
            return redirect('flight_planning:aircraft_movement_list')
    else:
        form = AircraftMovementForm(instance=movement)

    context = {
        'title': f'Редактирование перемещения: {movement.aircraft.registration_number}',
        'form': form,
        'movement': movement,
        'is_edit': True,
        'is_planner': True,
        'can_view_reports': can_view_flight_reports(request.user),
    }
    return render(request, 'flight_planning/aircraft_movement_form.html', context)


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def aircraft_movement_delete_view(request, pk):
    """Удаление записи о перемещении ВС. Поддерживает как стандартный POST, так и AJAX запрос.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор записи перемещения.

    Returns:
        HttpResponse: JSON-ответ или редирект на список перемещений.
    """
    movement = get_object_or_404(AircraftMovement, pk=pk)
    aircraft_reg = movement.aircraft.registration_number
    mpd_name = movement.mpd.name
    movement.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success', 'message': 'Запись успешно удалена'})

    messages.success(request, f"Запись о перемещении борта {aircraft_reg} на {mpd_name} удалена.")
    return redirect('flight_planning:aircraft_movement_list')


@login_required
@require_http_methods(["GET"])
def get_aircraft_locations_api(request):
    """
    API эндпоинт для получения текущего местоположения всех активных ВС по МПД.
    """
    date_str = request.GET.get('date')
    target_date = None
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Invalid date format, expected YYYY-MM-DD'}, status=400)

    mpd_map = get_mpd_aircraft_map(target_date)
    return JsonResponse({'status': 'success', 'locations': mpd_map})


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def save_crew_api(request):
    """API эндпоинт для создания или обновления летного экипажа (на день или диапазон дат).

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом экипажа.

    Returns:
        JsonResponse: Результат сохранения экипажа или список ошибок/конфликтов.
    """
    try:
        data = json.loads(request.body)
        crew_id = data.get('crew_id')
        mpd_id = data.get('mpd_id')
        aircraft_id = data.get('aircraft_id')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date') or start_date_str
        flight_type = data.get('flight_type', 'standard')
        members = data.get('members', [])
        name = data.get('name', '')
        comment = data.get('comment', '')

        if not mpd_id or not start_date_str:
            return JsonResponse({'error': 'Не заполнены обязательные параметры (МПД, даты).'}, status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Некорректный формат даты (ожидается YYYY-MM-DD).'}, status=400)

        if start_date > end_date:
            return JsonResponse({'error': 'Начальная дата не может быть позже конечной.'}, status=400)

        if aircraft_id:
            try:
                aircraft_id = int(aircraft_id)
            except (ValueError, TypeError):
                aircraft_id = None

        force_override = data.get('force_override', False)

        if crew_id and start_date == end_date:
            # Редактирование конкретного экипажа на выбранную дату
            result = update_flight_crew(
                crew_id=int(crew_id),
                mpd_id=int(mpd_id),
                aircraft_id=aircraft_id,
                target_date=start_date,
                flight_type=flight_type,
                members=members,
                comment=comment,
                crew_name=name,
                force_override=force_override
            )
        else:
            result = create_or_update_flight_crew_range(
                mpd_id=int(mpd_id),
                aircraft_id=aircraft_id,
                start_date=start_date,
                end_date=end_date,
                flight_type=flight_type,
                members=members,
                created_by=request.user,
                comment=comment,
                crew_name=name,
                force_override=force_override
            )

        if result.get('status') == 'error':
            return JsonResponse({'error': ' '.join(result.get('errors', [])), 'errors': result.get('errors')}, status=400)
        elif result.get('status') == 'conflict':
            return JsonResponse({
                'status': 'conflict',
                'conflict_type': result.get('conflict_type', 'unknown'),
                'conflicts': result.get('conflicts', []),
                'can_override': result.get('can_override', False),
                'errors': result.get('errors', []),
                'error': ' '.join(result.get('errors', []))
            }, status=409)

        return JsonResponse({
            'status': 'success',
            'message': f"Экипаж успешно сохранен ({start_date_str} — {end_date_str})",
            'created_crews_count': result.get('created_crews_count', 1),
            'aircraft_name': result.get('aircraft_name', 'Резерв')
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_crew_detail_api(request, crew_id: int):
    """API получения детальной информации по конкретному экипажу для редактирования.

    Args:
        request (HttpRequest): HTTP-запрос.
        crew_id (int): Идентификатор экипажа.

    Returns:
        JsonResponse: Данные экипажа и его состава.
    """
    try:
        crew = FlightCrew.objects.select_related('aircraft', 'mpd').prefetch_related('members__member').get(id=crew_id)
        members_data = []
        for m in crew.members.all():
            job_name = ''
            try:
                if hasattr(m.member, 'user_work_profile') and m.member.user_work_profile and m.member.user_work_profile.job:
                    job_name = m.member.user_work_profile.job.name
            except Exception:
                pass

            members_data.append({
                'member_id': m.member_id,
                'name': m.member.title or m.member.username,
                'role': m.role,
                'role_label': m.get_role_display(),
                'job': job_name,
                'allowed_roles': get_pilot_allowed_roles(job_name)
            })

        return JsonResponse({
            'status': 'success',
            'crew': {
                'id': crew.id,
                'mpd_id': crew.mpd_id,
                'mpd_name': crew.mpd.name,
                'aircraft_id': crew.aircraft_id,
                'aircraft_number': crew.aircraft.registration_number if crew.aircraft else 'Резерв',
                'date': crew.date.isoformat(),
                'flight_type': crew.flight_type,
                'flight_type_label': crew.get_flight_type_display(),
                'name': crew.name,
                'comment': crew.comment,
                'members': members_data
            }
        })
    except FlightCrew.DoesNotExist:
        return JsonResponse({'error': 'Экипаж не найден.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def validate_crew_api(request):
    """API валидации состава экипажа на лету по авиационным правилам и проверка статусов занятости.

    Args:
        request (HttpRequest): HTTP-запрос с типом полета, списком членов и опциональными датами.

    Returns:
        JsonResponse: Результат валидации, список замечаний по составу и предупреждений о занятости.
    """
    try:
        data = json.loads(request.body)
        flight_type = data.get('flight_type', 'standard')
        members = data.get('members', [])
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')

        is_valid, errors = validate_crew_composition(flight_type, members)
        status_warnings = []

        if start_date_str and end_date_str and members:
            try:
                s_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                e_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                member_ids = [m.get('member_id') or m.get('pilot_id') for m in members if (m.get('member_id') or m.get('pilot_id'))]
                if member_ids:
                    records = EmployeeStatusRecord.objects.filter(
                        employee_id__in=member_ids,
                        status_type__is_blocking=True,
                        start_date__lte=e_date,
                        end_date__gte=s_date
                    ).select_related('employee', 'status_type')
                    for r in records:
                        p_name = r.employee.title or r.employee.username
                        doc_part = f" (док. №{r.document_number})" if r.document_number else ""
                        status_warnings.append(
                            f"{p_name}: статус «{r.status_type.name}» с {r.start_date.strftime('%d.%m.%Y')} по {r.end_date.strftime('%d.%m.%Y')}{doc_part}"
                        )
            except Exception:
                pass

        return JsonResponse({
            'status': 'success',
            'is_valid': is_valid,
            'errors': errors,
            'status_warnings': status_warnings
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def delete_crew_api(request):
    """API удаления экипажа по идентификатору.

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом (crew_id).

    Returns:
        JsonResponse: Результат операции удаления.
    """
    try:
        data = json.loads(request.body)
        crew_id = data.get('crew_id')
        if not crew_id:
            return JsonResponse({'error': 'Параметр crew_id обязателен.'}, status=400)

        success = delete_flight_crew(int(crew_id))
        if success:
            return JsonResponse({'status': 'success', 'message': 'Экипаж успешно удален.'})
        return JsonResponse({'error': 'Экипаж не найден.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_day_crew_info_api(request):
    """Возвращает полную информацию по МПД на выбранную дату (ВС, экипажи, назначения).

    Args:
        request (HttpRequest): HTTP-запрос с параметрами mpd_id и date.

    Returns:
        JsonResponse: Сводная информация по МПД на дату.
    """
    mpd_id = request.GET.get('mpd_id')
    date_str = request.GET.get('date')

    if not mpd_id or not date_str:
        return JsonResponse({'error': 'Параметры mpd_id и date обязательны.'}, status=400)

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        mpd_id = int(mpd_id)
    except ValueError:
        return JsonResponse({'error': 'Некорректные параметры.'}, status=400)

    # 1. Воздушные суда на МПД
    available_aircraft = get_available_aircraft_for_mpd(mpd_id, target_date)
    aircraft_list = [
        {
            'id': ac.id,
            'registration_number': ac.registration_number,
            'type_property': ac.type_property.type_property if ac.type_property else 'ВС',
            'is_decommissioned': ac.is_decommissioned
        }
        for ac in available_aircraft
    ]

    # 2. Существующие экипажи на МПД на эту дату
    crews = FlightCrew.objects.filter(
        mpd_id=mpd_id,
        date=target_date
    ).select_related('aircraft', 'aircraft__type_property').prefetch_related('members__member')

    crews_data = []
    for c in crews:
        members_data = [
            {
                'member_id': m.member_id,
                'name': m.member.title or m.member.username,
                'role': m.role,
                'role_label': m.get_role_display()
            }
            for m in c.members.all()
        ]
        crews_data.append({
            'id': c.id,
            'aircraft_id': c.aircraft_id,
            'aircraft_number': c.aircraft.registration_number if c.aircraft else 'Резерв',
            'aircraft_type': c.aircraft.type_property.type_property if (c.aircraft and c.aircraft.type_property) else '',
            'flight_type': c.flight_type,
            'flight_type_label': c.get_flight_type_display(),
            'name': c.name,
            'comment': c.comment,
            'members': members_data
        })

    # 3. Индивидуальные назначения вне экипажей
    standalone_assignments = PilotAssignment.objects.filter(
        mpd_id=mpd_id,
        date=target_date,
        crew__isnull=True
    ).select_related('pilot')

    standalone_data = [
        {
            'assignment_id': a.id,
            'pilot_id': a.pilot_id,
            'pilot_name': a.pilot.title or a.pilot.username,
        }
        for a in standalone_assignments
    ]

    return JsonResponse({
        'status': 'success',
        'mpd_id': mpd_id,
        'date': date_str,
        'aircraft_count': len(aircraft_list),
        'aircraft_list': aircraft_list,
        'crews_count': len(crews_data),
        'crews': crews_data,
        'standalone_assignments': standalone_data
    })


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def add_member_to_crew_api(request):
    """API добавления сотрудника в существующий экипаж с проверкой занятости на других МПД/экипажах.

    Args:
        request (HttpRequest): HTTP-запрос с JSON телом (crew_id, pilot_id, role, force_override).

    Returns:
        JsonResponse: Результат добавления в экипаж или список конфликтов.
    """
    try:
        data = json.loads(request.body)
        crew_id = data.get('crew_id')
        pilot_id = data.get('pilot_id')
        role = data.get('role')
        force_override = data.get('force_override', False)

        if not crew_id or not pilot_id or not role:
            return JsonResponse({'error': 'Не заполнены обязательные параметры (crew_id, pilot_id, role).'}, status=400)

        crew = get_object_or_404(FlightCrew, id=crew_id)
        pilot = get_object_or_404(DataBaseUser, id=pilot_id)

        # Проверка конфликтов занятости
        conflicts = check_crew_member_conflicts(
            mpd_id=crew.mpd_id,
            start_date=crew.date,
            end_date=crew.date,
            members=[{'member_id': int(pilot_id), 'role': role}],
            current_crew_id=crew.id
        )

        if conflicts and not force_override:
            return JsonResponse({
                'status': 'conflict',
                'conflict_type': 'members',
                'conflicts': conflicts,
                'can_override': True,
                'errors': [c['description'] for c in conflicts],
                'error': ' '.join([c['description'] for c in conflicts])
            }, status=409)

        with transaction.atomic():
            # Если сотрудник состоял в другом экипаже на эту дату, удаляем
            old_crews = FlightCrew.objects.filter(
                date=crew.date,
                members__member=pilot
            ).exclude(id=crew.id)
            for oc in old_crews:
                CrewMember.objects.filter(crew=oc, member=pilot).delete()

            CrewMember.objects.update_or_create(
                crew=crew,
                member=pilot,
                defaults={'role': role}
            )
            assignment, created = PilotAssignment.objects.get_or_create(
                pilot=pilot,
                date=crew.date,
                defaults={
                    'mpd': crew.mpd,
                    'crew': crew,
                    'role_in_crew': role,
                    'created_by': request.user
                }
            )
            if not created:
                assignment.mpd = crew.mpd
                assignment.crew = crew
                assignment.role_in_crew = role
                assignment.save()

        return JsonResponse({
            'status': 'success',
            'message': f"Сотрудник {pilot.title or pilot.username} успешно добавлен в экипаж."
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@login_required
@require_http_methods(["GET"])
def get_crew_notes_api(request, crew_id: int):
    """API получения списка пометок и сообщений к экипажу/полету.

    Args:
        request (HttpRequest): HTTP-запрос.
        crew_id (int): Идентификатор экипажа.

    Returns:
        JsonResponse: Список пометок и флаги прав текущего пользователя.
    """
    try:
        crew = FlightCrew.objects.select_related('aircraft', 'mpd').prefetch_related('notes__author', 'members__member').get(id=crew_id)
        today = timezone.now().date()
        min_editable_date = today - timedelta(days=1)
        max_editable_date = today + timedelta(days=1)
        is_date_allowed = (min_editable_date <= crew.date <= max_editable_date)

        # Проверка прав: только назначенный член экипажа (второй пилот/участник) или администратор/диспетчер
        membership = crew.members.filter(member=request.user).first()
        is_crew_member = membership is not None
        is_admin = is_flight_planner(request.user)
        can_user_add_note = (is_date_allowed and is_crew_member) or is_admin

        roles_dict = dict(CREW_ROLES)
        notes_data = [
            {
                'id': n.id,
                'author_id': n.author_id,
                'author_name': n.author.title if (n.author and n.author.title) else (n.author.username if n.author else 'Неизвестно'),
                'author_role': n.author_role,
                'author_role_label': roles_dict.get(n.author_role, n.author_role),
                'message': n.message,
                'created_at': n.created_at.strftime('%d.%m.%Y %H:%M'),
                'created_at_time': n.created_at.strftime('%H:%M'),
                'can_delete': (is_admin or (n.author_id == request.user.id))
            }
            for n in crew.notes.all()
        ]

        # Находим второго пилота (или членов экипажа)
        copilot_member = crew.members.filter(role='copilot').first()
        copilot_name = (copilot_member.member.title or copilot_member.member.username) if copilot_member else ""

        return JsonResponse({
            'status': 'success',
            'crew_id': crew.id,
            'mpd_name': crew.mpd.name,
            'aircraft_number': crew.aircraft.registration_number if crew.aircraft else 'Резерв',
            'date': crew.date.isoformat(),
            'date_formatted': crew.date.strftime('%d.%m.%Y'),
            'copilot_name': copilot_name,
            'is_date_allowed': is_date_allowed,
            'is_crew_member': is_crew_member,
            'is_admin': is_admin,
            'can_user_add_note': can_user_add_note,
            'min_editable_date': min_editable_date.strftime('%d.%m.%Y'),
            'max_editable_date': max_editable_date.strftime('%d.%m.%Y'),
            'notes': notes_data
        })
    except FlightCrew.DoesNotExist:
        return JsonResponse({'error': 'Экипаж не найден.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def save_crew_note_api(request, crew_id: int):
    """API добавления пометки к полету.

    Ограничение для экипажа: разрешено только для рейсов на вчера, сегодня и завтра (сегодня +- 1 день).
    Диспетчеры/планировщики могут добавлять пометки на любые даты.

    Args:
        request (HttpRequest): HTTP-запрос с текстом сообщения.
        crew_id (int): Идентификатор экипажа.

    Returns:
        JsonResponse: Сохраненная пометка или сообщение об ошибке.
    """
    try:
        crew = FlightCrew.objects.select_related('aircraft', 'mpd').prefetch_related('members').get(id=crew_id)
        today = timezone.now().date()
        min_editable_date = today - timedelta(days=1)
        max_editable_date = today + timedelta(days=1)

        # Проверка принадлежности к экипажу
        membership = crew.members.filter(member=request.user).first()
        is_crew_member = membership is not None
        is_admin = is_flight_planner(request.user)

        if not is_admin:
            if not is_crew_member:
                return JsonResponse({
                    'error': 'Вы не назначены в данный экипаж. Оставлять пометки к полету разрешено только назначенному члену экипажа или диспетчеру.'
                }, status=403)
            if not (min_editable_date <= crew.date <= max_editable_date):
                return JsonResponse({
                    'error': f'Ввод пометок летным составом разрешен только для рейсов на вчера ({min_editable_date.strftime("%d.%m.%Y")}), сегодня ({today.strftime("%d.%m.%Y")}) и завтра ({max_editable_date.strftime("%d.%m.%Y")}).'
                }, status=403)

        data = json.loads(request.body)
        message = data.get('message', '').strip()
        if not message:
            return JsonResponse({'error': 'Текст пометки не может быть пустым.'}, status=400)

        author_role = membership.role if membership else 'dispatcher'

        note = FlightCrewNote.objects.create(
            crew=crew,
            author=request.user,
            author_role=author_role,
            message=message
        )

        roles_dict = dict(CREW_ROLES)
        return JsonResponse({
            'status': 'success',
            'message': 'Пометка к полету успешно добавлена.',
            'note': {
                'id': note.id,
                'author_id': note.author_id,
                'author_name': note.author.title if (note.author and note.author.title) else (note.author.username if note.author else 'Неизвестно'),
                'author_role': note.author_role,
                'author_role_label': roles_dict.get(note.author_role, note.author_role),
                'message': note.message,
                'created_at': note.created_at.strftime('%d.%m.%Y %H:%M'),
                'created_at_time': note.created_at.strftime('%H:%M'),
                'can_delete': True
            }
        })
    except FlightCrew.DoesNotExist:
        return JsonResponse({'error': 'Экипаж не найден.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_crew_note_api(request, note_id: int):
    """API удаления пометки к полету (доступно автору или диспетчеру).

    Args:
        request (HttpRequest): HTTP-запрос.
        note_id (int): Идентификатор пометки.

    Returns:
        JsonResponse: Результат удаления пометки.
    """
    try:
        note = FlightCrewNote.objects.select_related('crew').get(id=note_id)
        is_admin = is_flight_planner(request.user)
        if not (is_admin or note.author_id == request.user.id):
            return JsonResponse({'error': 'У вас нет прав на удаление этой пометки.'}, status=403)

        crew_id = note.crew_id
        note.delete()
        remaining_notes_count = FlightCrewNote.objects.filter(crew_id=crew_id).count()

        return JsonResponse({
            'status': 'success',
            'message': 'Пометка удалена.',
            'crew_id': crew_id,
            'remaining_notes_count': remaining_notes_count
        })
    except FlightCrewNote.DoesNotExist:
        return JsonResponse({'error': 'Пометка не найдена.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@flight_reports_required
def personnel_utilization_report_view(request):
    """Отображает аналитический отчет по производственной загрузке летного состава.

    Группирует персонал по 4 авиационно-кадровым группам загрузки. Доступен руководству
    и сотрудникам планирования.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Страница аналитического отчета по загрузке персонала.
    """
    now = timezone.now()
    year = request.GET.get('year', now.year)
    month = request.GET.get('month', now.month)
    job_category = request.GET.get('job_category', '').strip()

    try:
        year = int(year)
        month = int(month)
    except (ValueError, TypeError):
        year = now.year
        month = now.month

    if month < 1:
        month = 1
    elif month > 12:
        month = 12

    # Получаем сгруппированные данные отчета
    report_data = get_personnel_utilization_report_data(year=year, month=month, job_category=job_category or None)

    # Навигация по месяцам
    first_day = datetime(year, month, 1).date()
    prev_month_date = first_day - timedelta(days=1)
    if month == 12:
        next_month_date = datetime(year + 1, 1, 1).date()
    else:
        next_month_date = datetime(year, month + 1, 1).date()

    MONTHS_LIST = [
        (1, 'Январь'), (2, 'Февраль'), (3, 'Март'), (4, 'Апрель'),
        (5, 'Май'), (6, 'Июнь'), (7, 'Июль'), (8, 'Август'),
        (9, 'Сентябрь'), (10, 'Октябрь'), (11, 'Ноябрь'), (12, 'Декабрь')
    ]

    JOB_CATEGORIES = [
        ('', 'Все должности летного состава'),
        ('commander', 'КВС (Командиры воздушных судов)'),
        ('copilot', 'Вторые пилоты'),
        ('instructor', 'Пилоты-инструкторы / Инспекторы'),
        ('flight_engineer', 'Бортмеханики / Бортинженеры'),
    ]

    years_list = list(range(now.year - 2, now.year + 3))

    user_division = ""
    try:
        if hasattr(request.user, 'user_work_profile') and request.user.user_work_profile and request.user.user_work_profile.divisions:
            user_division = request.user.user_work_profile.divisions.name or str(request.user.user_work_profile.divisions)
    except Exception:
        pass

    if not user_division:
        user_division = "Служба планирования и организации полетов"

    is_planner = is_flight_planner(request.user)

    context = {
        'report': report_data,
        'year': year,
        'month': month,
        'job_category': job_category,
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
        'months_list': MONTHS_LIST,
        'years_list': years_list,
        'job_categories': JOB_CATEGORIES,
        'generation_time': now.strftime('%d.%m.%Y %H:%M'),
        'current_user': request.user,
        'user_division': user_division,
        'is_planner': is_planner,
        'can_view_reports': True,
    }

    return render(request, 'flight_planning/utilization_report.html', context)


def generate_basing_excel_response(report_data: dict, company_name: str, user_division: str, author_name: str) -> HttpResponse:
    """Генерирует официальный файл Excel (.xlsx) с отчетом «Базирование ВС на дату».

    Args:
        report_data (dict): Словарь с агрегированными данными базирования.
        company_name (str): Наименование организации.
        user_division (str): Наименование подразделения.
        author_name (str): ФИО составителя отчета.

    Returns:
        HttpResponse: HTTP-ответ с прикрепленным Excel-файлом.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Базирование {report_data['target_date_short']}"
    ws.views.sheetView[0].showGridLines = True

    # Стили
    company_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
    division_font = Font(name="Calibri", size=10, italic=True, color="475569")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    header_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
    data_font = Font(name="Calibri", size=10, color="000000")
    bold_data_font = Font(name="Calibri", size=10, bold=True, color="000000")
    reserve_font = Font(name="Calibri", size=10, bold=True, color="B91C1C")
    footer_font = Font(name="Calibri", size=9, italic=True, color="64748B")

    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    mpd_bg_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    table_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    header_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=Side(border_style="medium", color="475569")
    )

    # 1. Шапка документа
    ws.merge_cells("A1:F1")
    ws["A1"] = company_name
    ws["A1"].font = company_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = user_division
    ws["A2"].font = division_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:F4")
    ws["A4"] = f"БАЗИРОВАНИЕ ВС НА {report_data['target_date_formatted']} год"
    ws["A4"].font = title_font
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28

    # 2. Заголовки колонок
    headers = ["№", "МПД", "Тип ВС", "№", "Дата прибытия", "Примечания"]
    header_row_idx = 6

    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row_idx].height = 24

    current_row = header_row_idx + 1

    # 3. Данные по МПД
    for group in report_data['mpd_groups']:
        aircrafts = group['aircrafts']

        for i, ac in enumerate(aircrafts):
            is_first = (i == 0)

            # Колонка 1: Порядковый номер МПД
            cell_num = ws.cell(row=current_row, column=1, value=f"{group['index']}." if is_first else "")
            cell_num.font = bold_data_font if is_first else data_font
            cell_num.alignment = Alignment(horizontal="center", vertical="top")
            cell_num.border = table_border
            if is_first:
                cell_num.fill = mpd_bg_fill

            # Колонка 2: Название МПД
            cell_mpd = ws.cell(row=current_row, column=2, value=group['mpd_name'] if is_first else "")
            cell_mpd.font = bold_data_font if is_first else data_font
            cell_mpd.alignment = Alignment(horizontal="left", vertical="top")
            cell_mpd.border = table_border
            if is_first:
                cell_mpd.fill = mpd_bg_fill

            # Колонка 3: Тип ВС
            cell_type = ws.cell(row=current_row, column=3, value=ac['type_name'])
            cell_type.font = data_font
            cell_type.alignment = Alignment(horizontal="center", vertical="center")
            cell_type.border = table_border

            # Колонка 4: Номер ВС
            cell_reg = ws.cell(row=current_row, column=4, value=ac['registration_number'])
            cell_reg.font = bold_data_font
            cell_reg.alignment = Alignment(horizontal="center", vertical="center")
            cell_reg.border = table_border

            # Колонка 5: Дата прибытия
            cell_date = ws.cell(row=current_row, column=5, value=ac['arrival_date_formatted'])
            cell_date.font = data_font
            cell_date.alignment = Alignment(horizontal="center", vertical="center")
            cell_date.border = table_border

            # Колонка 6: Примечания
            cell_comment = ws.cell(row=current_row, column=6, value=ac['comment'])
            cell_comment.font = reserve_font if ac['is_reserve'] else data_font
            cell_comment.alignment = Alignment(horizontal="left", vertical="center")
            cell_comment.border = table_border

            ws.row_dimensions[current_row].height = 20
            current_row += 1

    # 4. Итоговая строка
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    type_str = ", ".join([f"{t['type_name']}: {t['count']}" for t in report_data['type_counts']]) if report_data['type_counts'] else "нет данных"
    summary_text = (
        f"ИТОГО: ВС на базировании: {report_data['total_aircrafts']} ед. "
        f"(Задействовано МПД: {report_data['total_mpds']}, В резерве: {report_data['total_reserve']}). "
        f"Распределение по типам: {type_str}"
    )
    cell_summary = ws.cell(row=current_row, column=1, value=summary_text)
    cell_summary.font = bold_data_font
    cell_summary.alignment = Alignment(horizontal="left", vertical="center")

    current_row += 2
    footer_text = (
        f"Составитель: {author_name} ({user_division}) | "
        f"Дата и время формирования: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
    )
    ws.cell(row=current_row, column=1, value=footer_text).font = footer_font

    # Автоматическая настройка ширины колонок
    col_widths = {1: 8, 2: 38, 3: 16, 4: 16, 5: 18, 6: 32}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"basing_aircraft_{report_data['target_date_formatted']}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@flight_reports_required
def aircraft_basing_report_view(request):
    """Отображает официальный отчет «Базирование ВС на дату».

    Группирует ВС по МПД, показывает тип, бортовой номер, дату прибытия и примечания.
    Поддерживает фильтрацию и экспорт в Excel. Доступен руководству и планировщикам.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Страница отчета базирования или скачиваемый файл Excel.
    """
    date_str = request.GET.get('date', '').strip()
    mpd_id = request.GET.get('mpd', '').strip()
    type_id = request.GET.get('type', '').strip()
    export_format = request.GET.get('export', '').strip().lower()

    # Определение даты среза
    target_date = timezone.now().date()
    if date_str:
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                target_date = datetime.strptime(date_str, fmt).date()
                break
            except (ValueError, TypeError):
                continue

    try:
        mpd_id_int = int(mpd_id) if mpd_id else None
    except ValueError:
        mpd_id_int = None

    try:
        type_id_int = int(type_id) if type_id else None
    except ValueError:
        type_id_int = None

    # Получаем сгруппированные данные отчета через селектор
    report_data = get_aircraft_basing_report_data(
        target_date=target_date,
        mpd_id=mpd_id_int,
        aircraft_type_id=type_id_int
    )

    # Определение подразделения пользователя
    user_division = ""
    try:
        if hasattr(request.user, 'user_work_profile') and request.user.user_work_profile and request.user.user_work_profile.divisions:
            user_division = request.user.user_work_profile.divisions.name or str(request.user.user_work_profile.divisions)
    except Exception:
        pass

    if not user_division:
        user_division = "Служба планирования и организации полетов"

    author_name = (
        request.user.title
        if (hasattr(request.user, 'title') and request.user.title)
        else (request.user.get_full_name() or request.user.username)
    )
    company_name = "ООО «Авиакомпания «БАРКОЛ»"

    # Экспорт в Excel при наличии параметра ?export=excel
    if export_format == 'excel':
        return generate_basing_excel_response(
            report_data=report_data,
            company_name=company_name,
            user_division=user_division,
            author_name=author_name
        )

    # Справочники для фильтрации
    mpds_list = PlaceProductionActivity.objects.filter(in_planning=True).order_by('name')
    types_list = TypeProperty.objects.filter(estate__isnull=False).distinct().order_by('type_property')

    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)

    is_planner = is_flight_planner(request.user)

    context = {
        'report': report_data,
        'target_date': target_date,
        'target_date_str': target_date.strftime('%Y-%m-%d'),
        'target_date_formatted': target_date.strftime('%d.%m.%Y'),
        'target_date_short': target_date.strftime('%d.%m.%y'),
        'prev_date_str': prev_date.strftime('%Y-%m-%d'),
        'next_date_str': next_date.strftime('%Y-%m-%d'),
        'mpds_list': mpds_list,
        'types_list': types_list,
        'selected_mpd_id': mpd_id_int,
        'selected_type_id': type_id_int,
        'user_division': user_division,
        'author_name': author_name,
        'company_name': company_name,
        'generation_time': timezone.now().strftime('%d.%m.%Y %H:%M'),
        'title': f"Базирование ВС на {target_date.strftime('%d.%m.%Y')} год",
        'is_planner': is_planner,
        'can_view_reports': True,
    }

    return render(request, 'flight_planning/aircraft_basing_report.html', context)


@login_required
@flight_planner_required
@csrf_exempt
@require_http_methods(["POST"])
def batch_swap_aircraft_api(request):
    """API эндпоинт для пакетной замены борта ВС в экипажах на интервал дат.

    Args:
        request (HttpRequest): HTTP-запрос с параметрами замены ВС.

    Returns:
        JsonResponse: Результат пакетного обновления экипажей.
    """
    try:
        data = json.loads(request.body)
        mpd_id = data.get('mpd_id')
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date') or start_date_str
        old_aircraft_id = data.get('old_aircraft_id')
        new_aircraft_id = data.get('new_aircraft_id')

        if not mpd_id or not start_date_str or not end_date_str:
            return JsonResponse({'error': 'Необходимо указать МПД, начальную и конечную дату.'}, status=400)

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({'error': 'Некорректный формат даты (ожидается YYYY-MM-DD).'}, status=400)

        result = batch_swap_aircraft(
            mpd_id=int(mpd_id),
            start_date=start_date,
            end_date=end_date,
            old_aircraft_id=old_aircraft_id,
            new_aircraft_id=new_aircraft_id,
            created_by=request.user
        )

        if result.get('status') == 'error':
            return JsonResponse({'error': ' '.join(result.get('errors', [])), 'errors': result.get('errors')}, status=400)

        return JsonResponse({
            'status': 'success',
            'message': result.get('message', 'Замена борта выполнена успешно.'),
            'updated_count': result.get('updated_count', 0)
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ========================================================
# ПЕРИОДИЧЕСКИЕ МЕРОПРИЯТИЯ (ЖУРНАЛ И СПРАВОЧНИКИ)
# ========================================================

@login_required
@flight_planning_view_required
def periodic_check_list_view(request):
    """Отображает журнал периодических мероприятий сотрудников, матрицу квалификации и справочник видов мероприятий.

    Args:
        request (HttpRequest): Объект HTTP-запроса с параметрами фильтрации:
            - tab: активная вкладка ('records', 'matrix', 'types');
            - employee_id: фильтр по сотруднику;
            - aircraft_type_id: фильтр по типу ВС;
            - check_type_id: фильтр по виду мероприятия;
            - status: фильтр по статусу ('valid', 'warning', 'expired');
            - q: поисковый запрос (по ФИО).

    Returns:
        HttpResponse: Отрендеренная страница журнала периодических мероприятий.
    """
    active_tab = request.GET.get('tab', 'records')
    employee_id = request.GET.get('employee_id')
    aircraft_type_id = request.GET.get('aircraft_type_id')
    check_type_id = request.GET.get('check_type_id')
    status_filter = request.GET.get('status')
    search_query = request.GET.get('q', '').strip()

    today = timezone.now().date()
    warning_threshold_date = today + timedelta(days=30)

    # Список разрешенных сотрудников согласно роли и принадлежности текущего пользователя
    pilots_list = get_allowed_staff_queryset(user=request.user)

    # Базовая выборка всех записей для доступного состава сотрудников
    all_records = PeriodicCheckRecord.objects.filter(
        employee__in=pilots_list
    ).select_related(
        'employee', 'employee__user_work_profile__job',
        'check_type', 'aircraft_type', 'created_by'
    ).order_by('-end_date', '-start_date', '-id')

    # Группировка для вычисления актуальных vs продленных (архивных) записей
    grouped_by_emp_check: Dict[tuple, List[PeriodicCheckRecord]] = {}
    for r in all_records:
        key = (r.employee_id, r.check_type_id, r.aircraft_type_id or 0)
        if key not in grouped_by_emp_check:
            grouped_by_emp_check[key] = []
        grouped_by_emp_check[key].append(r)

    latest_record_ids = set()
    superseded_record_map: Dict[int, PeriodicCheckRecord] = {}

    for key, rec_list in grouped_by_emp_check.items():
        latest_rec = rec_list[0]
        latest_record_ids.add(latest_rec.id)
        for older_rec in rec_list[1:]:
            superseded_record_map[older_rec.id] = latest_rec

    # Подсчет KPI для актуальных (последних) записей квалификации сотрудников
    latest_records_list = [r for r in all_records if r.id in latest_record_ids]
    total_records_count = all_records.count()
    expired_records_count = sum(1 for r in latest_records_list if r.end_date < today)
    warning_records_count = sum(1 for r in latest_records_list if today <= r.end_date <= warning_threshold_date)
    valid_records_count = sum(1 for r in latest_records_list if r.end_date > warning_threshold_date)
    renewed_records_count = len(superseded_record_map)

    # Фильтрация реестра записей
    records_qs = all_records

    if employee_id:
        try:
            records_qs = records_qs.filter(employee_id=int(employee_id))
        except ValueError:
            pass

    if aircraft_type_id:
        try:
            records_qs = records_qs.filter(aircraft_type_id=int(aircraft_type_id))
        except ValueError:
            pass

    if check_type_id:
        try:
            records_qs = records_qs.filter(check_type_id=int(check_type_id))
        except ValueError:
            pass

    if search_query:
        records_qs = records_qs.filter(
            Q(employee__last_name__icontains=search_query) |
            Q(employee__first_name__icontains=search_query) |
            Q(employee__title__icontains=search_query) |
            Q(document_number__icontains=search_query) |
            Q(issued_by__icontains=search_query)
        )

    if status_filter == 'expired':
        # Только актуальные непройденные/просроченные проверки (без продления)
        records_qs = [r for r in records_qs if r.id in latest_record_ids and r.end_date < today]
    elif status_filter == 'warning':
        # Только актуальные истекающие в течение 30 дней
        records_qs = [r for r in records_qs if r.id in latest_record_ids and today <= r.end_date <= warning_threshold_date]
    elif status_filter == 'valid':
        # Только актуальные действующие проверки
        records_qs = [r for r in records_qs if r.id in latest_record_ids and r.end_date > warning_threshold_date]
    elif status_filter == 'renewed':
        # Продленные архивные записи
        records_qs = [r for r in records_qs if r.id in superseded_record_map]
    else:
        records_qs = list(records_qs)

    # Кэшируем информацию о наличии новой проверки для быстрого рендеринга шаблона
    for r in records_qs:
        successor = superseded_record_map.get(r.id)
        r._is_superseded_cached = (successor is not None)
        r._successor_cached = successor

    # Список типов проверок
    check_types = PeriodicCheckType.objects.all().select_related('aircraft_type').order_by('order', 'name')

    # Формирование данных сводной матрицы персонала
    active_check_types = [ct for ct in check_types if ct.is_active]
    matrix_data = []

    # Пакетная выгрузка последних записей для матрицы
    all_pilot_records = [r for r in all_records if r.check_type_id in [ct.id for ct in active_check_types]]

    pilot_check_dict: Dict[int, Dict[int, PeriodicCheckRecord]] = {}
    for r in all_pilot_records:
        if r.employee_id not in pilot_check_dict:
            pilot_check_dict[r.employee_id] = {}
        if r.check_type_id not in pilot_check_dict[r.employee_id]:
            pilot_check_dict[r.employee_id][r.check_type_id] = r

    # Пакетная выгрузка закрепленных проверок для отображения в матрице
    pilot_ids = [p.id for p in pilots_list]
    assignments_map = get_batch_employee_check_assignments(pilot_ids)

    for p in pilots_list:
        row_cells = []
        pilot_has_expired = False
        pilot_has_warning = False

        job_name = (p.user_work_profile.job.name.lower()) if (hasattr(p, 'user_work_profile') and p.user_work_profile and p.user_work_profile.job) else ''
        is_pilot = ('пилот' in job_name or 'командир' in job_name or 'квс' in job_name)
        is_fe = ('механик' in job_name or 'инженер' in job_name or 'бм' in job_name)
        is_tech = ('техник' in job_name)

        has_custom_assignments = (p.id in assignments_map)
        assigned_type_ids = assignments_map.get(p.id)

        for ct in active_check_types:
            if has_custom_assignments:
                is_required = (assigned_type_ids is not None and ct.id in assigned_type_ids)
            else:
                # По умолчанию согласно категории должности
                is_required = True
                if ct.applies_to == 'pilots' and not is_pilot and (is_fe or is_tech):
                    is_required = False
                elif ct.applies_to == 'flight_engineers' and not is_fe and (is_pilot or is_tech):
                    is_required = False
                elif ct.applies_to == 'technicians' and not is_tech and (is_pilot or is_fe):
                    is_required = False
                elif ct.applies_to == 'crew' and not is_pilot and not is_fe and is_tech:
                    is_required = False

            rec = pilot_check_dict.get(p.id, {}).get(ct.id)
            if not is_required:
                cell_status = 'not_required'
                if rec:
                    st = rec.status_on_date(today)
                    cell_text = f"до {rec.end_date.strftime('%d.%m.%y')} (не обяз.)"
                    cell_class = 'matrix-cell-not-required'
                    days_left = (rec.end_date - today).days
                else:
                    cell_text = 'Не требуется'
                    cell_class = 'matrix-cell-not-required'
                    days_left = None
            else:
                if not rec:
                    cell_status = 'missing'
                    cell_text = '+ Внести'
                    cell_class = 'matrix-cell-missing'
                    pilot_has_expired = True
                    days_left = None
                else:
                    st = rec.status_on_date(today)
                    days_left = (rec.end_date - today).days
                    if st == 'expired':
                        cell_status = 'expired'
                        cell_text = f"до {rec.end_date.strftime('%d.%m.%y')}"
                        cell_class = 'matrix-cell-expired'
                        pilot_has_expired = True
                    elif st == 'warning':
                        cell_status = 'warning'
                        cell_text = f"до {rec.end_date.strftime('%d.%m.%y')}"
                        cell_class = 'matrix-cell-warning'
                        pilot_has_warning = True
                    else:
                        cell_status = 'valid'
                        cell_text = f"до {rec.end_date.strftime('%d.%m.%y')}"
                        cell_class = 'matrix-cell-valid'

            row_cells.append({
                'check_type': ct,
                'record': rec,
                'status': cell_status,
                'is_required': is_required,
                'text': cell_text,
                'css_class': cell_class,
                'days_left': days_left
            })

        matrix_data.append({
            'pilot': p,
            'job': p.user_work_profile.job.name if (hasattr(p, 'user_work_profile') and p.user_work_profile and p.user_work_profile.job) else '',
            'cells': row_cells,
            'has_expired': pilot_has_expired,
            'has_warning': pilot_has_warning,
            'has_custom_assignments': has_custom_assignments
        })

    # Формы для модальных окон быстрого добавления
    record_form = PeriodicCheckRecordForm(user=request.user)
    type_form = PeriodicCheckTypeForm()

    aircraft_types = TypeProperty.objects.all().order_by('type_property')

    context = {
        'title': 'Журнал периодических мероприятий',
        'active_tab': active_tab,
        'records': records_qs,
        'check_types': check_types,
        'matrix_data': matrix_data,
        'active_check_types': active_check_types,
        'aircraft_types': aircraft_types,
        'pilots_list': pilots_list,
        'record_form': record_form,
        'type_form': type_form,
        'total_records_count': total_records_count,
        'expired_records_count': expired_records_count,
        'warning_records_count': warning_records_count,
        'valid_records_count': valid_records_count,
        'renewed_records_count': renewed_records_count,
        'is_planner': is_flight_planner(request.user),
        'today': today,
        'filter_employee_id': employee_id,
        'filter_aircraft_type_id': aircraft_type_id,
        'filter_check_type_id': check_type_id,
        'filter_status': status_filter,
        'search_query': search_query,
    }
    return render(request, 'flight_planning/periodic_checks/check_list.html', context)


@login_required
@flight_planner_required
def periodic_check_create_view(request):
    """Создание новой записи о прохождении периодического мероприятия сотрудником.

    Args:
        request (HttpRequest): Объект HTTP-запроса (GET или POST с файлом).

    Returns:
        HttpResponse: Перенаправление на журнал или отрендеренная форма.
    """
    if request.method == 'POST':
        form = PeriodicCheckRecordForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            record = form.save(commit=False)
            record.created_by = request.user
            record.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': f"Запись мероприятия «{record.check_type.name}» для {record.employee.title} успешно создана!",
                    'record_id': record.id
                })

            messages.success(request, f"Запись мероприятия «{record.check_type.name}» для {record.employee.title} успешно создана!")
            return redirect('flight_planning:periodic_check_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'error',
                    'errors': form.errors,
                    'error': 'Пожалуйста, проверьте правильность заполнения формы.'
                }, status=400)
    else:
        initial_data = {}
        if request.GET.get('employee_id'):
            initial_data['employee'] = request.GET.get('employee_id')
        if request.GET.get('check_type_id'):
            initial_data['check_type'] = request.GET.get('check_type_id')
        form = PeriodicCheckRecordForm(initial=initial_data, user=request.user)

    return render(request, 'flight_planning/periodic_checks/check_record_form.html', {
        'form': form,
        'title': 'Добавление записи о периодическом мероприятии',
        'is_create': True
    })


@login_required
@flight_planner_required
def periodic_check_update_view(request, pk: int):
    """Редактирование записи о прохождении периодического мероприятия.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Первичный ключ записи.

    Returns:
        HttpResponse: Перенаправление на журнал или страница редактирования.
    """
    record = get_object_or_404(PeriodicCheckRecord.objects.select_related('employee', 'check_type'), pk=pk)

    if request.method == 'POST':
        form = PeriodicCheckRecordForm(request.POST, request.FILES, instance=record, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Запись мероприятия «{record.check_type.name}» для {record.employee.title} успешно обновлена!")
            return redirect('flight_planning:periodic_check_list')
    else:
        form = PeriodicCheckRecordForm(instance=record, user=request.user)

    return render(request, 'flight_planning/periodic_checks/check_record_form.html', {
        'form': form,
        'record': record,
        'title': f"Редактирование записи мероприятия: {record.check_type.name}",
        'is_create': False
    })


@login_required
@flight_planner_required
def periodic_check_delete_view(request, pk: int):
    """Удаление записи о прохождении мероприятия.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Первичный ключ записи.

    Returns:
        HttpResponse: Перенаправление на список записей.
    """
    record = get_object_or_404(PeriodicCheckRecord, pk=pk)
    name = f"{record.check_type.name} ({record.employee.title})"
    record.delete()
    messages.success(request, f"Запись мероприятия «{name}» успешно удалена.")
    return redirect('flight_planning:periodic_check_list')


@login_required
@flight_planner_required
def periodic_check_type_create_view(request):
    """Создание нового вида периодического мероприятия.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Перенаправление на справочник или страница с формой.
    """
    if request.method == 'POST':
        form = PeriodicCheckTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Новый вид периодического мероприятия успешно создан!")
            return redirect(f"{reverse('flight_planning:periodic_check_list')}?tab=types")
    else:
        form = PeriodicCheckTypeForm()

    return render(request, 'flight_planning/periodic_checks/check_type_form.html', {
        'form': form,
        'title': 'Создание вида периодического мероприятия',
        'is_create': True
    })


@login_required
@flight_planner_required
def periodic_check_type_update_view(request, pk: int):
    """Редактирование вида периодического мероприятия.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор редактируемого вида мероприятия.

    Returns:
        HttpResponse: Перенаправление на справочник или страница с формой.
    """
    check_type = get_object_or_404(PeriodicCheckType, pk=pk)
    if request.method == 'POST':
        form = PeriodicCheckTypeForm(request.POST, instance=check_type)
        if form.is_valid():
            form.save()
            messages.success(request, f"Вид мероприятия «{check_type.name}» успешно обновлен!")
            return redirect(f"{reverse('flight_planning:periodic_check_list')}?tab=types")
    else:
        form = PeriodicCheckTypeForm(instance=check_type)

    return render(request, 'flight_planning/periodic_checks/check_type_form.html', {
        'form': form,
        'check_type': check_type,
        'title': f'Редактирование вида мероприятия: {check_type.name}',
        'is_create': False
    })


@login_required
@flight_planner_required
def periodic_check_type_delete_view(request, pk: int):
    """Удаление вида периодического мероприятия (если нет связанных записей).

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Первичный ключ вида мероприятия.

    Returns:
        HttpResponse: Перенаправление на справочник с сообщением.
    """
    check_type = get_object_or_404(PeriodicCheckType, pk=pk)
    records_count = check_type.records.count()
    if records_count > 0:
        messages.error(request, f"Невозможно удалить вид мероприятия «{check_type.name}», так как с ним связано {records_count} записей в журнале. Сначала удалите или переместите эти записи.")
    else:
        name = check_type.name
        check_type.delete()
        messages.success(request, f"Вид мероприятия «{name}» успешно удален.")
    return redirect(f"{reverse('flight_planning:periodic_check_list')}?tab=types")


@login_required
@require_http_methods(["GET"])
def get_pilot_checks_api(request, pilot_id: int):
    """REST API получения статуса проверок конкретного пилота/сотрудника.

    Args:
        request (HttpRequest): HTTP-запрос.
        pilot_id (int): Идентификатор сотрудника.

    Returns:
        JsonResponse: Статус всех проверок пилота.
    """
    target_date_str = request.GET.get('date')
    aircraft_type_id = request.GET.get('aircraft_type_id')

    target_date = None
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = None

    ac_id = int(aircraft_type_id) if aircraft_type_id and aircraft_type_id.isdigit() else None

    status_data = get_pilot_periodic_check_status(
        pilot_id=pilot_id,
        target_date=target_date,
        aircraft_type_id=ac_id
    )
    return JsonResponse({'status': 'success', 'check_status': status_data})


@login_required
@require_http_methods(["GET"])
def calculate_check_date_api(request):
    """API автоматического расчета даты окончания проверки по виду проверки и дате начала.

    Args:
        request (HttpRequest): HTTP-запрос с параметрами:
            - check_type_id: ID вида проверки;
            - start_date: дата начала (YYYY-MM-DD).

    Returns:
        JsonResponse: Расчетная дата окончания (end_date).
    """
    check_type_id = request.GET.get('check_type_id')
    start_date_str = request.GET.get('start_date')

    if not check_type_id or not start_date_str:
        return JsonResponse({'error': 'Параметры check_type_id и start_date обязательны.'}, status=400)

    try:
        check_type = PeriodicCheckType.objects.get(id=int(check_type_id))
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = calculate_check_end_date(
            start_date=start_date,
            validity_months=check_type.validity_months,
            validity_days=check_type.validity_days
        )
        return JsonResponse({
            'status': 'success',
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'validity_months': check_type.validity_months,
            'aircraft_type_id': check_type.aircraft_type_id,
            'aircraft_type_name': check_type.aircraft_display
        })
    except PeriodicCheckType.DoesNotExist:
        return JsonResponse({'error': 'Вид проверки не найден.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_check_history_api(request):
    """API получения истории прохождения периодических мероприятий сотрудником.

    Args:
        request (HttpRequest): HTTP-запрос с GET-параметрами:
            - employee_id (int, обязательный): ID сотрудника;
            - check_type_id (int, опциональный): ID вида мероприятия для фильтрации.

    Returns:
        JsonResponse: Список исторических записей мероприятий, текущий статус и метаданные.
    """
    employee_id = request.GET.get('employee_id')
    check_type_id = request.GET.get('check_type_id')

    if not employee_id:
        return JsonResponse({'status': 'error', 'error': 'Параметр employee_id обязателен.'}, status=400)

    try:
        employee = DataBaseUser.objects.select_related('user_work_profile__job').get(id=int(employee_id))
    except (DataBaseUser.DoesNotExist, ValueError):
        return JsonResponse({'status': 'error', 'error': 'Сотрудник не найден.'}, status=404)

    today = timezone.now().date()
    records_qs = PeriodicCheckRecord.objects.filter(
        employee_id=employee.id
    ).select_related(
        'check_type', 'aircraft_type', 'created_by'
    ).order_by('-end_date', '-start_date', '-id')

    target_check_type = None
    if check_type_id:
        try:
            target_check_type = PeriodicCheckType.objects.get(id=int(check_type_id))
            records_qs = records_qs.filter(check_type_id=target_check_type.id)
        except (PeriodicCheckType.DoesNotExist, ValueError):
            pass

    history_items = []
    for idx, r in enumerate(records_qs):
        is_latest = (idx == 0)
        is_superseded = not is_latest
        if is_superseded:
            st = 'renewed'
        else:
            st = r.status_on_date(today)

        history_items.append({
            'id': r.id,
            'check_type_id': r.check_type_id,
            'check_type_name': r.check_type.name,
            'aircraft_display': r.check_type.aircraft_display,
            'aircraft_type_name': r.aircraft_type.type_property if r.aircraft_type else "Все типы ВС (*)",
            'aircraft_type_id': r.aircraft_type_id,
            'start_date': r.start_date.strftime('%d.%m.%Y'),
            'end_date': r.end_date.strftime('%d.%m.%Y'),
            'start_date_iso': r.start_date.isoformat(),
            'end_date_iso': r.end_date.isoformat(),
            'status': st,
            'is_latest': is_latest,
            'is_superseded': is_superseded,
            'days_remaining': r.days_remaining,
            'days_overdue': r.days_overdue,
            'document_number': r.document_number or "—",
            'issued_by': r.issued_by or "—",
            'notes': r.notes or "",
            'scan_url': r.scan_file.url if r.scan_file else None,
            'scan_name': r.scan_file.name.split('/')[-1] if r.scan_file else None
        })

    latest_rec = history_items[0] if history_items else None

    return JsonResponse({
        'status': 'success',
        'employee_id': employee.id,
        'employee_name': employee.title or f"{employee.last_name} {employee.first_name}".strip() or employee.username,
        'employee_job': employee.user_work_profile.job.name if hasattr(employee, 'user_work_profile') and employee.user_work_profile and employee.user_work_profile.job else "",
        'check_type_id': target_check_type.id if target_check_type else None,
        'check_type_name': target_check_type.name if target_check_type else "Все мероприятия",
        'history_count': len(history_items),
        'latest_record': latest_rec,
        'records': history_items
    })


@login_required
@require_http_methods(["GET"])
def get_employee_check_assignments_api(request, employee_id: int):
    """API получения перечня закрепленных за сотрудником обязательных мероприятий.

    Args:
        request (HttpRequest): HTTP-запрос.
        employee_id (int): Идентификатор сотрудника.

    Returns:
        JsonResponse: JSON с текущими закреплениями и всеми доступными видами мероприятий.
    """
    data = get_employee_check_assignments(employee_id)
    if data.get('status') == 'error':
        return JsonResponse(data, status=404)
    return JsonResponse(data)


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def save_employee_check_assignments_api(request):
    """API сохранения индивидуального закрепления обязательных мероприятий за сотрудником.

    Args:
        request (HttpRequest): HTTP-запрос (JSON или form-data) с параметрами:
            - employee_id (int): Идентификатор сотрудника;
            - check_type_ids (List[int]): Массив ID закрепляемых мероприятий.

    Returns:
        JsonResponse: Результат сохранения.
    """
    if request.content_type == 'application/json':
        try:
            body = json.loads(request.body)
            employee_id = body.get('employee_id')
            check_type_ids = body.get('check_type_ids', [])
        except Exception:
            return JsonResponse({'status': 'error', 'error': 'Некорректный формат JSON.'}, status=400)
    else:
        employee_id = request.POST.get('employee_id')
        check_type_ids = request.POST.getlist('check_type_ids')

    if not employee_id:
        return JsonResponse({'status': 'error', 'error': 'Параметр employee_id обязателен.'}, status=400)

    try:
        emp_id = int(employee_id)
        cid_list = [int(cid) for cid in check_type_ids]
    except (ValueError, TypeError):
        return JsonResponse({'status': 'error', 'error': 'Некорректные идентификаторы мероприятий.'}, status=400)

    save_employee_check_assignments(
        employee_id=emp_id,
        check_type_ids=cid_list,
        assigned_by=request.user
    )

    return JsonResponse({
        'status': 'success',
        'message': 'Индивидуальный перечень обязательных мероприятий успешно сохранен!'
    })


# ========================================================
# ПРЕДСТАВЛЕНИЯ МОДУЛЯ «СОСТОЯНИЯ И СТАТУСЫ ПЕРСОНАЛА»
# ========================================================

@login_required
@flight_planning_view_required
def employee_status_list_view(request):
    """Дашборд и журнал учета состояний/статусов сотрудников (Отпуск, Больничный, Резерв, КПК, ВЛЭК).

    Включает 3 вкладки:
    1. «Журнал записей» — полный реестр с фильтрами по сотрудникам, статусам и датам.
    2. «Сводная сетка занятости» — матрица сотрудников по дням выбранного месяца.
    3. «Справочник видов состояний» — настройка видов статусов и их цветов.

    Args:
        request (HttpRequest): Объект HTTP-запроса с GET-параметрами фильтрации.

    Returns:
        HttpResponse: Отрендеренная страница дашборда состояний сотрудников.
    """
    active_tab = request.GET.get('tab', 'records')
    today = timezone.now().date()

    # Фильтры журнала
    employee_id = request.GET.get('employee')
    status_type_id = request.GET.get('status_type')
    month = request.GET.get('month', today.month)
    year = request.GET.get('year', today.year)
    search_query = request.GET.get('q', '').strip()

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        month = today.month
        year = today.year

    start_of_month = date(year, month, 1)
    if month == 12:
        end_of_month = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_of_month = date(year, month + 1, 1) - timedelta(days=1)

    records_qs = EmployeeStatusRecord.objects.all().select_related(
        'employee', 'employee__user_work_profile', 'employee__user_work_profile__job',
        'status_type', 'created_by'
    )

    # Список разрешенных сотрудников согласно роли и принадлежности текущего пользователя
    pilots_list = get_allowed_staff_queryset(user=request.user)

    # Ограничение выборки записей доступными сотрудниками
    records_qs = records_qs.filter(employee__in=pilots_list)

    if employee_id:
        try:
            records_qs = records_qs.filter(employee_id=int(employee_id))
        except ValueError:
            pass

    if status_type_id:
        try:
            records_qs = records_qs.filter(status_type_id=int(status_type_id))
        except ValueError:
            pass

    if search_query:
        records_qs = records_qs.filter(
            Q(employee__last_name__icontains=search_query) |
            Q(employee__first_name__icontains=search_query) |
            Q(employee__title__icontains=search_query) |
            Q(document_number__icontains=search_query) |
            Q(notes__icontains=search_query)
        )

    # KPI счетчики на сегодняшнюю дату (для доступного состава)
    all_active_today = EmployeeStatusRecord.objects.filter(
        employee__in=pilots_list,
        start_date__lte=today,
        end_date__gte=today
    )
    total_active_today_count = all_active_today.count()
    sick_leave_today_count = all_active_today.filter(
        Q(status_type__code='SICK_LEAVE') | Q(status_type__name__icontains='Больничный')
    ).count()
    vacation_today_count = all_active_today.filter(
        Q(status_type__code__in=['VACATION', 'EXTRA_VACATION']) | Q(status_type__name__icontains='Отпуск')
    ).count()
    reserve_today_count = all_active_today.filter(
        Q(status_type__code='RESERVE') | Q(status_type__name__icontains='Резерв')
    ).count()
    training_today_count = all_active_today.filter(
        Q(status_type__code__in=['KPK', 'VLEK', 'MEDICAL_EXAM']) | Q(status_type__name__icontains='КПК') | Q(status_type__name__icontains='ВЛЭК')
    ).count()

    total_records_count = EmployeeStatusRecord.objects.filter(employee__in=pilots_list).count()

    # Справочник видов статусов
    status_types = EmployeeStatusType.objects.all().order_by('order', 'name')

    days_in_month = (end_of_month - start_of_month).days + 1
    days_list = [date(year, month, d) for d in range(1, days_in_month + 1)]

    # Выборка записей для матрицы месяца (только для доступного персонала)
    month_records = EmployeeStatusRecord.objects.filter(
        employee__in=pilots_list,
        start_date__lte=end_of_month,
        end_date__gte=start_of_month
    ).select_related('employee', 'status_type')

    # Индексация записей: (employee_id, date) -> status_record
    status_matrix_lookup = {}
    for r in month_records:
        curr = max(start_of_month, r.start_date)
        last_d = min(end_of_month, r.end_date)
        while curr <= last_d:
            status_matrix_lookup[(r.employee_id, curr)] = r
            curr += timedelta(days=1)

    matrix_rows = []
    for p in pilots_list:
        p_name = p.title or f"{p.last_name} {p.first_name}".strip() or p.username
        job_name = p.user_work_profile.job.name if (hasattr(p, 'user_work_profile') and p.user_work_profile and p.user_work_profile.job) else ''
        p_cells = []
        has_any_status = False

        for d in days_list:
            rec = status_matrix_lookup.get((p.id, d))
            if rec:
                has_any_status = True
                # Вычисляем краткую аббревиатуру для компактного вывода в клетке
                code = rec.status_type.code
                if code == 'VACATION':
                    abbr = 'ОТ'
                elif code == 'EXTRA_VACATION':
                    abbr = 'ДО'
                elif code == 'SICK_LEAVE':
                    abbr = 'Б'
                elif code == 'RESERVE':
                    abbr = 'Р'
                elif code == 'KPK':
                    abbr = 'КПК'
                elif code == 'VLEK':
                    abbr = 'ВЛ'
                elif code == 'MEDICAL_EXAM':
                    abbr = 'МО'
                elif code == 'BUSINESS_TRIP':
                    abbr = 'КМ'
                elif code == 'DAY_OFF':
                    abbr = 'В'
                else:
                    abbr = rec.status_type.name[:2].upper()

                p_cells.append({
                    'date': d,
                    'is_today': (d == today),
                    'has_status': True,
                    'record': rec,
                    'abbr': abbr,
                    'name': rec.status_type.name,
                    'color': rec.status_type.color,
                    'is_blocking': rec.status_type.is_blocking,
                    'tooltip': f"{rec.status_type.name} (c {rec.start_date.strftime('%d.%m')} по {rec.end_date.strftime('%d.%m')})"
                })
            else:
                p_cells.append({
                    'date': d,
                    'is_today': (d == today),
                    'has_status': False,
                    'record': None,
                    'abbr': '',
                    'name': '',
                    'color': '',
                    'is_blocking': False,
                    'tooltip': ''
                })

        matrix_rows.append({
            'pilot': p,
            'pilot_name': p_name,
            'job': job_name,
            'cells': p_cells,
            'has_any_status': has_any_status
        })

    prev_month_date = start_of_month - timedelta(days=1)
    next_month_date = end_of_month + timedelta(days=1)

    record_form = EmployeeStatusRecordForm(user=request.user)
    type_form = EmployeeStatusTypeForm()

    context = {
        'title': 'Состояния и статусы сотрудников',
        'active_tab': active_tab,
        'records': records_qs,
        'status_types': status_types,
        'matrix_rows': matrix_rows,
        'days_list': days_list,
        'pilots_list': pilots_list,
        'record_form': record_form,
        'type_form': type_form,
        'year': year,
        'month': month,
        'month_name': start_of_month.strftime('%B %Y'),
        'prev_year': prev_month_date.year,
        'prev_month': prev_month_date.month,
        'next_year': next_month_date.year,
        'next_month': next_month_date.month,
        'today': today,
        'total_active_today_count': total_active_today_count,
        'sick_leave_today_count': sick_leave_today_count,
        'vacation_today_count': vacation_today_count,
        'reserve_today_count': reserve_today_count,
        'training_today_count': training_today_count,
        'total_records_count': total_records_count,
        'is_planner': is_flight_planner(request.user),
        'filter_employee_id': employee_id,
        'filter_status_type_id': status_type_id,
        'search_query': search_query,
    }
    return render(request, 'flight_planning/employee_statuses/status_list.html', context)


@login_required
@flight_planner_required
def employee_status_create_view(request):
    """Создание новой записи о состоянии/статусе сотрудника.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Перенаправление или JSON-ответ при AJAX.
    """
    if request.method == 'POST':
        form = EmployeeStatusRecordForm(request.POST, user=request.user)
        if form.is_valid():
            record = form.save(commit=False)
            record.created_by = request.user
            record.save()

            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': f"Запись состояния «{record.status_type.name}» для {record.employee.title} успешно создана!",
                    'record_id': record.id
                })

            messages.success(request, f"Запись состояния «{record.status_type.name}» для {record.employee.title} успешно создана!")
            return redirect('flight_planning:employee_status_list')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
    else:
        form = EmployeeStatusRecordForm(user=request.user)

    return render(request, 'flight_planning/employee_statuses/status_record_form.html', {
        'form': form,
        'title': 'Добавить запись о состоянии сотрудника'
    })


@login_required
@flight_planner_required
def employee_status_update_view(request, pk: int):
    """Редактирование записи о состоянии сотрудника.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор редактируемой записи.

    Returns:
        HttpResponse: Перенаправление на список или страница с формой.
    """
    record = get_object_or_404(EmployeeStatusRecord, pk=pk)
    if request.method == 'POST':
        form = EmployeeStatusRecordForm(request.POST, instance=record, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Запись состояния для {record.employee.title} успешно обновлена!")
            return redirect('flight_planning:employee_status_list')
    else:
        form = EmployeeStatusRecordForm(instance=record, user=request.user)

    return render(request, 'flight_planning/employee_statuses/status_record_form.html', {
        'form': form,
        'record': record,
        'title': f'Редактирование состояния: {record.employee.title} — {record.status_type.name}'
    })


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def employee_status_delete_view(request, pk: int):
    """Удаление записи о состоянии сотрудника.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор удаляемой записи.

    Returns:
        HttpResponse: Перенаправление или JSON-ответ.
    """
    record = get_object_or_404(EmployeeStatusRecord, pk=pk)
    emp_name = record.employee.title or record.employee.username
    st_name = record.status_type.name
    record.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'message': f"Запись «{st_name}» для сотрудника {emp_name} успешно удалена."
        })

    messages.success(request, f"Запись «{st_name}» для сотрудника {emp_name} успешно удалена.")
    return redirect('flight_planning:employee_status_list')


@login_required
@flight_planner_required
def employee_status_type_create_view(request):
    """Создание нового вида состояния сотрудника.

    Args:
        request (HttpRequest): Объект HTTP-запроса.

    Returns:
        HttpResponse: Перенаправление или страница с формой.
    """
    if request.method == 'POST':
        form = EmployeeStatusTypeForm(request.POST)
        if form.is_valid():
            st_type = form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'status': 'success',
                    'message': f"Вид состояния «{st_type.name}» успешно добавлен!",
                    'type_id': st_type.id
                })

            messages.success(request, f"Вид состояния «{st_type.name}» успешно добавлен!")
            return redirect(reverse('flight_planning:employee_status_list') + '?tab=types')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'errors': form.errors}, status=400)
    else:
        form = EmployeeStatusTypeForm()

    return render(request, 'flight_planning/employee_statuses/status_type_form.html', {
        'form': form,
        'title': 'Добавить вид состояния сотрудника'
    })


@login_required
@flight_planner_required
def employee_status_type_update_view(request, pk: int):
    """Редактирование существующего вида состояния.

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор вида состояния.

    Returns:
        HttpResponse: Перенаправление на список или страница с формой.
    """
    st_type = get_object_or_404(EmployeeStatusType, pk=pk)
    if request.method == 'POST':
        form = EmployeeStatusTypeForm(request.POST, instance=st_type)
        if form.is_valid():
            form.save()
            messages.success(request, f"Вид состояния «{st_type.name}» успешно обновлен!")
            return redirect(reverse('flight_planning:employee_status_list') + '?tab=types')
    else:
        form = EmployeeStatusTypeForm(instance=st_type)

    return render(request, 'flight_planning/employee_statuses/status_type_form.html', {
        'form': form,
        'st_type': st_type,
        'title': f'Редактирование вида состояния: {st_type.name}'
    })


@login_required
@flight_planner_required
@require_http_methods(["POST"])
def employee_status_type_delete_view(request, pk: int):
    """Удаление вида состояния сотрудника (с защитой от удаления связанных записей).

    Args:
        request (HttpRequest): Объект HTTP-запроса.
        pk (int): Идентификатор вида состояния.

    Returns:
        HttpResponse: Перенаправление на страницу справочника.
    """
    st_type = get_object_or_404(EmployeeStatusType, pk=pk)
    if st_type.records.exists():
        msg = f"Нельзя удалить вид состояния «{st_type.name}», так как по нему есть зарегистрированные записи ({st_type.records.count()} шт.)."
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'error', 'error': msg}, status=400)
        messages.error(request, msg)
        return redirect(reverse('flight_planning:employee_status_list') + '?tab=types')

    name = st_type.name
    st_type.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'status': 'success',
            'message': f"Вид состояния «{name}» успешно удален."
        })

    messages.success(request, f"Вид состояния «{name}» успешно удален.")
    return redirect(reverse('flight_planning:employee_status_list') + '?tab=types')


@login_required
@require_http_methods(["GET"])
def get_pilot_employee_statuses_api(request, pilot_id: int):
    """API получения информации о состояниях/статусах сотрудника.

    Args:
        request (HttpRequest): HTTP-запрос.
        pilot_id (int): Идентификатор сотрудника.

    Returns:
        JsonResponse: Статус и список записей состояний сотрудника.
    """
    status_data = get_pilot_employee_statuses(pilot_id)
    return JsonResponse({'status': 'success', 'data': status_data})









