import datetime
import os
import pathlib
import time
import uuid
from gc import get_objects

from dateutil import rrule
from dateutil.relativedelta import relativedelta
from dateutil.rrule import DAILY
from django.contrib.contenttypes.fields import GenericRelation, GenericForeignKey
from django.contrib.contenttypes.models import ContentType
from django.core.mail import EmailMultiAlternatives
from django.core.validators import FileExtensionValidator
from django.db import models
from django.db.models import Q, Choices, Max
from django.db.models.expressions import result
from django.db.models.signals import post_save, pre_save
from django.dispatch import receiver
from django.template.loader import render_to_string
from django.urls import reverse
from django_ckeditor_5.fields import CKEditor5Field
from docx import Document
from docxtpl import DocxTemplate, RichText, Listing
from htmldocx import HtmlToDocx
from loguru import logger

from administration_app.utils import (
    ending_day,
    format_name_initials,
    timedelta_to_time,
    change_approval_status,
)
from contracts_app.models import CompanyProperty, Estate, TypeProperty
from customers_app.models import (
    DataBaseUser,
    Counteragent,
    HarmfulWorkingConditions,
    Division,
    Job,
    AccessLevel,
    HistoryChange,
    Affiliation,
)
from djangoProject.settings import BASE_DIR, EMAIL_HOST_USER, MEDIA_URL
from library_app.models import DocumentForm
from telegram_app.models import TelegramNotification, ChatID


# logger.add("debug.json", format=config('LOG_FORMAT'), level=config('LOG_LEVEL'),
#            rotation=config('LOG_ROTATION'), compression=config('LOG_COMPRESSION'),
#            serialize=config('LOG_SERIALIZE'))


def filename_creator(instance, filename, prefix: str, datefield, filetype: str):
    """
    Генерирует уникальное имя файла на основе переданных параметров.

    Используется для формирования имен файлов при загрузке, основываясь на префиксе,
    дате, типе файла и уникальном идентификаторе. Уникальный идентификатор вычисляется
    на основе максимального значения поля `pk` в модели.

    Args:
        instance (Model): Экземпляр модели, к которой относится файл.
        filename (str): Исходное имя файла, используется для извлечения расширения.
        prefix (str): Префикс, добавляемый к имени файла.
        datefield (Union[datetime.date, Any]): Дата или значение, используемое для
                                               формирования части имени файла.
        filetype (str): Тип файла, добавляемый к имени.

    Returns:
        str: Сформированное имя файла в формате:
             '{prefix}-{date_str}-{filetype}-{uid}.{ext}'
    """
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'

    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    if isinstance(datefield, datetime.date):
        date_str = datefield.strftime("%Y%m%d")
    else:
        date_str = str(datefield)

    return f"{prefix}-{date_str}-{filetype}-{uid}.{ext}"



# Create your models here.
def contract_directory_path(instance, filename):
    return f"hr/medical/{filename}"

# def mdc_directory_path_pmo(instance, filename):
#     prefix = "MED"
#     datefield = instance.date_entry
#     max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
#     uid = f'{max_pk:07}'
#     user_uid = f"{instance.executor.pk:07}"
#
#     custom_name = (
#         f"{prefix}-{uid}-{str(instance.working_status)}-{str(datefield)}-{user_uid}.docx"
#     )
#     return os.path.join("hr", "medical", str(user_uid), custom_name)
#
# def mdc_directory_path_po(instance, filename):
#     prefix = "MED"
#     datefield = instance.date_entry
#     max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
#     uid = f'{max_pk:07}'
#     user_uid = f"{instance.executor.pk:07}"
#
#     custom_name = (
#         f"{prefix}-{uid}-{str(instance.working_status)}-{str(datefield)}-{user_uid}PO.docx"
#     )
#     return os.path.join("hr", "medical", str(user_uid), custom_name)

# @receiver(post_save, sender=Medical)
# def rename_file_name(sender, instance, **kwargs):
#     try:
#         change = 0
#         # Формируем уникальное окончание файла. Длинна в 7 символов. В окончании номер записи: рк, спереди дополняющие нули
#         uid = f"{instance.pk:07}"
#         user_uid = f"{instance.person.pk:07}"
#         filename_pmo = (
#             f"MED-{uid}-{instance.working_status}-{instance.date_entry}-{uid}.docx"
#         )
#         filename_po = (
#             f"MED-{uid}-{instance.working_status}-{instance.date_entry}-{uid}PO.docx"
#         )
#         Med(
#             instance,
#             f"media/hr/medical/{user_uid}",
#             filename_pmo,
#             filename_po,
#             user_uid,
#         )
#         if f"hr/medical/{user_uid}/{filename_pmo}" != instance.medical_direction:
#             change = 1
#             instance.medical_direction = f"hr/medical/{user_uid}/{filename_pmo}"
#         if f"hr/medical/{user_uid}/{filename_po}" != instance.medical_direction2:
#             change = 1
#             instance.medical_direction2 = f"hr/medical/{user_uid}/{filename_po}"
#         if change == 1:
#             instance.save()
#     except Exception as _ex:
#         logger.error(f"Ошибка при переименовании файла {_ex}")



def jds_directory_path(instance, filename):
    prefix = "JDS"
    datefield = instance.document_date
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{instance.document_division.code}-"
        f"{instance.document_job.code}-{uid}-{str(datefield)}.{ext}"
    )

    return os.path.join("docs", "JDS", str(instance.document_division.code), custom_name)

def ins_directory_path(instance, filename):
    prefix = "INS"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-DRAFT-{user_uid}.{ext}"
    )
    return os.path.join("docs", "INS", str(year), custom_name)

def ins_directory_path_scan(instance, filename):
    prefix = "INS"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-SCAN-{user_uid}.{ext}"
    )
    return os.path.join("docs", "INS", str(year), custom_name)



def prv_directory_path(instance, filename):
    prefix = "PRV"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-DRAFT-{user_uid}.{ext}"
    )
    return os.path.join("docs", "PRV", str(year), custom_name)


def prv_directory_path_scan(instance, filename):
    prefix = "PRV"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-SCAN-{user_uid}.{ext}"
    )
    return os.path.join("docs", "PRV", str(year), custom_name)

def gdc_directory_path(instance, filename):
    prefix = "GDC"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-DRAFT-{user_uid}.{ext}"
    )
    return os.path.join("docs", "GDC", str(year), custom_name)


def gdc_directory_path_scan(instance, filename):
    prefix = "GDC"
    datefield = instance.date_entry
    max_pk = (instance.__class__.objects.aggregate(Max('pk'))['pk__max'] or 0) + 1
    uid = f'{max_pk:07}'
    user_uid = f"{instance.executor.pk:07}"
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]
    ext = os.path.splitext(filename)[1].lstrip('.') or 'pdf'  # fallback если расширения нет

    custom_name = (
        f"{prefix}-{uid}-{str(datefield)}-SCAN-{user_uid}.{ext}"
    )
    return os.path.join("docs", "GDC", str(year), custom_name)

def brf_directory_path_doc(instance, filename):
    prefix = "BRF"
    filetype = "DOCS"  # или подставь по логике, можно и вытянуть из instance, если надо
    datefield = instance.date_entry

    custom_name = filename_creator(instance, filename, prefix, datefield, filetype)
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]

    return os.path.join("docs", "BRF", str(year), custom_name)


def brf_directory_path_scan(instance, filename):
    prefix = "BRF"
    filetype = "SCAN"  # или подставь по логике, можно и вытянуть из instance, если надо
    datefield = instance.date_entry

    custom_name = filename_creator(instance, filename, prefix, datefield, filetype)
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]

    return os.path.join("docs", "BRF", str(year), custom_name)


def opr_directory_path_scan(instance, filename):
    prefix = "OPR"
    filetype = "SCAN"  # или подставь по логике, можно и вытянуть из instance, если надо
    datefield = instance.date_entry

    custom_name = filename_creator(instance, filename, prefix, datefield, filetype)
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]

    return os.path.join("docs", "BRF", str(year), custom_name)

def ord_directory_path(instance, filename):
    year = instance.document_date
    return f"docs/ORD/{year.year}/{filename}"


def team_directory_path(instance, filename):
    year = instance.date_create
    return f"docs/ORD/{year.year}/{year.month}/{filename}"


def outfit_directory_path(instance, filename):
    prefix = "CARD"
    filetype = "SCAN"  # или подставь по логике, можно и вытянуть из instance, если надо
    datefield = instance.outfit_card_date

    custom_name = filename_creator(instance, filename, prefix, datefield, filetype)
    year = datefield.year if isinstance(datefield, datetime.date) else str(datefield)[:4]

    return os.path.join("docs", "CARD", str(year), custom_name)


class Documents(models.Model):
    class Meta:
        abstract = True

    ref_key = models.UUIDField(
        verbose_name="Уникальный номер", default=uuid.uuid4, unique=True
    )
    # type_of_document = models.ForeignKey(TypeDocuments, verbose_name='Тип документа', on_delete=models.SET_NULL,
    #                                      null=True)
    date_entry = models.DateField(
        verbose_name="Дата ввода информации", auto_now_add=True
    )
    executor = models.ForeignKey(
        DataBaseUser,
        verbose_name="Исполнитель",
        on_delete=models.SET_NULL,
        null=True,
        related_name="%(app_label)s_%(class)s_executor",
    )
    document_date = models.DateField(
        verbose_name="Дата документа", default=datetime.datetime.now
    )
    document_name = models.CharField(
        verbose_name="Наименование документа", max_length=200, default=""
    )
    document_number = models.CharField(
        verbose_name="Номер документа", max_length=18, default=""
    )

    access = models.ForeignKey(
        AccessLevel,
        verbose_name="Уровень доступа к документу",
        on_delete=models.SET_NULL,
        null=True,
        default=5,
    )
    employee = models.ManyToManyField(
        DataBaseUser,
        verbose_name="Ответственное лицо",
        blank=True,
        related_name="%(app_label)s_%(class)s_employee",
    )
    allowed_placed = models.BooleanField(
        verbose_name="Разрешение на публикацию", default=False
    )
    validity_period_start = models.DateField(
        verbose_name="Документ действует с", blank=True, null=True
    )
    validity_period_end = models.DateField(
        verbose_name="Документ действует по", blank=True, null=True
    )
    actuality = models.BooleanField(verbose_name="Актуальность", default=False)
    previous_document = models.URLField(
        verbose_name="Примечание к предшествующему документу", blank=True
    )
    parent_document = models.ForeignKey(
        "self", verbose_name="Предшествующий документ", on_delete=models.SET_NULL, null=True, blank=True
    )
    applying_for_job = models.BooleanField(
        verbose_name="Обязательно к ознакомлению при приеме на работу", default=False
    )
    document_updated_at = models.DateTimeField(auto_now=True)

    def __str__(self):
        return f'№ {self.document_number} от {self.document_date.strftime("%d.%m.%Y")}'


class Purpose(models.Model):
    class Meta:
        verbose_name = "Цель служебной записки"
        verbose_name_plural = "Цели служебной записки"

    title = models.CharField(verbose_name="Наименование", max_length=300)

    def __str__(self):
        return self.title

    @staticmethod
    def get_absolute_url():
        return reverse("hrdepartment_app:purpose_list")

    def get_data(self):
        return {
            "pk": self.pk,
            "title": self.title,
        }


class MedicalOrganisation(models.Model):
    class Meta:
        verbose_name = "Медицинская организация"
        verbose_name_plural = "Медицинскик организации"

    ref_key = models.CharField(
        verbose_name="Уникальный номер", max_length=37, default=""
    )
    description = models.CharField(
        verbose_name="Наименование", max_length=200, default=""
    )
    ogrn = models.CharField(verbose_name="ОГРН", max_length=13, default="")
    address = models.CharField(verbose_name="Адрес", max_length=250, default="")
    email = models.CharField(verbose_name="Email", max_length=150, default="")
    phone = models.CharField(verbose_name="Телефон", max_length=150, default="")

    def __str__(self):
        return self.description

    def get_title(self):
        return self.description

    @staticmethod
    def get_absolute_url():
        return reverse("hrdepartment_app:medicalorg_list")

    def get_data(self):
        return {
            "pk": self.pk,
            "description": self.description,
            "ogrn": self.ogrn,
            "address": self.address,
        }


def Med(obj_model, filepath, filename_pmo, filename_po, request):
    inspection_type = [
        ("1", "Предварительный"),
        ("2", "Периодический"),
        ("3", "Внеплановый"),
    ]

    if obj_model.person.user_work_profile.job.division_affiliation.pk == 2:
        doc = DocxTemplate(
            pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/med.docx")
        )
    else:
        doc = DocxTemplate(
            pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/med2.docx")
        )
    doc2 = DocxTemplate(
        pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/med3.docx")
    )
    if obj_model.person.gender == "male":
        gender = "муж."
    else:
        gender = "жен."
    try:
        harmful = list()
        for items in obj_model.harmful.iterator():
            harmful.append(f"{items.code}: {items.name}")
        if obj_model.person.user_work_profile.divisions.address:
            division = str(obj_model.person.user_work_profile.divisions)
            div_address = (
                f"Адрес обособленного подразделения места производственной деятельности {division[6:]} "
                f"(далее – {obj_model.person.user_work_profile.divisions}): "
                f"{obj_model.person.user_work_profile.divisions.address}."
            )
        else:
            div_address = ""
        context = {
            "gender": gender,
            "title": next(
                x[1] for x in inspection_type if x[0] == obj_model.type_inspection
            ).lower(),
            "number": obj_model.number,
            "birthday": obj_model.person.birthday.strftime("%d.%m.%Y"),
            "division": obj_model.person.user_work_profile.divisions,
            "job": obj_model.person.user_work_profile.job,
            "FIO": obj_model.person,
            "snils": obj_model.person.user_profile.snils,
            "oms": obj_model.person.user_profile.oms,
            "status": obj_model.get_working_status_display(),
            "harmful": ", ".join(harmful),
            "organisation": obj_model.organisation,
            "ogrn": obj_model.organisation.ogrn,
            "email": obj_model.organisation.email,
            "tel": obj_model.organisation.phone,
            "address": obj_model.organisation.address,
            "div_address": div_address,
        }
        context2 = {
            "gender": gender,
            "number": obj_model.number,
            "birthday": obj_model.person.birthday.strftime("%d.%m.%Y"),
            "division": obj_model.person.user_work_profile.divisions,
            "job": obj_model.person.user_work_profile.job,
            "FIO": obj_model.person,
            "snils": obj_model.person.user_profile.snils,
            "oms": obj_model.person.user_profile.oms,
            "div_address": div_address,
        }
    except Exception as _ex:
        DataBaseUser.objects.get(pk=request)
        logger.debug(
            f"Ошибка заполнения файла {filename_pmo}: {DataBaseUser.objects.get(pk=request)} {_ex}"
        )
        context = {}
    doc.render(context)
    doc2.render(context2)
    path_obj = pathlib.Path.joinpath(pathlib.Path.joinpath(BASE_DIR, filepath))
    if not path_obj.exists():
        path_obj.mkdir(parents=True)
    doc.save(pathlib.Path.joinpath(path_obj, filename_pmo))
    doc2.save(pathlib.Path.joinpath(path_obj, filename_po))
    # ToDo: Попытка конвертации docx в pdf в Linux. Не работает
    # convert(filename, (filename[:-4]+'pdf'))
    # convert(filepath)


class Medical(models.Model):
    class Meta:
        verbose_name = "Медицинское направление"
        verbose_name_plural = "Медицинские направления"
        ordering = ["-date_entry"]

    type_of = [("1", "Поступающий на работу"), ("2", "Работающий")]

    inspection_view = [
        ("1", "Медицинский осмотр"),
        ("2", "Психиатрическое освидетельствование"),
    ]

    inspection_type = [
        ("1", "Предварительный"),
        ("2", "Периодический"),
        ("3", "Внеплановый"),
    ]

    ref_key = models.CharField(
        verbose_name="Уникальный номер", max_length=37, default=""
    )
    number = models.CharField(verbose_name="Номер", max_length=11, default="")
    person = models.ForeignKey(
        DataBaseUser, verbose_name="Сотрудник", on_delete=models.SET_NULL, null=True
    )
    date_entry = models.DateField(verbose_name="Дата ввода информации", null=True)
    date_of_inspection = models.DateField(verbose_name="Дата осмотра", null=True)
    organisation = models.ForeignKey(
        MedicalOrganisation,
        verbose_name="Медицинская организация",
        on_delete=models.SET_NULL,
        null=True,
    )
    working_status = models.CharField(
        verbose_name="Статус",
        max_length=40,
        choices=type_of,
        help_text="",
        blank=True,
        default="",
    )
    view_inspection = models.CharField(
        verbose_name="Вид осмотра",
        max_length=40,
        choices=inspection_view,
        help_text="",
        blank=True,
        default="",
    )
    type_inspection = models.CharField(
        verbose_name="Тип осмотра",
        max_length=15,
        choices=inspection_type,
        help_text="",
        blank=True,
        default="",
    )
    medical_direction = models.FileField(verbose_name="Файл ПМО", blank=True) #upload_to=contract_directory_path,
    medical_direction2 = models.FileField(verbose_name="Файл ПО", blank=True) #upload_to=contract_directory_path,
    harmful = models.ManyToManyField(
        HarmfulWorkingConditions, verbose_name="Вредные условия труда"
    )
    updated_at = models.DateTimeField(auto_now=True)

    def get_data(self):
        return {
            "pk": self.pk,
            "number": self.number,
            "date_entry": f"{self.date_entry:%d.%m.%Y} г.",
            "person": self.person.get_title(),
            "organisation": self.organisation.get_title(),
            "working_status": self.get_working_status_display(),
            "view_inspection": self.get_view_inspection_display(),
            "type_inspection": self.get_type_inspection_display(),
        }

    def __str__(self):
        return f"{self.number} {self.person}"

    def generate_med_files(self):
        uid = f"{self.pk:07}"
        user_uid = f"{self.person.pk:07}"
        filename_pmo = f"MED-{uid}-{self.working_status}-{self.date_entry}-{uid}.docx"
        filename_po = f"MED-{uid}-{self.working_status}-{self.date_entry}-{uid}PO.docx"

        # Генерация файлов
        Med(self, f"media/hr/medical/{user_uid}", filename_pmo, filename_po, user_uid)

        # Обновляем поля путей
        self.medical_direction = f"hr/medical/{user_uid}/{filename_pmo}"
        self.medical_direction2 = f"hr/medical/{user_uid}/{filename_po}"

    def save(self, *args, **kwargs):
        is_new = self.pk is None
        super().save(*args, **kwargs)  # сначала сохраняем, чтобы был pk
        self.generate_med_files()
        super().save(update_fields=["medical_direction", "medical_direction2"])  # обновляем только пути
#
#
# @receiver(post_save, sender=Medical)
# def rename_file_name(sender, instance, **kwargs):
#     try:
#         change = 0
#         # Формируем уникальное окончание файла. Длинна в 7 символов. В окончании номер записи: рк, спереди дополняющие нули
#         uid = f"{instance.pk:07}"
#         user_uid = f"{instance.person.pk:07}"
#         filename_pmo = (
#             f"MED-{uid}-{instance.working_status}-{instance.date_entry}-{uid}.docx"
#         )
#         filename_po = (
#             f"MED-{uid}-{instance.working_status}-{instance.date_entry}-{uid}PO.docx"
#         )
#         Med(
#             instance,
#             f"media/hr/medical/{user_uid}",
#             filename_pmo,
#             filename_po,
#             user_uid,
#         )
#         if f"hr/medical/{user_uid}/{filename_pmo}" != instance.medical_direction:
#             change = 1
#             instance.medical_direction = f"hr/medical/{user_uid}/{filename_pmo}"
#         if f"hr/medical/{user_uid}/{filename_po}" != instance.medical_direction2:
#             change = 1
#             instance.medical_direction2 = f"hr/medical/{user_uid}/{filename_po}"
#         if change == 1:
#             instance.save()
#     except Exception as _ex:
#         logger.error(f"Ошибка при переименовании файла {_ex}")


class PlaceProductionActivity(models.Model):
    class Meta:
        verbose_name = "Место назначения"
        verbose_name_plural = "Места назначения"

    name = models.CharField(verbose_name="Наименование", max_length=250)
    address = models.CharField(verbose_name="Адрес", max_length=250, blank=True)
    short_name = models.CharField(verbose_name="Краткое наименование", max_length=30, default="", blank=True)
    use_team_orders = models.BooleanField(verbose_name="Использовать в приказах",
                                          default=False)  # Использовать командные orders
    additional_payment = models.DecimalField(verbose_name="Дополнительная оплата", default=0, blank=True,
                                             decimal_places=2, max_digits=10)
    email = models.EmailField(verbose_name="Электронная почта", max_length=250, blank=True)
    work_email_password = models.CharField(
        verbose_name="Пароль от корпоративной почты",
        max_length=50,
        blank=True,
        default="",
    )

    def __str__(self):
        return str(self.name)

    def get_data(self):
        return {
            "pk": self.pk,
            "name": self.name,
            "address": self.address,
        }

    @staticmethod
    def get_absolute_url():
        return reverse("hrdepartment_app:place_list")


class ReasonForCancellation(models.Model):
    class Meta:
        verbose_name = "Причина отмены"
        verbose_name_plural = "Причины отмены"

    name = models.CharField(verbose_name="Наименование", max_length=250, default="")

    def __str__(self):
        return self.name

    def get_title(self):
        return str(self.name)


class OfficialMemo(models.Model):
    class Meta:
        verbose_name = "Служебная записка"
        verbose_name_plural = "Служебные записки"
        ordering = ["-date_of_creation"]

    type_of_accommodation = [("1", "Квартира"), ("2", "Гостиница")]

    type_of_trip = [("1", "Служебная поездка"), ("2", "Командировка")]

    memo_type = [
        ("1", "Направление"),
        ("2", "Продление"),
        ("3", "Без выезда"),
    ]
    document_extension = models.ForeignKey(
        "self",
        verbose_name="Документ основания",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="extension",
    )
    date_of_creation = models.DateTimeField(
        verbose_name="Дата и время создания", auto_now_add=True
    )  # При миграции указать 1 и вставить timezone.now()
    official_memo_type = models.CharField(
        verbose_name="Тип СП",
        max_length=9,
        choices=memo_type,
        help_text="",
        default="1",
    )
    person = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник",
        on_delete=models.SET_NULL,
        null=True,
        related_name="employee",
    )
    purpose_trip = models.ForeignKey(
        Purpose,
        verbose_name="Цель",
        on_delete=models.SET_NULL,
        null=True,
    )
    period_from = models.DateField(verbose_name="Дата начала", null=True)
    period_for = models.DateField(verbose_name="Дата окончания", null=True)
    place_departure = models.ForeignKey(
        PlaceProductionActivity,
        verbose_name="Место отправления",
        on_delete=models.SET_NULL,
        null=True,
        related_name="place_departure",
    )
    place_production_activity = models.ManyToManyField(
        PlaceProductionActivity,
        verbose_name="МПД",
        related_name="place_production_activity",
    )
    accommodation = models.CharField(
        verbose_name="Проживание",
        max_length=9,
        choices=type_of_accommodation,
        help_text="",
        blank=True,
        default="",
    )
    type_trip = models.CharField(
        verbose_name="Тип поездки",
        max_length=9,
        choices=type_of_trip,
        help_text="",
        blank=True,
        default="",
    )
    order = models.ForeignKey(
        "DocumentsOrder",
        verbose_name="Приказ",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
    )
    comments = models.CharField(
        verbose_name="Примечание", max_length=250, default="", blank=True
    )
    document_accepted = models.BooleanField(
        verbose_name="Документ принят", default=False
    )
    responsible = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник",
        on_delete=models.SET_NULL,
        null=True,
        related_name="responsible",
    )
    cancellation = models.BooleanField(verbose_name="Отмена", default=False)
    reason_cancellation = models.ForeignKey(
        ReasonForCancellation,
        verbose_name="Причина отмены",
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
    )
    expenses = models.BooleanField(verbose_name="Пометка выплаты", default=False)
    expenses_summ = models.DecimalField(
        verbose_name="Сумма аванса",
        default=0,
        max_digits=10,
        decimal_places=2,
    )
    history_change = GenericRelation(HistoryChange)
    title = models.CharField(
        verbose_name="Наименование", max_length=200, default="", blank=True
    )
    creation_retroactively = models.BooleanField(
        verbose_name="Документ введен задним числом", default=False
    )

    def __str__(self):
        return self.title

    def get_title(self):
        return self.title

    def get_data(self):
        place = [str(item) for item in self.place_production_activity.iterator()]
        if self.type_trip == "1":
            if self.official_memo_type == "1":
                type_trip = "СП"
            elif self.official_memo_type == "2":
                type_trip = "СП+"
            else:
                type_trip = "БВ"
        else:
            if self.official_memo_type == "1":
                type_trip = "К"
            else:
                type_trip = "К+"

        return {
            "pk": self.pk,
            "type_trip": type_trip,
            "person": str(self.person),
            "job": str(self.person.user_work_profile.job),
            "place_production_activity": "; ".join(place),
            "purpose_trip": str(self.purpose_trip),
            "period_from": f"{self.period_from:%d.%m.%Y} г.",  # .strftime(""),
            "period_for": f"{self.period_for:%d.%m.%Y} г.",  # .strftime(""),
            "accommodation": str(self.get_accommodation_display()),
            "order": str(self.order) if self.order else "",
            "comments": str(self.comments),
            "cancellation": self.cancellation,
            "document_accepted": self.document_accepted,
            "date_order": self.period_from,
            "expenses_summ": self.expenses_summ,
        }


@receiver(pre_save, sender=OfficialMemo)
def fill_title(sender, instance, **kwargs):
    if instance.official_memo_type == "1":
        type_memo = "(СП):" if instance.type_trip == "1" else "(К):"
    elif instance.official_memo_type == "2":
        type_memo = "(СП+):" if instance.type_trip == "1" else "(К+):"
    else:
        type_memo = "(БВ)"
    instance.title = f'{type_memo} {format_name_initials(instance.person) if instance.person else "None"} с {instance.period_from.strftime("%d.%m.%Y")} по {instance.period_for.strftime("%d.%m.%Y")}'


class ApprovalProcess(models.Model):
    """
    Служебная записка
    """

    class Meta:
        abstract = True

    date_of_creation = models.DateTimeField(
        verbose_name="Дата и время создания", auto_now_add=True
    )
    person_executor = models.ForeignKey(
        DataBaseUser,
        verbose_name="Исполнитель",
        on_delete=models.SET_NULL,
        null=True,
        related_name="person_executor",
    )
    submit_for_approval = models.BooleanField(
        verbose_name="Передан на согласование", default=False
    )
    comments_for_approval = models.CharField(
        verbose_name="Комментарий для согласования",
        max_length=200,
        help_text="",
        blank=True,
        default="",
    )
    person_agreement = models.ForeignKey(
        DataBaseUser,
        verbose_name="Согласующее лицо",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_agreement",
    )
    document_not_agreed = models.BooleanField(
        verbose_name="Документ согласован", default=False
    )
    reason_for_approval = models.CharField(
        verbose_name="Примечание к согласованию",
        max_length=200,
        help_text="",
        blank=True,
        default="",
    )
    person_distributor = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник НО",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_distributor",
    )
    location_selected = models.BooleanField(
        verbose_name="Выбрано место проживания", default=False
    )
    person_department_staff = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник ОК",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_department_staff",
    )
    process_accepted = models.BooleanField(verbose_name="Издан приказ", default=False)
    person_clerk = models.ForeignKey(
        DataBaseUser,
        verbose_name="Делопроизводитель",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_clerk",
    )
    originals_received = models.BooleanField(
        verbose_name="Получены оригиналы", default=False
    )
    person_hr = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник ОК",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_hr",
    )
    hr_accepted = models.BooleanField(verbose_name="Документы проверены", default=False)
    person_accounting = models.ForeignKey(
        DataBaseUser,
        verbose_name="Сотрудник Бухгалтерии",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="person_accounting",
    )
    accepted_accounting = models.BooleanField(
        verbose_name="Принято в бухгалтерии", default=False
    )
    history_change = GenericRelation(HistoryChange)


class ApprovalOficialMemoProcess(ApprovalProcess):
    """
    Бизнес-процесс служебной записки
    """

    class Meta:
        verbose_name = "Служебная записка по служебной поездке"
        verbose_name_plural = "Служебные записки по служебным поездкам"
        ordering = ["-document__period_from"]

    type_of = [("1", "Квартира"), ("2", "Гостиница")]
    # ref_key = models.CharField(default=uuid.uuid4, max_length=37, null=True, blank=True)
    document = models.OneToOneField(
        OfficialMemo,
        verbose_name="Документ",
        on_delete=models.CASCADE,
        null=True,
        related_name="docs",
    )
    accommodation = models.CharField(
        verbose_name="Проживание",
        max_length=9,
        choices=type_of,
        help_text="",
        blank=True,
        default="",
    )
    order = models.ForeignKey(
        "DocumentsOrder",
        verbose_name="Приказ",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
    )
    email_send = models.BooleanField(verbose_name="Письмо отправлено", default=False)
    cancellation = models.BooleanField(verbose_name="Отмена", default=False)
    reason_cancellation = models.ForeignKey(
        ReasonForCancellation,
        verbose_name="Причина отмены",
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
    )
    date_receipt_original = models.DateField(
        verbose_name="Дата получения", null=True, blank=True
    )
    originals_docs_comment = models.CharField(
        verbose_name="Примечание", max_length=100, help_text="", blank=True, default=""
    )
    submitted_for_signature = models.DateField(
        verbose_name="Дата передачи на подпись", null=True, blank=True
    )
    date_transfer_hr = models.DateField(
        verbose_name="Дата передачи в ОК", null=True, blank=True
    )
    number_business_trip_days = models.IntegerField(verbose_name="Дни СП", default=0)
    number_flight_days = models.IntegerField(verbose_name="Дни ЛД", default=0)
    start_date_trip = models.DateField(
        verbose_name="Дата начала по СЗ", null=True, blank=True
    )
    end_date_trip = models.DateField(
        verbose_name="Дата окончания по СЗ", null=True, blank=True
    )
    date_transfer_accounting = models.DateField(
        verbose_name="Дата передачи в бухгалтерию", null=True, blank=True
    )
    prepaid_expense = models.CharField(
        verbose_name="Пометка выплаты",
        max_length=100,
        help_text="",
        blank=True,
        default="",
    )
    prepaid_expense_summ = models.DecimalField(
        verbose_name="Сумма авансового отчета",
        default=0,
        max_digits=10,
        decimal_places=2,
    )

    def __init__(self, *args, **kwargs):
        super(ApprovalOficialMemoProcess, self).__init__(*args, **kwargs)

    def __str__(self):
        return str(self.document)

    def get_data(self):
        if self.document.official_memo_type == "3":
            location_selected = "--//--"
            process_accepted = "--//--"
        else:
            location_selected = (
                format_name_initials(self.person_distributor, self)
                if self.location_selected
                else ""
            )
            process_accepted = (
                format_name_initials(self.person_department_staff, self)
                if self.process_accepted
                else ""
            )
        return {
            "pk": self.pk,
            "document_type": "К" if self.document.type_trip == "2" else "СП",
            "document": str(self.document.title),
            "submit_for_approval": format_name_initials(self.person_executor, self) if self.submit_for_approval else "",
            "document_not_agreed": format_name_initials(self.person_agreement, self)
            if self.document_not_agreed
            else "",
            "location_selected": location_selected,
            "process_accepted": process_accepted,
            "accepted_accounting": format_name_initials(self.person_accounting, self)
            if self.accepted_accounting
            else "",
            "accommodation": str(self.get_accommodation_display()),
            "order": str(self.order) if self.order else "",
            "comments": str(self.document.comments),
            "cancellation": self.cancellation,
            "originals_received": True if self.originals_received and self.date_transfer_hr else False,
            "expenses_summ": self.document.expenses_summ if self.process_accepted and self.document.expenses_summ > 0 else "",
            "expenses_summ_check": self.document.expenses if self.process_accepted and self.document.expenses_summ > 0 else "-",
        }

    @staticmethod
    def get_absolute_url():
        return reverse("hrdepartment_app:bpmemo_list")

    def send_mail(self, title, trigger=0):
        # Отмена СП или СК
        if self.cancellation and trigger == 0:
            mail_to = self.document.person.email
            mail_to_copy_first = (
                self.person_executor.email if self.person_executor else ""
            )
            mail_to_copy_second = (
                self.person_distributor.email if self.person_distributor else ""
            )
            mail_to_copy_third = (
                self.person_department_staff.email
                if self.person_department_staff
                else ""
            )
            subject_mail = title

            current_context = {
                "title": self.document.title,
                "order_number": str(self.order.document_number)
                if self.order
                else "--//--",
                "order_date": str(self.order.document_date) if self.order else "--//--",
                "reason_cancellation": str(self.reason_cancellation),
                "person_executor": str(self.person_executor)
                if self.person_executor
                else "",
                "person_distributor": str(self.person_distributor)
                if self.person_distributor
                else "",
                "person_department_staff": str(self.person_department_staff)
                if self.person_department_staff
                else "",
                "mail_to_copy": str(self.person_executor.email)
                if self.person_executor
                else "",
            }
            logger.debug(f"Email string: {current_context}")
            text_content = render_to_string(
                "hrdepartment_app/email_cancel_bpmemo.html", current_context
            )
            html_content = render_to_string(
                "hrdepartment_app/email_cancel_bpmemo.html", current_context
            )
            first_msg = EmailMultiAlternatives(
                subject_mail,
                text_content,
                EMAIL_HOST_USER,
                [mail_to, mail_to_copy_first],
            )
            second_msg = EmailMultiAlternatives(
                subject_mail,
                text_content,
                EMAIL_HOST_USER,
                [mail_to_copy_second, mail_to_copy_third],
            )
            first_msg.attach_alternative(html_content, "text/html")
            second_msg.attach_alternative(html_content, "text/html")

            try:
                # send_mass_mail((first_msg, second_msg), fail_silently=False)
                first_msg.send()
            except Exception as _ex:
                logger.debug(
                    f"Failed to send first email to: {mail_to} {mail_to_copy_first}. {_ex}"
                )
            try:
                second_msg.send()
            except Exception as _ex:
                logger.debug(
                    f"Failed to send second email to {mail_to_copy_second} {mail_to_copy_third}. {_ex}"
                )
        if trigger == 1 or trigger == 2:
            # Повторное уведомление об СП или СК
            type_of = ["Служебная квартира", "Гостиница"]

            if self.process_accepted:
                from openpyxl import load_workbook

                delta = self.document.period_for - self.document.period_from
                try:
                    place = [
                        item.name
                        for item in self.document.place_production_activity.all()
                    ]
                except Exception as _ex:
                    place = []
                # Получаем ссылку на файл шаблона
                if (
                        self.document.person.user_work_profile.job.division_affiliation.pk
                        == 2
                ):
                    if self.document.type_trip == "2":
                        filepath_name = "spk.xlsx"
                    else:
                        filepath_name = "sp.xlsx"
                else:
                    if self.document.type_trip == "2":
                        filepath_name = "sp2k.xlsx"
                    else:
                        filepath_name = "sp2.xlsx"
                filepath = pathlib.Path.joinpath(
                    pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates"),
                    filepath_name,
                )
                wb = load_workbook(filepath)
                ws = wb.active
                ws["C3"] = str(self.document.person)
                ws["M3"] = str(self.document.person.service_number)
                ws["C4"] = str(self.document.person.user_work_profile.job)
                ws["C5"] = str(self.document.person.user_work_profile.divisions)
                ws["C6"] = "Приказ № " + str(self.order.document_number)
                ws["F6"] = self.order.document_date.strftime("%d.%m.%y")
                ws["H6"] = "на " + ending_day(int(delta.days) + 1)
                ws["L6"] = self.document.period_from.strftime("%d.%m.%y")
                ws["O6"] = self.document.period_for.strftime("%d.%m.%y")
                ws["C8"] = ", ".join(place)
                ws["C9"] = str(self.document.purpose_trip)
                ws["A90"] = (
                        str(self.person_agreement.user_work_profile.job)
                        + ", "
                        + format_name_initials(self.person_agreement)
                )

                wb.save(
                    pathlib.Path.joinpath(
                        pathlib.Path.joinpath(BASE_DIR, "media"), filepath_name
                    )
                )
                wb.close()
                if trigger == 2:
                    mail_to = self.person_executor.email
                else:
                    mail_to = self.document.person.email
                # mail_to_copy = self.person_executor.email
                type_trip = (
                    "поездку" if self.document.type_trip == "1" else "командировку"
                )
                official_memo_type = self.document.official_memo_type
                if official_memo_type == "1":
                    # Конвертируем xlsx в pdf
                    # Удалить
                    from msoffice2pdf import convert

                    source = str(
                        pathlib.Path.joinpath(
                            pathlib.Path.joinpath(BASE_DIR, "media"), filepath_name
                        )
                    )
                    output_dir = str(pathlib.Path.joinpath(BASE_DIR, "media"))
                    file_name = convert(source=source, output_dir=output_dir, soft=0)
                    subject_mail = "Направление в служебную " + type_trip
                    type_trip_title = "Вы направляетесь в служебную " + type_trip
                    type_trip_variant = "направлении в служебную " + type_trip
                    type_trip_variant_second = "направление в служебную " + type_trip
                    type_trip_extension = ""
                else:
                    subject_mail = (
                            "Продление служебной "
                            + type_trip[0:-1]
                            + "и: с "
                            + str(self.document.period_from.strftime("%d.%m.%Y"))
                            + " г. по "
                            + str(self.document.period_for.strftime("%d.%m.%Y"))
                            + " г. "
                            + str(self.document.document_extension.order)
                    )
                    type_trip_title = "Вам продлена служебная " + type_trip[0:-1] + "а"
                    type_trip_variant = "продлении служебной " + type_trip[0:-1] + "и"
                    type_trip_variant_second = (
                            "продление служебной " + type_trip[0:-1] + "и"
                    )
                    type_trip_extension = "Внимание! При продлении служебной поездки или служебной командировки, новое служебное задание не высылается. Отметки и печати о выбытии и прибытии в пункты назначения проставляются в основном служебном задании."

                if self.accommodation == "1":
                    accommodation = "Квартира"
                else:
                    accommodation = "Гостиница"
                print(str(place).strip("['']"), place)
                current_context = {
                    "greetings": "Уважаемый"
                    if self.document.person.gender == "male"
                    else "Уважаемая",
                    "person": str(self.document.person),
                    "place": ", ".join(place),
                    "type_trip": type_trip_title,
                    "type_trip_variant": type_trip_variant,
                    "type_trip_variant_second": type_trip_variant_second,
                    "type_trip_second": "поездки"
                    if self.document.type_trip == "1"
                    else "командировки",
                    "purpose_trip": str(self.document.purpose_trip),
                    "order_number": str(self.order.document_number),
                    "order_date": self.order.document_date.strftime("%d.%m.%Y"),
                    "delta": str(ending_day(int(delta.days) + 1)),
                    "period_from": self.document.period_from.strftime("%d.%m.%Y"),
                    "period_for": self.document.period_for.strftime("%d.%m.%Y"),
                    "accommodation": accommodation,
                    "person_executor": format_name_initials(self.person_executor),
                    "mail_to_copy": str(self.person_executor.email),
                    "person_distributor": format_name_initials(self.person_distributor),
                    "Year": str(datetime.datetime.today().year),
                    "type_trip_extension": type_trip_extension,
                }
                logger.debug(f"Email string: {current_context}")
                text_content = render_to_string(
                    "hrdepartment_app/email_template.html", current_context
                )
                html_content = render_to_string(
                    "hrdepartment_app/email_template.html", current_context
                )

                msg = EmailMultiAlternatives(
                    subject_mail,
                    text_content,
                    EMAIL_HOST_USER,
                    [
                        mail_to,
                    ],
                )
                msg.attach_alternative(html_content, "text/html")
                if self.document.official_memo_type == "1":
                    msg.attach_file(str(file_name))
                try:
                    res = msg.send()
                    self.email_send = True
                    self.save()
                except Exception as _ex:
                    logger.debug(f"Failed to send email. {_ex}")


def create_xlsx(instance):
    from openpyxl import load_workbook

    filepath = pathlib.Path.joinpath(MEDIA_URL, "wb.xlsx")
    wb = load_workbook(filepath)
    ws = wb.active()
    ws["C3"] = instance.document.person
    filepath2 = pathlib.Path.joinpath(MEDIA_URL, "wb-1.xlsx")
    ws.save(filepath2)


@receiver(pre_save, sender=ApprovalOficialMemoProcess)
def hr_accepted(sender, instance, **kwargs):
    obj_list = ReportCard.objects.filter(
        Q(doc_ref_key=instance.pk) & Q(employee=instance.document.person)
    )
    for item in obj_list:
        item.delete()
    if not instance.cancellation and instance.pk:
        if instance.start_date_trip and instance.end_date_trip:
            interval = list(
                rrule.rrule(
                    rrule.DAILY,
                    dtstart=instance.start_date_trip,
                    until=instance.end_date_trip,
                )
            )
        else:
            interval = list(
                rrule.rrule(
                    rrule.DAILY,
                    dtstart=instance.document.period_from,
                    until=instance.document.period_for,
                )
            )
        if len(interval) > 0:
            for date in interval:
                if instance.document.type_trip == "1":
                    record_type = "14"
                else:
                    record_type = "15"
                # Проверяем день на выходные дни и праздничные дни
                start_time, end_time, type_of_day = check_day(
                    date,
                    datetime.datetime(1, 1, 1, 9, 30).time(),
                    datetime.datetime(1, 1, 1, 18, 0).time(),
                    int(record_type)
                )
                report_kwargs = {
                    "report_card_day": date,
                    "rec_no": instance.pk + instance.document.person.pk,
                    "employee": instance.document.person,
                    "start_time": start_time,
                    "end_time": end_time,
                    "record_type": record_type,
                    "reason_adjustment": str(instance.document),
                    "doc_ref_key": instance.pk,
                    "confirmed": True if instance.hr_accepted else False,
                }
                obj, created = ReportCard.objects.update_or_create(
                    report_card_day=date,
                    doc_ref_key=instance.pk,
                    employee=instance.document.person,
                    defaults=report_kwargs,
                )
                obj.place_report_card.set(
                    instance.document.place_production_activity.all()
                )
                obj.save()


@receiver(post_save, sender=ApprovalOficialMemoProcess)
def create_report(sender, instance: ApprovalOficialMemoProcess, **kwargs):
    change_approval_status(instance)
    type_of = ["Служебная квартира", "Гостиница"]
    if (
            instance.submit_for_approval
            and not instance.document_not_agreed
            and not instance.email_send
    ):
        business_process = BusinessProcessDirection.objects.filter(
            person_executor=instance.person_executor.user_work_profile.job
        )
        person_agreement_job_list = []
        person_agreement_list = []
        for item in business_process:
            for job in item.person_agreement.all():
                person_agreement_job_list.append(job)
        for item in DataBaseUser.objects.filter(
                user_work_profile__job__name__in=set(person_agreement_job_list)
        ):
            if item.telegram_id:
                person_agreement_list.append(
                    ChatID.objects.filter(chat_id=item.telegram_id).first()
                )
        kwargs_obj = {
            "message": f"Необходимо согласовать документ: {instance.document}",
            "document_url": f"https://corp.barkol.ru/hr/bpmemo/{instance.pk}/update/",
            "document_id": f"{instance.pk}",
            "sending_counter": 3,
            "send_time": datetime.datetime.now() + relativedelta(minutes=1),
            "send_date": datetime.datetime.today(),
        }
        tn, created = TelegramNotification.objects.update_or_create(
            document_id=instance.pk, defaults=kwargs_obj
        )
        tn.respondents.set(person_agreement_list)
    if (
            instance.document_not_agreed
            and not instance.location_selected
            and not instance.email_send
            and instance.document.official_memo_type in ["1", "2"]
    ):
        person_agreement_list = []
        for item in DataBaseUser.objects.filter(
                Q(user_work_profile__divisions__type_of_role="1")
                & Q(user_work_profile__job__right_to_approval=True)
        ):
            if item.telegram_id:
                person_agreement_list.append(
                    ChatID.objects.filter(chat_id=item.telegram_id).first()
                )
        kwargs_obj = {
            "message": f"Необходимо утвердить место проживания: {instance.document}",
            "document_url": f"https://corp.barkol.ru/hr/bpmemo/{instance.pk}/update/",
            "document_id": f"{instance.pk}",
            "sending_counter": 3,
            "send_time": datetime.datetime.now() + relativedelta(minutes=1),
            "send_date": datetime.datetime.today(),
        }
        tn, created = TelegramNotification.objects.update_or_create(
            document_id=instance.pk, defaults=kwargs_obj
        )
        tn.respondents.set(person_agreement_list)
    if (
            instance.location_selected
            and not instance.process_accepted
            and not instance.email_send
    ):
        person_agreement_list = []
        for item in DataBaseUser.objects.filter(
                Q(user_work_profile__divisions__type_of_role="2")
                & Q(user_work_profile__job__right_to_approval=True)
        ):
            if item.telegram_id:
                person_agreement_list.append(
                    ChatID.objects.filter(chat_id=item.telegram_id).first()
                )
        kwargs_obj = {
            "message": f"Необходимо издать приказ: {instance.document}",
            "document_url": f"https://corp.barkol.ru/hr/bpmemo/{instance.pk}/update/",
            "document_id": f"{instance.pk}",
            "sending_counter": 3,
            "send_time": datetime.datetime.now() + relativedelta(minutes=1),
            "send_date": datetime.datetime.today(),
        }
        tn, created = TelegramNotification.objects.update_or_create(
            document_id=instance.pk, defaults=kwargs_obj
        )
        tn.respondents.set(person_agreement_list)
    if instance.process_accepted and not instance.email_send:
        tn = TelegramNotification.objects.filter(document_id=instance.pk)
        for item in tn:
            item.delete()
        from openpyxl import load_workbook

        delta = instance.document.period_for - instance.document.period_from
        try:
            place = [
                item.name for item in instance.document.place_production_activity.all()
            ]
        except Exception as _ex:
            place = []
        # Получаем ссылку на файл шаблона
        if instance.document.person.user_work_profile.job.division_affiliation.pk == 2:
            if instance.document.type_trip == "2":
                filepath_name = "spk.xlsx"
            else:
                filepath_name = "sp.xlsx"
        else:
            if instance.document.type_trip == "2":
                filepath_name = "sp2k.xlsx"
            else:
                filepath_name = "sp2.xlsx"
        filepath = pathlib.Path.joinpath(
            pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates"), filepath_name
        )
        wb = load_workbook(filepath)
        ws = wb.active
        ws["C3"] = str(instance.document.person)
        ws["M3"] = str(instance.document.person.service_number)
        ws["C4"] = str(instance.document.person.user_work_profile.job)
        ws["C5"] = str(instance.document.person.user_work_profile.divisions)
        ws["C6"] = "Приказ № " + str(instance.order.document_number)
        ws["F6"] = instance.order.document_date.strftime("%d.%m.%y")
        ws["H6"] = "на " + ending_day(int(delta.days) + 1)
        ws["L6"] = instance.document.period_from.strftime("%d.%m.%y")
        ws["O6"] = instance.document.period_for.strftime("%d.%m.%y")
        ws["C8"] = ", ".join(place)
        ws["C9"] = str(instance.document.purpose_trip)
        if instance.document.purpose_trip.title == "Дежурства на ПСР":
            ws["H86"] = ", из них ПСР"
            ws["K86"] = "__________"
        ws["A90"] = (
                str(instance.person_agreement.user_work_profile.job)
                + ", "
                + format_name_initials(instance.person_agreement)
        )

        wb.save(
            pathlib.Path.joinpath(
                pathlib.Path.joinpath(BASE_DIR, "media"), filepath_name
            )
        )
        wb.close()

        mail_to = instance.document.person.email
        mail_to_copy = instance.person_executor.email
        type_trip = "поездку" if instance.document.type_trip == "1" else "командировку"

        official_memo_type = instance.document.official_memo_type
        if official_memo_type == "1":
            # Конвертируем xlsx в pdf
            # Удалить
            from msoffice2pdf import convert

            source = str(
                pathlib.Path.joinpath(
                    pathlib.Path.joinpath(BASE_DIR, "media"), filepath_name
                )
            )
            output_dir = str(pathlib.Path.joinpath(BASE_DIR, "media"))
            file_name = convert(source=source, output_dir=output_dir, soft=0)
            subject_mail = "Направление в служебную " + type_trip
            type_trip_title = "Вы направляетесь в служебную " + type_trip
            type_trip_variant = "направлении в служебную " + type_trip
            type_trip_variant_second = "направление в служебную " + type_trip
            type_trip_extension = ""
        else:
            subject_mail = (
                    "Продление служебной "
                    + type_trip[0:-1]
                    + "и: с "
                    + str(instance.document.period_from.strftime("%d.%m.%Y"))
                    + " г. по "
                    + str(instance.document.period_for.strftime("%d.%m.%Y"))
                    + " г. Приказ:  "
                    + str(instance.document.order)
            )
            type_trip_title = "Вам продлена служебная " + type_trip[0:-1] + "а"
            type_trip_variant = "продлении служебной " + type_trip[0:-1] + "и"
            type_trip_variant_second = "продление служебной " + type_trip[0:-1] + "и"
            type_trip_extension = "Внимание! При продлении служебной поездки или служебной командировки, новое служебное задание не высылается. Отметки и печати о выбытии и прибытии в пункты назначения проставляются в основном служебном задании."

        if instance.accommodation == "1":
            accommodation = "Квартира"
        else:
            accommodation = "Гостиница"
        current_context = {
            "greetings": "Уважаемый"
            if instance.document.person.gender == "male"
            else "Уважаемая",
            "person": str(instance.document.person),
            "place": ", ".join(place),
            "type_trip": type_trip_title,
            "type_trip_variant": type_trip_variant,
            "type_trip_variant_second": type_trip_variant_second,
            "type_trip_second": "поездки"
            if instance.document.type_trip == "1"
            else "командировки",
            "purpose_trip": str(instance.document.purpose_trip),
            "order_number": str(instance.order.document_number),
            "order_date": instance.order.document_date.strftime("%d.%m.%Y"),
            "delta": str(ending_day(int(delta.days) + 1)),
            "period_from": instance.document.period_from.strftime("%d.%m.%Y"),
            "period_for": instance.document.period_for.strftime("%d.%m.%Y"),
            "accommodation": accommodation,
            "person_executor": format_name_initials(instance.person_executor),
            "mail_to_copy": str(instance.person_executor.email),
            "person_distributor": format_name_initials(instance.person_distributor),
            "Year": str(datetime.datetime.today().year),
            "type_trip_extension": type_trip_extension,
        }
        logger.debug(f"Email string: {current_context}")
        text_content = render_to_string(
            "hrdepartment_app/email_template.html", current_context
        )
        html_content = render_to_string(
            "hrdepartment_app/email_template.html", current_context
        )

        msg = EmailMultiAlternatives(
            subject_mail,
            text_content,
            EMAIL_HOST_USER,
            [
                mail_to,
                mail_to_copy,
            ],
        )
        msg.attach_alternative(html_content, "text/html")
        if instance.document.official_memo_type == "1":
            msg.attach_file(str(file_name))
        try:
            res = msg.send()
            instance.email_send = True
            instance.save()
        except Exception as _ex:
            logger.debug(f"Failed to send email. {_ex}")


class BusinessProcessDirection(models.Model):
    type_of = [("1", "Служебная поездка"), ("2", "Приказы о старших бригадах")]

    class Meta:
        verbose_name = "Направление бизнес процесса"
        verbose_name_plural = "Направления бизнес процессов"

    business_process_type = models.CharField(
        verbose_name="Тип бизнес процесса",
        max_length=5,
        default="",
        blank=True,
        choices=type_of,
    )
    person_executor = models.ManyToManyField(
        Job, verbose_name="Исполнитель", related_name="person_executor"
    )
    person_agreement = models.ManyToManyField(
        Job, verbose_name="Согласующее лицо", related_name="person_agreement"
    )
    clerk = models.ManyToManyField(
        Job, verbose_name="Делопроизводитель", related_name="clerk"
    )
    person_hr = models.ManyToManyField(
        Job, verbose_name="Сотрудник ОК", related_name="person_hr"
    )
    date_start = models.DateField(verbose_name="Дата начала", null=True, blank=True)
    date_end = models.DateField(verbose_name="Дата окончания", null=True, blank=True)

    @staticmethod
    def get_absolute_url():
        return reverse("hrdepartment_app:bptrip_list")


class OrderDescription(models.Model):
    """
    Модель, представляющая описание приказа.

    Атрибуты:
        name (str): Название приказа
        affiliation (Affiliation): Принадлежность приказа к службе

    """

    class Meta:
        verbose_name = "Наименование приказа"
        verbose_name_plural = "Наименования приказов"

    name = models.CharField(
        verbose_name="",
        max_length=250,
        blank=True,
    )
    affiliation = models.ForeignKey(
        Affiliation,
        verbose_name="Принадлежность",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
    )

    def __str__(self):
        return self.name


class DocumentsOrder(Documents):
    type_of_order = [("1", "Общая деятельность"), ("2", "Личный состав")]

    class Meta:
        verbose_name = "Приказ"
        verbose_name_plural = "Приказы"
        ordering = ["-document_date"]

    document_name = models.ForeignKey(
        OrderDescription,
        verbose_name="Наименование документа",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        default=None,
    )
    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=ord_directory_path, blank=True
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=ord_directory_path, blank=True
    )
    document_order_type = models.CharField(
        verbose_name="Тип приказа", max_length=18, choices=type_of_order
    )
    document_foundation = models.ForeignKey(
        OfficialMemo,
        verbose_name="Документ основание",
        on_delete=models.SET_NULL,
        null=True,
        blank=True,
        related_name="doc_foundation",
    )
    description = CKEditor5Field("Содержание", config_name="extends", blank=True)
    approved = models.BooleanField(verbose_name="Утверждён", default=False)
    cancellation = models.BooleanField(verbose_name="Отмена", default=False)
    reason_cancellation = models.ForeignKey(
        ReasonForCancellation,
        verbose_name="Причина отмены",
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
    )

    def get_data(self):
        status = ""
        dt = datetime.datetime.today()

        if (
                self.validity_period_end
                and datetime.date(dt.year, dt.month, dt.day) > self.validity_period_end
        ):
            status = "Действие завершил"
        else:
            status = "Действует"

        if self.cancellation:
            status = "Отменён"
        return {
            "pk": self.pk,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",  # .strftime(""),
            "document_name": self.document_name.name,
            "person": format_name_initials(self.document_foundation.person.get_title())
            if self.document_foundation
            else "",
            "approved": status,
            "cancellation": self.cancellation,
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:order_list")

    def __str__(self):
        return f'Пр. № {self.document_number} от {self.document_date.strftime("%d.%m.%Y")} г.'


def order_doc(obj_model: DocumentsOrder, filepath: str, filename: str, request):
    sub_doc_file = ""
    if obj_model.document_foundation:
        if obj_model.document_foundation.type_trip == "1":
            if (
                    "Командир воздушного судна"
                    in obj_model.document_foundation.person.user_work_profile.job.get_title()
            ):
                doc = DocxTemplate(
                    pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/aom2.docx")
                )
            else:
                doc = DocxTemplate(
                    pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/aom.docx")
                )
        else:
            doc = DocxTemplate(
                pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/aom3.docx")
            )

        delta = (
                obj_model.document_foundation.period_for
                - obj_model.document_foundation.period_from
        )
        place = [
            item.name
            for item in obj_model.document_foundation.place_production_activity.all()
        ]
        context = {
            "Number": obj_model.document_number,
            "DateDoc": f'{obj_model.document_date.strftime("%d.%m.%Y")} г.',
            "FIO": obj_model.document_foundation.person,
            "ServiceNum": obj_model.document_foundation.person.service_number,
            "Division": obj_model.document_foundation.person.user_work_profile.divisions,
            "Job": obj_model.document_foundation.person.user_work_profile.job,
            "Place": ", ".join(place),
            "DateCount": str(int(delta.days) + 1),
            "DateFrom": f'{obj_model.document_foundation.period_from.strftime("%d.%m.%Y")} г.',
            "DateFor": f'{obj_model.document_foundation.period_for.strftime("%d.%m.%Y")} г.',
            "Purpose": obj_model.document_foundation.purpose_trip,
            "DateAcquaintance": f'{obj_model.document_date.strftime("%d.%m.%Y")} г.',
        }

    else:
        doc = DocxTemplate(
            pathlib.Path.joinpath(BASE_DIR, "static/DocxTemplates/ord.docx")
        )
        sub_doc_file = pathlib.Path.joinpath(
            pathlib.Path.joinpath(BASE_DIR, filepath), f"subdoc-{filename}"
        )
        desc_document = Document()
        new_parser = HtmlToDocx()
        new_parser.add_html_to_document(obj_model.description, desc_document)
        desc_result_path = sub_doc_file
        desc_document.save(desc_result_path)
        sub_doc = doc.new_subdoc(desc_result_path)
        context = {
            "Number": obj_model.document_number,
            "DateDoc": f'{obj_model.document_date.strftime("%d.%m.%Y")} г.',
            # "Title": obj_model.document_name,
            "Description": sub_doc,
            # "Description": sub_doc,
        }

    try:
        doc.render(context, autoescape=True)
    except Exception as _ex:
        # DataBaseUser.objects.get(pk=request)
        logger.debug(f"Ошибка заполнения файла {filename}: {_ex}")
        context = {}

    path_obj = pathlib.Path.joinpath(pathlib.Path.joinpath(BASE_DIR, filepath))
    if not path_obj.exists():
        path_obj.mkdir(parents=True)
    doc.save(pathlib.Path.joinpath(path_obj, filename))
    if os.path.isfile(sub_doc_file):
        os.remove(sub_doc_file)

    from msoffice2pdf import convert

    try:
        var = convert(
            source=str(pathlib.Path.joinpath(path_obj, filename)),
            output_dir=str(path_obj),
            soft=0,
        )
        logger.debug(
            f"Файл: {str(pathlib.Path.joinpath(path_obj, filename))}, Путь: {str(path_obj)}"
        )
        return var
    except Exception as _ex:
        logger.error(f"Ошибка сохранения файла в pdf {filename}: {_ex}")


@receiver(post_save, sender=DocumentsOrder)
def rename_order_file_name(sender, instance: DocumentsOrder, **kwargs):
    if not instance.cancellation:
        # Формируем уникальное окончание файла. Длинна в 7 символов. В окончании номер записи: рк, спереди дополняющие нули

        # ext_scan = str(instance.scan_file).split('.')[-1]
        uid = f"{instance.pk:07}"
        filename = (
            f"ORD-{instance.document_order_type}-{instance.document_date}-{uid}.docx"
        )
        scanname = (
            f"ORD-{instance.document_order_type}-{instance.document_date}-{uid}.pdf"
        )
        date_doc = instance.document_date
        created_pdf = order_doc(
            instance,
            f"media/docs/ORD/{date_doc.year}/{date_doc.month}",
            filename,
            instance.document_order_type,
        )
        scan_name = pathlib.Path(created_pdf).name
        if f"docs/ORD/{date_doc.year}/{date_doc.month}/{filename}" != instance.doc_file:
            DocumentsOrder.objects.filter(pk=instance.pk).update(
                doc_file=f"docs/ORD/{date_doc.year}/{date_doc.month}/{filename}"
            )
        # if scanname != scan_name:
        try:
            pathlib.Path.rename(
                pathlib.Path.joinpath(
                    BASE_DIR,
                    "media",
                    f"docs/ORD/{date_doc.year}/{date_doc.month}",
                    scan_name,
                ),
                pathlib.Path.joinpath(
                    BASE_DIR,
                    "media",
                    f"docs/ORD/{date_doc.year}/{date_doc.month}",
                    scanname,
                ),
            )
        except Exception as _ex0:
            logger.error(f"Ошибка переименования файла: {_ex0}")
        DocumentsOrder.objects.filter(pk=instance.pk).update(
            scan_file=f"docs/ORD/{date_doc.year}/{date_doc.month}/{scanname}"
        )


class CreatingTeam(models.Model):
    class Meta:
        verbose_name = "Создание бригады"
        verbose_name_plural = "Создание бригад"
        ordering = ["-id"]

    doc_type = [
        ("0", "Новый документ"),
        ("1", "Замещающий документ"),
    ]
    document_type = models.CharField(verbose_name='Тип документа', max_length=1, choices=doc_type, default="0")
    replaceable_document = models.ForeignKey('self', verbose_name='Отменяемый документ', null=True, blank=True,
                                             on_delete=models.SET_NULL, )
    senior_brigade = models.ForeignKey(DataBaseUser, verbose_name="Старший бригады", on_delete=models.SET_NULL,
                                       null=True, related_name='senior_brigade')
    team_brigade = models.ManyToManyField(DataBaseUser, verbose_name="Состав бригады", related_name='team_brigade',
                                          blank=True)
    executor_person = models.ForeignKey(DataBaseUser, verbose_name="Исполнитель", on_delete=models.SET_NULL,
                                        null=True, related_name='executor_person')
    approving_person = models.ForeignKey(DataBaseUser, verbose_name="Согласующее лицо", on_delete=models.SET_NULL,
                                         null=True, related_name='approving_person')
    date_start = models.DateField(verbose_name="Дата начала", null=True, blank=True,
                                  default=datetime.date.today)  # Дата начала бригады.
    date_end = models.DateField(verbose_name="Дата окончания", null=True, blank=True,
                                default=datetime.date.today)  # Дата окончания бригады.
    date_create = models.DateField(verbose_name="Дата приказа", null=True, blank=True,
                                   default=datetime.date.today)  # Дата создания бригады.
    number = models.CharField(verbose_name="Номер приказа", max_length=255, blank=True, default='')
    place = models.ForeignKey(PlaceProductionActivity, verbose_name="МПД", on_delete=models.SET_NULL, null=True,
                              related_name='place')
    company_property = models.ManyToManyField(Estate, verbose_name="Задание на полет", related_name='company_property')
    agreed = models.BooleanField(verbose_name="Согласовано", default=False)
    doc_file = models.FileField(verbose_name="Файл документа", upload_to=team_directory_path, blank=True)
    scan_file = models.FileField(verbose_name="Скан документа", upload_to=team_directory_path, blank=True)
    cancellation = models.BooleanField(verbose_name="Отмена", default=False)
    email_send = models.BooleanField(verbose_name="Письмо отправлено", default=False)
    email_cancellation_send = models.BooleanField(verbose_name="Письмо от отмене отправлено", default=False)
    history_change = GenericRelation(HistoryChange)
    updated_at = models.DateTimeField(auto_now=True)

    def change_status(self, item: int, status: bool):
        match item:
            case 0:
                self.email_send = status
            case 1:
                self.email_cancellation_send = status

    def __str__(self):
        return f"{format_name_initials(self.senior_brigade)} - с: {self.date_start.strftime('%d.%m.%Y')} по: {self.date_end.strftime('%d.%m.%Y')}"

    def get_absolute_url(self):
        return reverse("hrdepartment_app:team_list")

    def get_data(self):
        status = ""
        dt = datetime.datetime.today()

        if (
                self.date_end
                and datetime.date(dt.year, dt.month, dt.day) > self.date_end
        ):
            status = "Действие завершил"
        else:
            status = "Действует"

        if self.cancellation:
            status = "Отменён"

        return {
            "pk": self.pk,
            "document_name": format_name_initials(self.senior_brigade),  # self.senior_brigade,
            "date_start": self.date_start.strftime("%d.%m.%Y"),
            "date_end": self.date_end.strftime("%d.%m.%Y"),
            "document_number": self.number,
            "document_date": f"{self.date_create:%d.%m.%Y} г.",  # .strftime(""),
            "document_division": self.place.name,
            "agreed": "Согласовано" if self.agreed else "Не согласовано",  # Согласовано, Не согласовано
            "actuality": status,
            "executor": format_name_initials(self.executor_person),
            "email_send": "Да" if self.email_send else "Нет",
        }


def ias_order(obj_model: CreatingTeam, filepath: str, filename: str, request, single=True, cancel=False):
    ordteam = "static/DocxTemplates/ord-ias-single.docx" if single else "static/DocxTemplates/ord-ias.docx"
    if cancel:
        ordteam = "static/DocxTemplates/ord-ias-single-cancel.docx" if single else "static/DocxTemplates/ord-ias-cancel.docx"
    doc = DocxTemplate(
        pathlib.Path.joinpath(BASE_DIR, ordteam)
    )
    sub_doc_file = pathlib.Path.joinpath(
        pathlib.Path.joinpath(BASE_DIR, filepath), f"subdoc-{filename}"
    )
    desc_document = Document()
    # new_parser = HtmlToDocx()
    # new_parser.add_html_to_document(obj_model.description, desc_document)
    desc_result_path = sub_doc_file
    desc_document.save(desc_result_path)
    sub_doc = doc.new_subdoc(desc_result_path)
    team_brigade_list = f"- {format_name_initials(obj_model.senior_brigade)} - {obj_model.senior_brigade.user_work_profile.job}\a"
    for item in obj_model.team_brigade.all():
        team_brigade_list += f"- {format_name_initials(item)} - {item.user_work_profile.job}\a"
    company_property = ''
    for item in obj_model.company_property.all():
        company_property += f" {item.type_property} {item},"
    context = {
        "DocNumber": '____' if obj_model.number == '' else obj_model.number,
        "DateDoc": f'{obj_model.date_create.strftime("%d.%m.%Y")} г.',
        "DateDocOrder": f'{obj_model.date_start.strftime("%d.%m.%Y")} г.',
        "DateStart": obj_model.date_start.strftime("%d.%m.%Y"),
        "DateEnd": obj_model.date_end.strftime("%d.%m.%Y"),
        "Place": obj_model.place,
        "CompanyProperty": company_property,
        "team_brigade": obj_model.senior_brigade,
        "team_brigade_job": obj_model.senior_brigade.user_work_profile.job,
        "ShortName": obj_model.place,
        "team_brigade_list": Listing(f"{team_brigade_list[:-1]}"),
        "additional_payment": str(obj_model.place.additional_payment),
        "number_order": obj_model.replaceable_document.number if obj_model.replaceable_document else '',
        "date_order": f'{obj_model.replaceable_document.date_create.strftime("%d.%m.%Y")} г.' if obj_model.replaceable_document else '',
    }
    doc.render(context, autoescape=True)
    path_obj = pathlib.Path.joinpath(pathlib.Path.joinpath(BASE_DIR, filepath))
    if not path_obj.exists():
        path_obj.mkdir(parents=True)
    doc.save(pathlib.Path.joinpath(path_obj, filename))
    if os.path.isfile(sub_doc_file):
        os.remove(sub_doc_file)
    # from msoffice2pdf import convert
    #
    # try:
    #     var = convert(
    #         source=str(pathlib.Path.joinpath(path_obj, filename)),
    #         output_dir=str(path_obj),
    #         soft=0,
    #     )
    #     logger.debug(
    #         f"Файл: {str(pathlib.Path.joinpath(path_obj, filename))}, Путь: {str(path_obj)}"
    #     )
    #     return var
    # except Exception as _ex:
    #     logger.error(f"Ошибка сохранения файла в pdf {filename}: {_ex}")


@receiver(post_save, sender=CreatingTeam)
def rename_ias_order_file_name(sender, instance: CreatingTeam, **kwargs):
    if instance.agreed:
        tn = TelegramNotification.objects.filter(document_id=instance.pk)
        if tn:
            tn.delete()
    if instance.agreed and instance.number != '':
        # Формируем уникальное окончание файла. Длинна в 7 символов. В окончании номер записи: рк, спереди дополняющие нули
        # ext_scan = str(instance.scan_file).split('.')[-1]
        uid = f"{instance.pk:07}"
        filename = (f"ORD-3-{instance.date_create}-{uid}.docx")
        scanname = (f"ORD-3-{instance.date_create}-{uid}.pdf")
        date_doc = instance.date_create
        single = True if instance.team_brigade.count() < 1 else False
        cancel = True if instance.replaceable_document else False
        created_pdf = ias_order(instance, f"media/docs/ORD/{date_doc.year}/{date_doc.month}", filename, '3', single,
                                cancel)
        # scan_name = pathlib.Path(created_pdf).name
        if f"docs/ORD/{date_doc.year}/{date_doc.month}/{filename}" != instance.doc_file:
            CreatingTeam.objects.filter(pk=instance.pk).update(
                doc_file=f"docs/ORD/{date_doc.year}/{date_doc.month}/{filename}"
            )
        if instance.scan_file:
            if f"docs/ORD/{date_doc.year}/{date_doc.month}/{scanname}" != instance.scan_file.name:
                try:
                    pathlib.Path.rename(
                        pathlib.Path.joinpath(
                            BASE_DIR,
                            "media",
                            instance.scan_file.name,
                        ),
                        pathlib.Path.joinpath(
                            BASE_DIR,
                            "media",
                            f"docs/ORD/{date_doc.year}/{date_doc.month}",
                            scanname,
                        ),
                    )
                except Exception as _ex0:
                    logger.error(f"Ошибка переименования файла: {_ex0}")
                CreatingTeam.objects.filter(pk=instance.pk).update(
                    scan_file=f"docs/ORD/{date_doc.year}/{date_doc.month}/{scanname}"
                )
            print(instance.scan_file.name)
        #
        # if not instance.email_send:

    else:
        business_process = BusinessProcessDirection.objects.filter(
            person_executor=instance.executor_person.user_work_profile.job
        )
        person_agreement_job_list = []
        person_agreement_list = []
        for item in business_process:
            for job in item.person_agreement.all():
                person_agreement_job_list.append(job)
        for item in DataBaseUser.objects.filter(
                user_work_profile__job__name__in=set(person_agreement_job_list)
        ):
            if item.telegram_id:
                person_agreement_list.append(
                    ChatID.objects.filter(chat_id=item.telegram_id).first()
                )
        kwargs_obj = {
            "message": f"Необходимо согласовать документ: {instance}",
            "document_url": f"https://corp.barkol.ru/hr/team/{instance.pk}/agreed/",
            "document_id": f"{instance.pk}",
            "sending_counter": 3,
            "send_time": datetime.datetime.now() + relativedelta(minutes=1),
            "send_date": datetime.datetime.today(),
        }
        tn, created = TelegramNotification.objects.update_or_create(
            document_id=instance.pk, defaults=kwargs_obj
        )
        tn.respondents.set(person_agreement_list)


class DocumentsJobDescription(Documents):
    class Meta:
        verbose_name = "Должностная инструкция"
        verbose_name_plural = "Должностные инструкции"
        ordering = ("-document_date",)

    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=jds_directory_path, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['doc', 'docx']),
        ]
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=jds_directory_path, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['pdf']),
        ]
    )
    document_division = models.ForeignKey(
        Division, verbose_name="Подразделение", on_delete=models.SET_NULL, null=True
    )
    document_job = models.ForeignKey(
        Job, verbose_name="Должность", on_delete=models.SET_NULL, null=True
    )
    document_order = models.ForeignKey(
        DocumentsOrder, verbose_name="Приказ", on_delete=models.SET_NULL, null=True
    )

    def get_data(self):
        if (DocumentsJobDescription.objects.filter(parent_document=self.pk).count() == 0) and (self.actuality):
            get_actual = 0
        else:
            get_actual = 1

        return {
            "pk": self.pk,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",  # .strftime(""),
            "document_job": str(self.document_job),
            "document_division": str(self.document_division),
            "document_order": str(self.document_order),
            "actuality": "Да" if get_actual == 0 else "Нет",
            "executor": str(self.executor),
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:jobdescription_list")

    def __str__(self):
        return f'ДИ {self.document_name} №{self.document_number} от {self.document_date.strftime("%d.%m.%Y")}'

@receiver(pre_save, sender=DocumentsJobDescription)
def delete_old_file_on_change_jds(sender, instance, **kwargs):
    if not instance.pk:
        return  # новый объект

    try:
        old_instance = DocumentsJobDescription.objects.get(pk=instance.pk)
    except DocumentsJobDescription.DoesNotExist:
        return

    # Список полей, которые нужно сравнивать и очищать
    file_fields = ['doc_file', 'scan_file']

    for field in file_fields:
        old_file = getattr(old_instance, field)
        new_file = getattr(instance, field)

        if old_file and old_file.name != getattr(new_file, 'name', None):
            try:
                if os.path.isfile(old_file.path):
                    os.remove(old_file.path)
            except Exception as e:
                print(f"Ошибка удаления старого файла в поле {field}: {e}")


class TimeSheet(models.Model):
    """
    Табель учета рабочего времени
    """
    date = models.DateField(verbose_name="Дата табеля", null=True, blank=True)
    employee = models.ForeignKey(
        DataBaseUser, verbose_name="Ответственный", on_delete=models.SET_NULL, null=True, blank=True
    )
    time_sheets_place = models.ForeignKey(
        PlaceProductionActivity, verbose_name="МПД", on_delete=models.SET_NULL, null=True, blank=True,
        related_name="time_sheets_place")
    notes = models.TextField(verbose_name="Примечания", blank=True)

    class Meta:
        verbose_name = "Табель учета рабочего времени"
        verbose_name_plural = "Табели учета рабочего времени"
        ordering = ("-date",)

    def __str__(self):
        return f"Табель от {self.date} для {self.employee}"

    def get_absolute_url(self):
        return reverse('hrdepartment_app:timesheet', kwargs={'pk': self.pk})

    def get_data(self):
        """
        Получает данные из экземпляра ReportCard.

        :return: словарь, содержащий следующие данные:
            - "pk": первичный ключ экземпляра ReportCard.
            - "employee": форматированные инициалы имени сотрудника.
            - "report_card_day": день табеля в формате "ДД.ММ.ГГГГ"
            - "start_time": время начала в формате "ЧЧ:ММ"
            - "end_time": время окончания в формате "ЧЧ:ММ"
            - "reason_adjustment": причина корректировки.
            - "record_type": отображение типа записи.
        """
        return {
            "pk": self.pk,
            "date": f"{self.date:%d.%m.%Y} г.",  # .strftime(''),
            "employee": format_name_initials(self.employee.title),
            "time_sheets_place": self.time_sheets_place.name,
            "notes": self.notes,
        }


class OperationalWork(models.Model):
    """
    Оперативные работы
    """

    class Meta:
        verbose_name = "Оперативная работа"
        verbose_name_plural = "Оперативные работы"
        ordering = ("name",)

    name = models.TextField(verbose_name="Наименование", blank=True)
    code = models.TextField(verbose_name="Код", blank=True)
    description = models.TextField(verbose_name="Описание", blank=True)
    air_bord_type = models.ForeignKey(TypeProperty, verbose_name="Тип", on_delete=models.SET_NULL, null=True,
                                      blank=True)

    def __str__(self):
        return self.code


class PeriodicWork(models.Model):
    """
    Периодические работы
    """

    class Meta:
        verbose_name = "Периодическая работа"
        verbose_name_plural = "Периодические работы"
        ordering = ("name",)

    name = models.TextField(verbose_name="Наименование", blank=True)
    code = models.TextField(verbose_name="Код", blank=True)
    ratio = models.FloatField(verbose_name="Норма-часы", blank=True)
    description = models.TextField(verbose_name="Описание", blank=True)
    air_bord_type = models.ForeignKey(TypeProperty, verbose_name="Тип", on_delete=models.SET_NULL, null=True,
                                      blank=True)

    def __str__(self):
        return self.code


class OutfitCard(models.Model):
    """
        Атрибуты:
        _________
        outfit_card_date: Дата;
        employee = Ответственный сотрудник;
        outfit_card_place = МПД;
        operational_work = Оперативные работы;
        periodic_work = Периодические работы;
        air_board = Воздушный борт;
        other_work = Другие работы;
        notes = Примечания;
        """

    class Meta:
        verbose_name = "Карта-наряд"
        verbose_name_plural = "Карты-наряды"
        ordering = ("-outfit_card_date",)

    outfit_card_date = models.DateField(verbose_name="Дата", null=True, blank=True)  # Дата
    outfit_card_number = models.CharField(verbose_name="Номер", max_length=20, default="", blank=True)
    employee = models.ForeignKey(DataBaseUser, verbose_name="Ответственный", on_delete=models.SET_NULL, null=True,
                                 blank=True)
    outfit_card_place = models.ForeignKey(PlaceProductionActivity, verbose_name="МПД", on_delete=models.SET_NULL,
                                          null=True, blank=True, related_name="outfit_card_place")
    air_board = models.ForeignKey(Estate, verbose_name="Воздушный борт", related_name='company_air_board',
                                  blank=True, on_delete=models.SET_NULL, null=True, )
    operational_work = models.ManyToManyField(OperationalWork, verbose_name="Оперативные работы",
                                              related_name="outfit_card_operational", blank=True)
    periodic_work = models.ManyToManyField(PeriodicWork, verbose_name="Периодические работы",
                                           related_name="outfit_card_periodic", blank=True)
    other_work = models.CharField(verbose_name="Другие работы", max_length=200, default="", blank=True)
    outfit_card_date_end = models.DateField(verbose_name="Дата окончания", null=True, blank=True)
    scan_document = models.FileField(verbose_name="Скан документа", upload_to=outfit_directory_path, null=True,
                                     blank=True)
    notes = models.TextField(verbose_name="Примечания", blank=True)
    updated_at = models.DateTimeField(auto_now=True)

    def __str__(self):
        return self.outfit_card_number

    def get_workers(self):
        workers = ReportCard.objects.filter(outfit_card=self)
        return ', '.join(set([format_name_initials(worker.employee.title) for worker in workers]))

    def get_works(self):
        works = [item.code for item in self.periodic_work.all()] + [item.code for item in self.operational_work.all()]
        if self.other_work:
            works.append(self.other_work)
        return ', '.join(works)

    def get_data(self):
        self.get_workers()
        return {
            "pk": self.pk,
            "outfit_card_date": f"{self.outfit_card_date:%d.%m.%Y} г.",  # .strftime(''),
            "air_board": f"{self.air_board.type_property} {self.air_board.registration_number}",
            "works": self.get_works(),
            "outfit_card_number": self.outfit_card_number,
            "workers": self.get_workers(),
            "employee": format_name_initials(self.employee.title),
        }


class ReportCard(models.Model):
    """
    Атрибуты:
    _________
    report_card_day: Дата;
    rec_no = Номер записи;
    employee = Сотрудник;
    start_time = Время прихода;
    end_time = Время ухода';
    record_type = Тип записи;
    manual_input = Ручной ввод;
    reason_adjustment = Причина ручной корректировки;
    doc_ref_key = Уникальный номер документа;
    current_intervals = Текущий интервал;
    timesheet = Табель учета рабочего времени;
    lunch_time = Время обеда;
    flight_hours = Летные часы;
    operational_work = Оперативные работы;
    periodic_work = Периодические работы;
    air_board = Воздушный борт;
    additional_work = Дополнительные работы;
    other_work = Другие работы;
    sign_report_card = Признак табеля рабочего времени;
    """

    type_of_report = [
        ("1", "Явка"),
        ("2", "Ежегодный"),
        ("3", "Дополнительный ежегодный отпуск"),
        ("4", "Отпуск за свой счет"),
        ("5", "Дополнительный учебный отпуск (оплачиваемый)"),
        ("6", "Отпуск по уходу за ребенком"),
        ("7", "Дополнительный неоплачиваемый отпуск пострадавшим в аварии на ЧАЭС"),
        ("8", "Отпуск по беременности и родам"),
        ("9", "Отпуск без оплаты согласно ТК РФ"),
        ("10", "Дополнительный отпуск"),
        ("11", "Дополнительный оплачиваемый отпуск пострадавшим в "),
        ("12", "Основной"),
        ("13", "Ручной ввод"),
        ("14", "Служебная поездка"),
        ("15", "Командировка"),
        ("16", "Больничный"),
        ("17", "Мед осмотр"),
        ("18", "График отпусков"),
        ("19", "Отпуск на санаторно курортное лечение"),
        ("20", "Отгул"),
    ]

    class Meta:
        verbose_name = "Рабочее время"
        verbose_name_plural = "Табель учета"
        ordering = ("-report_card_day",)

    report_card_day = models.DateField(verbose_name="Дата", null=True, blank=True)
    rec_no = models.IntegerField(verbose_name="Номер записи", default=0, blank=True)
    employee = models.ForeignKey(
        DataBaseUser, verbose_name="Сотрудник", on_delete=models.SET_NULL, null=True, blank=True
    )
    start_time = models.TimeField(verbose_name="Время прихода", null=True, blank=True)
    end_time = models.TimeField(verbose_name="Время ухода", null=True, blank=True)
    record_type = models.CharField(
        verbose_name="Тип записи",
        max_length=100,
        choices=type_of_report,
        default="",
        blank=True,
    )
    manual_input = models.BooleanField(verbose_name="Ручной ввод", default=False)
    reason_adjustment = models.TextField(
        verbose_name="Причина ручной корректировки", blank=True
    )
    doc_ref_key = models.CharField(
        verbose_name="Уникальный номер документа", max_length=37, default="", blank=True
    )
    current_intervals = models.BooleanField(
        verbose_name="Текущий интервал", default=True
    )
    confirmed = models.BooleanField(verbose_name="Подтвержденная СП", default=False)
    place_report_card = models.ManyToManyField(
        PlaceProductionActivity, verbose_name="МПД", related_name="place_report_card", blank=True
    )
    timesheet = models.ForeignKey(
        TimeSheet, verbose_name="Табель учета рабочего времени", on_delete=models.CASCADE, related_name="report_cards",
        null=True, blank=True
    )
    lunch_time = models.IntegerField(verbose_name="Время обеда", null=True, blank=True)
    flight_hours = models.IntegerField(verbose_name="Летные часы", null=True, blank=True)
    outfit_card = models.ManyToManyField(
        OutfitCard, verbose_name="Оперативные работы", related_name="outfit_card_report_card", blank=True
    )
    additional_work = models.CharField(verbose_name="Дополнительные работы", max_length=200, default="", blank=True)

    sign_report_card = models.BooleanField(verbose_name="Признак табеля рабочего времени", default=False)

    def get_data(self):
        """
        Получает данные из экземпляра ReportCard.

        :return: словарь, содержащий следующие данные:
            - "pk": первичный ключ экземпляра ReportCard.
            - "employee": форматированные инициалы имени сотрудника.
            - "report_card_day": день табеля в формате "ДД.ММ.ГГГГ"
            - "start_time": время начала в формате "ЧЧ:ММ"
            - "end_time": время окончания в формате "ЧЧ:ММ"
            - "reason_adjustment": причина корректировки.
            - "record_type": отображение типа записи.
        """
        return {
            "pk": self.pk,
            "employee": format_name_initials(self.employee.title),
            "report_card_day": f"{self.report_card_day:%d.%m.%Y} г.",  # .strftime(''),
            "start_time": f"{self.start_time:%H:%M}",  # .strftime(''),
            "end_time": f"{self.end_time:%H:%M}",  # .strftime(''),
            "reason_adjustment": self.reason_adjustment,
            "record_type": self.get_record_type_display(),
        }

    def __str__(self):
        return f"{self.employee}: {self.report_card_day} : {self.record_type}"


class PreHolidayDay(models.Model):
    """
    Обозначает предпраздничный день.

    Атрибуты:
        preholiday_day (DateField): дата предпраздничного дня.
        work_time (TimeField): рабочее время предпраздничного дня.
    """

    class Meta:
        verbose_name = "Предпраздничный день"
        verbose_name_plural = "Предпраздничные дни"
        ordering = ["-preholiday_day"]

    preholiday_day = models.DateField(verbose_name="Дата", null=True, blank=True)
    work_time = models.TimeField(verbose_name="Рабочее время", null=True, blank=True)

    def __str__(self):
        return str(self.preholiday_day)


class WeekendDay(models.Model):
    """
    Праздничные дни и выходные дни в связи с праздником
    Атрибуты:
    weekend_day - Дата, description - Описание, weekend_type - Тип дня (1 - Праздник, 2 - Выходной)
    """

    type_of_weekend = [
        ("1", "Праздник"),
        ("2", "Выходной"),
    ]

    class Meta:
        verbose_name = "Праздничный день"
        verbose_name_plural = "Праздничные дни"
        ordering = ["-weekend_day"]

    weekend_day = models.DateField(verbose_name="Дата", null=True, blank=True)
    description = models.CharField(
        verbose_name="Описание", max_length=200, default="", blank=True
    )
    weekend_type = models.CharField(
        verbose_name="Тип дня",
        max_length=8,
        choices=type_of_weekend,
        blank=True,
        null=True,
    )

    def __str__(self):
        return str(self.weekend_day)


# def get_norm_time_at_custom_day(day, trigger=False, type_of_day=None):
#     """
#     Подсчет количества рабочих часов в указанном дне
#     Если переменная trigger=False, то возвращается количество рабочих часов в указанном дне, иначе - возвращается тип
#     дня (П - Праздник, В - Выходной, НБ - Не было на работе)
#     Если передается переменная type_of_day, то если она равна 14 и 15, то возвращается обычный рабочий день
#     """
#     preholiday_day_count = PreHolidayDay.objects.filter(preholiday_day=day)
#     weekend_day_count = WeekendDay.objects.filter(weekend_day=day).count()
#     if weekend_day_count == 0:
#         if preholiday_day_count.count() > 0:
#             if trigger:
#                 return 'Не было на работе'
#             return preholiday_day_count[0].work_time.hour * 3600 + preholiday_day_count[0].work_time.minute * 60
#         else:
#             if day.weekday() == 4:
#                 if trigger:
#                     return 'Не было на работе'
#                 return 27000
#             elif day.weekday() < 4:
#                 if trigger:
#                     return 'Не было на работе'
#                 return 30600
#             else:
#                 if trigger:
#                     return 'Выходной'
#                 if type_of_day in [14, 15]:
#                     return 30600
#                 return 0
#     else:
#         if trigger:
#             return 'Праздничный день'
#         if type_of_day in [14, 15]:
#             if preholiday_day_count.count() > 0:
#                 return preholiday_day_count[0].work_time.hour * 3600 + preholiday_day_count[0].work_time.minute * 60
#             else:
#                 return 30600
#         if type_of_day == None:
#             return 30600
#         # return 0
def get_norm_time_at_custom_day(day, trigger=False, type_of_day=0):
    """
    Подсчет количества рабочих часов в указанном дне
    Если переменная trigger=False, то возвращается количество рабочих часов в указанном дне, иначе - возвращается тип
    дня (П - Праздник, В - Выходной, НБ - Не было на работе)
    Если передается переменная type_of_day, то если она равна 14 и 15, то возвращается обычный рабочий день
    """
    # Константы для времени в секундах
    WORK_DAY_TIME = 30600  # 8 часов 30 минут
    SHORT_DAY_TIME = 27000  # 7 часов 30 минут
    # Проверка на праздничный день
    if WeekendDay.objects.filter(weekend_day=day).exists():
        if trigger:
            return 'Праздничный день'
        return WORK_DAY_TIME if type_of_day in [14, 15] else 0

    # Проверка на предпраздничный день
    preholiday_day = PreHolidayDay.objects.filter(preholiday_day=day).first()
    if preholiday_day:
        preholiday_result = preholiday_day.work_time.hour * 3600 + preholiday_day.work_time.minute * 60
        if trigger:
            return 'Не было на работе'
        return -preholiday_result if type_of_day == 100 else preholiday_result

    # Проверка на обычный рабочий день
    if day.weekday() < 5:  # Понедельник - Пятница
        weekday_result = SHORT_DAY_TIME if day.weekday() == 4 else WORK_DAY_TIME
        if trigger:
            return 'Не было на работе'
        return -weekday_result if type_of_day == 100 else weekday_result

    # Выходной день
    if trigger:
        return 'Выходной'
    return WORK_DAY_TIME if type_of_day in [14, 15] else 0

class ProductionCalendar(models.Model):
    """
    Месяц в производственном календаре.
    Атрибуты:
    ____________
    calendar_month - Месяц, number_calendar_days - Количество календарных дней,
    number_working_days - Количество рабочих дней,
    number_days_off_and_holidays - Количество выходных и празднечных дней, description - Описание
    Методы:
    ____________
    get_friday_count - Подсчитывает количество пятниц в месяце
    get_norm_time - Подсчет количества рабочих часов в месяце
    """

    class Meta:
        verbose_name = "Месяц в производственом календаре"
        verbose_name_plural = "Производственный календарь"
        ordering = ["-calendar_month"]

    calendar_month = models.DateField(verbose_name="Месяц", null=True, blank=True)
    number_calendar_days = models.PositiveIntegerField(
        verbose_name="Количество календарных дней", default=0, null=True, blank=True
    )
    number_working_days = models.PositiveIntegerField(
        verbose_name="Количество рабочих дней", default=0, null=True, blank=True
    )
    number_days_off_and_holidays = models.PositiveIntegerField(
        verbose_name="Количество выходных и празднечных дней",
        default=0,
        null=True,
        blank=True,
    )
    description = models.CharField(
        verbose_name="Описание", max_length=200, default="", blank=True
    )

    def get_friday_count(self):
        """
        Подсчитывает количество пятниц в месяце
        :return: количество пятниц
        """
        first_day = self.calendar_month + relativedelta(day=1)
        last_day = self.calendar_month + relativedelta(day=31)
        friday = 0
        for item in range(first_day.day, last_day.day + 1):
            date_obj = first_day + datetime.timedelta(days=item - 1)
            if date_obj.weekday() == 4:
                if WeekendDay.objects.filter(weekend_day=date_obj).count() == 0:
                    friday += 1
        return friday

    def get_norm_time_on_day(self):
        last_day = datetime.datetime.today() - datetime.timedelta(days=1)
        first_day = self.calendar_month.replace(day=1)
        # Инициализация переменных
        total_hours = 0
        friday_count = 0
        preholiday_time = 0
        # Получаем все дни в диапазоне
        days_in_month = list(rrule.rrule(DAILY, dtstart=first_day, until=last_day))
        # Получаем предпраздничные и выходные дни из базы данных за один запрос
        preholiday_days = {
            item.preholiday_day: item.work_time
            for item in PreHolidayDay.objects.filter(preholiday_day__in=days_in_month)
        }
        weekend_days = {
            item.weekend_day
            for item in WeekendDay.objects.filter(weekend_day__in=days_in_month)
        }
        # Итерация по дням месяца
        for day in days_in_month:
            day_date = day.date()

            if day_date in weekend_days:
                continue  # Пропускаем выходные и праздничные дни

            if day_date in preholiday_days:
                # Обработка предпраздничных дней
                work_time = preholiday_days[day_date]
                preholiday_time += work_time.hour + work_time.minute / 60
            else:
                # Обработка обычных рабочих дней
                if day.weekday() < 5:  # Понедельник-пятница
                    total_hours += 8.5
                    if day.weekday() == 4:  # Пятница
                        friday_count += 1

        # Корректировка нормы времени
        norm_time = total_hours - friday_count + preholiday_time
        return norm_time

    def get_norm_time_at_day(self):
        """
        Подсчет количества рабочих часов в день
        :return: количество рабочих часов в день
        """
        first_day = self.calendar_month + relativedelta(day=1)
        last_day = self.calendar_month + relativedelta(day=datetime.datetime.today().day)
        last_day = last_day - datetime.timedelta(days=1)
        preholiday_time, day_count, preholiday_day_count, friday_count = 0, 0, 0, 0
        preholiday_day_list = [item.preholiday_day for item in PreHolidayDay.objects.filter(
            preholiday_day__in=list(rrule.rrule(rrule.DAILY, dtstart=first_day, until=last_day)))]
        weekday_day_list = [item.weekend_day for item in WeekendDay.objects.filter(
            weekend_day__in=list(rrule.rrule(rrule.DAILY, dtstart=first_day, until=last_day)))]

        for days in rrule.rrule(rrule.DAILY, dtstart=first_day, until=last_day):
            if datetime.date(days.year, days.month, days.day) in preholiday_day_list:
                preholiday_day_count += 1

                day = PreHolidayDay.objects.get(preholiday_day=days)
                if day.preholiday_day.weekday() == 4:
                    friday_count += 1
                    preholiday_time += day.work_time.hour + 1 + day.work_time.minute / 60
                else:
                    preholiday_time += day.work_time.hour + day.work_time.minute / 60
            elif datetime.date(days.year, days.month, days.day) in weekday_day_list:
                pass
            else:
                if days.weekday() == 4:
                    friday_count += 1
                    day_count += 1
                elif days.weekday() < 4:
                    day_count += 1

        return (day_count * 8) + (day_count / 2) - friday_count - (preholiday_day_count * 8.5 - preholiday_time)

    def get_norm_time(self):
        """
        Подсчет количества рабочих часов в месяце
        :return: количество рабочих часов в месяце
        """
        first_day = self.calendar_month + relativedelta(day=1)
        last_day = self.calendar_month + relativedelta(day=31)
        preholiday_time = 0
        preholiday_day_count = 0

        preholiday_day = PreHolidayDay.objects.filter(
            preholiday_day__in=list(rrule.rrule(rrule.DAILY, dtstart=first_day, until=last_day)))

        for item in preholiday_day:
            preholiday_day_count += 1
            if item.preholiday_day.weekday() == 4:
                preholiday_time += item.work_time.hour + 1 + item.work_time.minute / 60
            else:
                preholiday_time += item.work_time.hour + item.work_time.minute / 60
        norm_time = (self.number_working_days * 8) + (self.number_working_days / 2) - self.get_friday_count() - (
                preholiday_day_count * 8.5 - preholiday_time)

        return norm_time

    def __str__(self):
        return str(self.calendar_month)


def check_day(date: datetime.date, time_start: datetime.time, time_end: datetime.time, days_type=0):
    """
    Функция определяющая время начала и окончания рабочего дня на заданную дату
    :param date: дата
    :param time_start: время начала
    :param time_end: время окончания
    :return: три значения: время начала, время окончания и тип дня (Р - рабочий, В - выходной, П - праздник)
    """
    type_of_day = ""
    weekend = WeekendDay.objects.filter(weekend_day=date.date()).exists()
    preholiday = PreHolidayDay.objects.filter(preholiday_day=date.date()).exists()
    check_time_end = time_end
    check_time_start = time_start
    if not weekend:
        if date.weekday() in [0, 1, 2, 3]:
            if not preholiday:
                check_time_end = datetime.timedelta(
                    hours=time_end.hour, minutes=time_end.minute
                )
            else:
                preholiday_time = PreHolidayDay.objects.get(preholiday_day=date.date())
                check_time_end = datetime.timedelta(
                    hours=time_start.hour, minutes=time_start.minute
                ) + datetime.timedelta(
                    hours=preholiday_time.work_time.hour,
                    minutes=preholiday_time.work_time.minute,
                )
            type_of_day = "Р"
        elif date.weekday() == 4:
            if not preholiday:
                check_time_end = datetime.timedelta(
                    hours=time_end.hour, minutes=time_end.minute
                ) - datetime.timedelta(hours=1)
            else:
                preholiday_time = PreHolidayDay.objects.get(preholiday_day=date.date())
                check_time_end = datetime.timedelta(
                    hours=time_start.hour, minutes=time_start.minute
                ) + datetime.timedelta(
                    hours=preholiday_time.work_time.hour,
                    minutes=preholiday_time.work_time.minute,
                )
            type_of_day = "Р"
        else:
            if days_type in [14, 15]:
                check_time_end = datetime.timedelta(
                    hours=time_end.hour, minutes=time_end.minute
                )
            else:
                check_time_end = datetime.timedelta(hours=0, minutes=0)
                check_time_start = datetime.timedelta(hours=0, minutes=0)
            type_of_day = "В"
    else:
        if days_type in [14, 15]:
            print(type_of_day, "Сработало")
            check_time_end = datetime.timedelta(
                hours=time_end.hour, minutes=time_end.minute
            )
        else:
            check_time_end = datetime.timedelta(hours=0, minutes=0)
            check_time_start = datetime.timedelta(hours=0, minutes=0)
        type_of_day = "П"

    return (
        timedelta_to_time(check_time_start),
        timedelta_to_time(check_time_end),
        type_of_day,
    )


class TypesUserworktime(models.Model):
    """
    url = /odata/standard.odata/Catalog_ВидыИспользованияРабочегоВремени?$format=application/json;odata=nometadata
    """

    class Meta:
        verbose_name = "Вид использования рабочего времени"
        verbose_name_plural = "Виды использования рабочего времени"

    ref_key = models.CharField(
        verbose_name="Уникальный номер", max_length=37, default=""
    )  # поле в 1с: Ref_Key
    description = models.CharField(
        verbose_name="Наименование", max_length=150, default=""
    )  # поле в 1с: Description
    letter_code = models.CharField(
        verbose_name="Буквенный код", max_length=5, default=""
    )  # поле в 1с: БуквенныйКод
    active = models.BooleanField(
        verbose_name="Используется", default=False
    )  # поле в 1с: БуквенныйКод

    def __str__(self):
        return self.description


class Instructions(Documents):
    class Meta:
        verbose_name = "Инструкция"
        verbose_name_plural = "Инструкции"

    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=ins_directory_path, blank=True
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=ins_directory_path_scan, blank=True
    )
    storage_location_division = models.ForeignKey(
        Division,
        verbose_name="Подразделение где хранится оригинал",
        on_delete=models.SET_NULL,
        null=True,
        related_name="instruction_location_division",
    )
    document_division = models.ManyToManyField(
        Division,
        verbose_name="Подразделения",
        related_name="instruction_document_division",
    )
    document_order = models.ForeignKey(
        DocumentsOrder, verbose_name="Приказ", on_delete=models.SET_NULL, null=True
    )
    document_form = models.ManyToManyField(
        DocumentForm, verbose_name="Бланки документов"
    )

    def get_data(self):
        """
        Return a dictionary representing the data of the object.

        :return: A dictionary with the following keys:
                 - "pk": The primary key of the object.
                 - "document_name": The name of the document.
                 - "document_number": The number of the document.
                 - "document_date": The formatted date of the document ("dd.mm.yyyy г.").
                 - "document_division": The storage location division of the document (converted to string).
                 - "document_order": The document order (converted to string).
                 - "actuality": The actuality of the document ("Да" if true, "Нет" if false).
                 - "executor": The formatted name initials of the executor.

        """
        return {
            "pk": self.pk,
            "document_name": self.document_name,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",
            "document_division": str(self.storage_location_division),
            "document_order": str(self.document_order),
            "actuality": "Да" if self.actuality else "Нет",
            "executor": format_name_initials(self.executor),
        }

    # def get_absolute_url(self):
    #     return reverse('hrdepartment_app:instructions_list')

    def __str__(self):
        """
        Returns the string representation of the object.

        :return: The string representation of the object.
        :rtype: str
        """
        return self.document_name

@receiver(pre_save, sender=Instructions)
def delete_old_file_on_change_ins(sender, instance, **kwargs):
    if not instance.pk:
        return  # новый объект

    try:
        old_instance = Instructions.objects.get(pk=instance.pk)
    except Instructions.DoesNotExist:
        return

    # Список полей, которые нужно сравнивать и очищать
    file_fields = ['doc_file', 'scan_file']

    for field in file_fields:
        old_file = getattr(old_instance, field)
        new_file = getattr(instance, field)

        if old_file and old_file.name != getattr(new_file, 'name', None):
            try:
                if os.path.isfile(old_file.path):
                    os.remove(old_file.path)
            except Exception as e:
                print(f"Ошибка удаления старого файла в поле {field}: {e}")


class Provisions(Documents):
    class Meta:
        verbose_name = "Положение"
        verbose_name_plural = "Положения"
        ordering = ['-document_date']

    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=prv_directory_path, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['doc', 'docx']),
        ]
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=prv_directory_path_scan, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['pdf']),
        ]
    )
    storage_location_division = models.ForeignKey(
        Division,
        verbose_name="Подразделение где хранится оригинал",
        on_delete=models.SET_NULL,
        null=True,
        related_name="provisions_location_division",
    )
    document_division = models.ManyToManyField(
        Division,
        verbose_name="Подразделения",
        related_name="provisions_document_division",
    )
    document_order = models.ForeignKey(
        DocumentsOrder, verbose_name="Приказ", on_delete=models.SET_NULL, null=True
    )
    document_form = models.ManyToManyField(
        DocumentForm, verbose_name="Бланки документов"
    )

    def get_data(self):
        try:
            get_date = False if self.validity_period_end < datetime.date.today() else True
        except TypeError:
            get_date = True
        get_actual = False if Provisions.objects.filter(parent_document=self.pk).count() > 0 else True
        return {
            "pk": self.pk,
            "document_name": self.document_name,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",
            "document_division": str(self.storage_location_division),
            "document_order": str(self.document_order),
            "actuality": "Да" if (get_actual and get_date) else "Нет",
            "executor": format_name_initials(self.executor),
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:provisions_list")

    def __str__(self):
        return f"{self.document_name} № {self.document_number} от {self.document_date.strftime('%d.%m.%Y')}"

    def get_scan_file_url(self):
        if not self.scan_file:
            return ""
        return f"{self.scan_file.url}?v={int(time.time())}"

@receiver(pre_save, sender=Provisions)
def delete_old_file_on_change_prv(sender, instance, **kwargs):
    if not instance.pk:
        return  # новый объект

    try:
        old_instance = Provisions.objects.get(pk=instance.pk)
    except Provisions.DoesNotExist:
        return

    # Список полей, которые нужно сравнивать и очищать
    file_fields = ['doc_file', 'scan_file']

    for field in file_fields:
        old_file = getattr(old_instance, field)
        new_file = getattr(instance, field)

        if old_file and old_file.name != getattr(new_file, 'name', None):
            try:
                if os.path.isfile(old_file.path):
                    os.remove(old_file.path)
            except Exception as e:
                print(f"Ошибка удаления старого файла в поле {field}: {e}")

class Briefings(Documents):
    class Meta:
        verbose_name = "Инструктаж"
        verbose_name_plural = "Инструктажи"
        ordering = ['-document_date']

    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=brf_directory_path_doc, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['doc', 'docx']),
        ]
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=brf_directory_path_scan, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['pdf']),
        ]
    )
    storage_location_division = models.ForeignKey(
        Division,
        verbose_name="Подразделение где хранится оригинал",
        on_delete=models.SET_NULL,
        null=True,
        related_name="briefings_location_division",
    )
    document_division = models.ManyToManyField(
        Division,
        verbose_name="Подразделения",
        related_name="briefings_document_division",
    )
    document_order = models.ForeignKey(
        DocumentsOrder, verbose_name="Приказ", on_delete=models.SET_NULL, null=True
    )
    document_form = models.ManyToManyField(
        DocumentForm, verbose_name="Бланки документов"
    )

    def get_data(self):
        try:
            get_date = False if self.validity_period_end < datetime.date.today() else True
        except TypeError:
            get_date = True
        get_actual = False if Briefings.objects.filter(parent_document=self.pk).count() > 0 else True
        return {
            "pk": self.pk,
            "document_name": self.document_name,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",
            "document_division": str(self.storage_location_division),
            "document_order": str(self.document_order),
            "actuality": "Да" if (get_actual and get_date) else "Нет",
            "executor": format_name_initials(self.executor),
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:briefings_list")

    def __str__(self):
        return f"{self.document_name} № {self.document_number} от {self.document_date.strftime('%d.%m.%Y')}"


class Operational(Documents):
    class Meta:
        verbose_name = "Нормативный акт"
        verbose_name_plural = "Нормативные акты"
        ordering = ['-document_date']

    access = None
    employee = None
    previous_document = None
    document_date = models.DateField(verbose_name="Дата документа", default=datetime.datetime.now, null=True, blank=True)
    document_number = models.CharField(verbose_name="Номер документа", max_length=18, default="", null=True, blank=True)

    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=opr_directory_path_scan, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['pdf']),
        ]
    )
    storage_location_division = models.ForeignKey(
        Division,
        verbose_name="Подразделение где хранится оригинал",
        on_delete=models.SET_NULL,
        null=True,
        related_name="operational_location_division",
    )
    document_division = models.ManyToManyField(
        Division,
        verbose_name="Подразделения",
        related_name="operational_document_division",
    )

    def get_data(self):
        try:
            get_date = False if self.validity_period_end < datetime.date.today() else True
        except TypeError:
            get_date = True
        get_actual = False if Operational.objects.filter(parent_document=self.pk).count() > 0 else True
        return {
            "pk": self.pk,
            "document_name": self.document_name,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",
            "document_division": str(self.storage_location_division),
            "actuality": "Да" if (get_actual and get_date) else "Нет",
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:operational_list")

    def __str__(self):
        return f"{self.document_name} № {self.document_number} от {self.document_date.strftime('%d.%m.%Y')}"


class GuidanceDocuments(Documents):
    class Meta:
        verbose_name = "Руководящий документ"
        verbose_name_plural = "Руководящие документы"

    doc_file = models.FileField(
        verbose_name="Файл документа", upload_to=gdc_directory_path, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['doc', 'docx']),
        ]
    )
    scan_file = models.FileField(
        verbose_name="Скан документа", upload_to=gdc_directory_path_scan, blank=True,
        validators=[
            FileExtensionValidator(allowed_extensions=['pdf']),
        ]
    )
    storage_location_division = models.ForeignKey(
        Division,
        verbose_name="Подразделение где хранится оригинал",
        on_delete=models.SET_NULL,
        null=True,
        related_name="guidance_documents_location_division",
    )
    document_division = models.ManyToManyField(
        Division,
        verbose_name="Подразделения",
        related_name="guidance_documents_document_division",
    )
    document_order = models.ForeignKey(
        DocumentsOrder, verbose_name="Приказ", on_delete=models.SET_NULL, null=True
    )

    def get_data(self):
        return {
            "pk": self.pk,
            "document_name": self.document_name,
            "document_number": self.document_number,
            "document_date": f"{self.document_date:%d.%m.%Y} г.",
            "document_division": str(self.storage_location_division),
            "document_order": str(self.document_order),
            "actuality": "Да" if self.actuality else "Нет",
            "executor": format_name_initials(self.executor),
        }

    def get_absolute_url(self):
        return reverse("hrdepartment_app:guidance_documents_list")

    def __str__(self):
        return self.document_name

@receiver(pre_save, sender=GuidanceDocuments)
def delete_old_file_on_change_gdc(sender, instance, **kwargs):
    if not instance.pk:
        return  # новый объект

    try:
        old_instance = GuidanceDocuments.objects.get(pk=instance.pk)
    except GuidanceDocuments.DoesNotExist:
        return

    # Список полей, которые нужно сравнивать и очищать
    file_fields = ['doc_file', 'scan_file']

    for field in file_fields:
        old_file = getattr(old_instance, field)
        new_file = getattr(instance, field)

        if old_file and old_file.name != getattr(new_file, 'name', None):
            try:
                if os.path.isfile(old_file.path):
                    os.remove(old_file.path)
            except Exception as e:
                print(f"Ошибка удаления старого файла в поле {field}: {e}")

class DocumentAcknowledgment(models.Model):
    class Meta:
        verbose_name = "Подтверждение документа"
        verbose_name_plural = "Подтверждения документов"

    document_type = models.ForeignKey(
        ContentType,
        on_delete=models.CASCADE,
        limit_choices_to={'model__in': (
        'documentsorder', 'creatingteam', 'documentsjobdescription', 'provisions', 'guidancedocuments', 'companyevent')}
    )
    document_id = models.PositiveIntegerField()
    document = GenericForeignKey('document_type', 'document_id')
    user = models.ForeignKey(DataBaseUser, on_delete=models.CASCADE, related_name='document_acknowledgments')
    acknowledgment_date = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return f"{self.user.username} ознакомился с документом '{self.document}'"

class DataBaseUserEvent(models.Model):
    class Meta:
        verbose_name = 'Отметка'
        verbose_name_plural = 'Отметки'
        indexes = [
            models.Index(fields=['person', 'date_marks']),
            models.Index(fields=['checked']),
        ]

    person = models.ForeignKey(DataBaseUser, on_delete=models.CASCADE, verbose_name='Сотрудник')
    date_marks = models.DateField(verbose_name='Дата отметки', null=True, blank=True)
    place = models.ForeignKey(PlaceProductionActivity, on_delete=models.SET_NULL, null=True, verbose_name='Место')
    checked = models.BooleanField(default=False, verbose_name='Проверено')
    road = models.BooleanField(default=False, verbose_name='В дороге')
    created_at = models.DateTimeField(auto_now_add=True, verbose_name='Дата создания')
    updated_at = models.DateTimeField(auto_now=True, verbose_name='Дата обновления')

    def __str__(self):
        return f"{self.person} - {self.date_marks} - {self.place}"

    def get_absolute_url(self):
        return reverse("hrdepartment_app:users_events_list")

    def get_data(self):
        return {
            "pk": self.pk,
            "date_marks": f"{self.date_marks:%d.%m.%Y} г.",
            "place": str(self.place),
            "checked": "Да" if self.checked else "Нет",
            "road": "Да" if self.road else "Нет",
            "created_at": f"{self.created_at:%d.%m.%Y} г.",
            "updated_at": f"{self.updated_at:%d.%m.%Y} г.",
            "executor": format_name_initials(self.person),
        }
