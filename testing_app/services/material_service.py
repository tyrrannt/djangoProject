"""Сервисный слой для работы с лекционными материалами, видеолекциями и учетом обращений сотрудников."""

import csv
import io
from typing import Dict, Any, Optional
from datetime import datetime

from django.conf import settings
from django.db import models, transaction
from django.db.models import F, Q, Sum, Count
from django.utils import timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from testing_app.models import LectureMaterial, VideoLecture, MaterialViewLog


def log_material_access(user, material, ip_address: Optional[str] = None) -> None:
    """Фиксирует факт обращения сотрудника к лекционному или видеоматериалу.

    Если сотрудник обращается к материалу впервые, создается новая запись в журнале.
    При повторных обращениях обновляется счетчик просмотров, дата последнего визита и IP.

    Args:
        user (settings.AUTH_USER_MODEL): Пользователь (сотрудник).
        material (Union[LectureMaterial, VideoLecture]): Объект лекции или видео.
        ip_address (Optional[str]): IP-адрес клиента (если доступен).
    """
    if not user or not user.is_authenticated:
        return

    now = timezone.now()

    if isinstance(material, LectureMaterial):
        mat_type = MaterialViewLog.MaterialType.LECTURE
        filter_kwargs = {"user": user, "lecture": material}
        defaults = {
            "material_type": mat_type,
            "views_count": 1,
            "last_ip": ip_address,
        }
    elif isinstance(material, VideoLecture):
        mat_type = MaterialViewLog.MaterialType.VIDEO
        filter_kwargs = {"user": user, "video_lecture": material}
        defaults = {
            "material_type": mat_type,
            "views_count": 1,
            "last_ip": ip_address,
        }
    else:
        return

    with transaction.atomic():
        log_obj, created = MaterialViewLog.objects.select_for_update().get_or_create(
            **filter_kwargs,
            defaults=defaults
        )
        if not created:
            log_obj.views_count = F("views_count") + 1
            log_obj.last_ip = ip_address or log_obj.last_ip
            log_obj.last_viewed_at = now
            log_obj.save(update_fields=["views_count", "last_ip", "last_viewed_at"])


def get_material_dashboard_stats() -> Dict[str, Any]:
    """Формирует сводную статистику по изучению материалов для дашборда тестирования.

    Returns:
        Dict[str, Any]: Словарь с показателями:
            - total_lectures: Всего лекций
            - active_lectures: Актуальных лекций
            - total_videos: Всего видеолекций
            - active_videos: Актуальных видеолекций
            - total_views: Суммарное число обращений
            - unique_users_count: Количество уникальных сотрудников, изучавших материалы
            - recent_views: Список последних 5 обращений
    """
    total_lectures = LectureMaterial.objects.count()
    active_lectures = LectureMaterial.objects.filter(is_actual=True).count()

    total_videos = VideoLecture.objects.count()
    active_videos = VideoLecture.objects.filter(is_actual=True).count()

    views_agg = MaterialViewLog.objects.aggregate(
        total_views=Sum("views_count"),
        unique_users=Count("user", distinct=True)
    )

    recent_views = (
        MaterialViewLog.objects.select_related(
            "user",
            "lecture",
            "video_lecture"
        )
        .order_by("-last_viewed_at")[:5]
    )

    return {
        "total_lectures": total_lectures,
        "active_lectures": active_lectures,
        "total_videos": total_videos,
        "active_videos": active_videos,
        "total_views": views_agg["total_views"] or 0,
        "unique_users_count": views_agg["unique_users"] or 0,
        "recent_views": recent_views,
    }


def get_material_access_report_qs(
    material_type: Optional[str] = None,
    material_id: Optional[str] = None,
    search_query: Optional[str] = None,
    division_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> models.QuerySet:
    """Возвращает отфильтрованный QuerySet журнала обращений сотрудников к материалам.

    Args:
        material_type (Optional[str]): Фильтр по типу: 'lecture', 'video' или пустая строка/None (все).
        material_id (Optional[str]): Идентификатор конкретного материала.
        search_query (Optional[str]): Строка поиска по ФИО или имени пользователя.
        division_id (Optional[str]): Идентификатор подразделения сотрудника.
        date_from (Optional[str]): Дата начала интервала (YYYY-MM-DD).
        date_to (Optional[str]): Дата окончания интервала (YYYY-MM-DD).

    Returns:
        models.QuerySet: QuerySet записей MaterialViewLog с оптимизированными связями.
    """
    qs = MaterialViewLog.objects.select_related(
        "user",
        "lecture",
        "video_lecture",
    ).prefetch_related(
        "user__user_work_profile__job",
        "user__user_work_profile__divisions",
    )

    if material_type in [MaterialViewLog.MaterialType.LECTURE, MaterialViewLog.MaterialType.VIDEO]:
        qs = qs.filter(material_type=material_type)

    if material_id:
        try:
            m_id = int(material_id)
            if material_type == MaterialViewLog.MaterialType.LECTURE:
                qs = qs.filter(lecture_id=m_id)
            elif material_type == MaterialViewLog.MaterialType.VIDEO:
                qs = qs.filter(video_lecture_id=m_id)
            else:
                qs = qs.filter(Q(lecture_id=m_id) | Q(video_lecture_id=m_id))
        except (ValueError, TypeError):
            pass

    if search_query:
        query = search_query.strip()
        qs = qs.filter(
            Q(user__last_name__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__username__icontains=query)
            | Q(lecture__title__icontains=query)
            | Q(video_lecture__title__icontains=query)
        )

    if division_id:
        try:
            d_id = int(division_id)
            qs = qs.filter(user__user_work_profile__divisions_id=d_id)
        except (ValueError, TypeError):
            pass

    if date_from:
        try:
            d_from = datetime.strptime(date_from, "%Y-%m-%d").date()
            qs = qs.filter(last_viewed_at__date__gte=d_from)
        except (ValueError, TypeError):
            pass

    if date_to:
        try:
            d_to = datetime.strptime(date_to, "%Y-%m-%d").date()
            qs = qs.filter(last_viewed_at__date__lte=d_to)
        except (ValueError, TypeError):
            pass

    return qs.order_by("-last_viewed_at")


def export_material_report_excel(queryset: models.QuerySet) -> openpyxl.Workbook:
    """Генерирует стилизованную Excel-книгу с отчетом об обращениях к материалам.

    Args:
        queryset (models.QuerySet): Отфильтрованный набор записей MaterialViewLog.

    Returns:
        openpyxl.Workbook: Сформированная рабочая книга Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал обращений"

    # Стили
    title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="000000")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # Шапка отчета
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="ОТЧЕТ ОБ ОБРАЩЕНИЯХ СОТРУДНИКОВ К ЛЕКЦИОННЫМ И ВИДЕОМАТЕРИАЛАМ")
    title_cell.font = title_font
    title_cell.alignment = align_left
    ws.row_dimensions[1].height = 30

    gen_date = timezone.now().strftime("%d.%m.%Y %H:%M")
    ws.merge_cells("A2:H2")
    sub_cell = ws.cell(row=2, column=1, value=f"Сформирован: {gen_date} | Всего записей: {queryset.count()}")
    sub_cell.font = Font(name="Calibri", size=9, italic=True, color="64748B")
    ws.row_dimensions[2].height = 20

    headers = [
        "№",
        "Сотрудник (ФИО)",
        "Должность",
        "Подразделение",
        "Тип материала",
        "Наименование материала",
        "Первое обращение",
        "Последнее обращение",
        "Всего обращений",
        "IP-адрес",
    ]

    ws.row_dimensions[4].height = 26
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    row_idx = 5
    for item in queryset:
        user = item.user
        fio = user.get_full_name() if user else "—"
        work_prof = getattr(user, "user_work_profile", None)
        job_title = str(work_prof.job) if work_prof and work_prof.job else "—"
        division_title = str(work_prof.divisions) if work_prof and work_prof.divisions else "—"

        mat_type = item.get_material_type_display()
        mat_title = item.lecture.title if item.lecture else (item.video_lecture.title if item.video_lecture else "—")
        first_date = timezone.localtime(item.first_viewed_at).strftime("%d.%m.%Y %H:%M") if item.first_viewed_at else "—"
        last_date = timezone.localtime(item.last_viewed_at).strftime("%d.%m.%Y %H:%M") if item.last_viewed_at else "—"

        row_data = [
            row_idx - 4,
            fio,
            job_title,
            division_title,
            mat_type,
            mat_title,
            first_date,
            last_date,
            item.views_count,
            item.last_ip or "—",
        ]

        ws.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            if col_idx in [1, 5, 7, 8, 10]:
                cell.alignment = align_center
            elif col_idx == 9:
                cell.alignment = align_right
            else:
                cell.alignment = align_left

        row_idx += 1

    # Автоподбор ширины колонок
    col_widths = {
        1: 6,
        2: 30,
        3: 25,
        4: 25,
        5: 22,
        6: 35,
        7: 18,
        8: 18,
        9: 16,
        10: 16,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"
    return wb


def export_material_report_csv(queryset: models.QuerySet) -> str:
    """Генерирует CSV-строку (UTF-8 с BOM) с отчетом об обращениях к материалам.

    Args:
        queryset (models.QuerySet): Отфильтрованный набор записей MaterialViewLog.

    Returns:
        str: Содержимое файла CSV в формате UTF-8.
    """
    output = io.StringIO()
    # Записываем BOM для корректного открытия в Microsoft Excel
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "№",
        "Сотрудник (ФИО)",
        "Должность",
        "Подразделение",
        "Тип материала",
        "Наименование материала",
        "Первое обращение",
        "Последнее обращение",
        "Всего обращений",
        "IP-адрес",
    ])

    for idx, item in enumerate(queryset, start=1):
        user = item.user
        fio = user.get_full_name() if user else "—"
        work_prof = getattr(user, "user_work_profile", None)
        job_title = str(work_prof.job) if work_prof and work_prof.job else "—"
        division_title = str(work_prof.divisions) if work_prof and work_prof.divisions else "—"

        mat_type = item.get_material_type_display()
        mat_title = item.lecture.title if item.lecture else (item.video_lecture.title if item.video_lecture else "—")
        first_date = timezone.localtime(item.first_viewed_at).strftime("%d.%m.%Y %H:%M") if item.first_viewed_at else "—"
        last_date = timezone.localtime(item.last_viewed_at).strftime("%d.%m.%Y %H:%M") if item.last_viewed_at else "—"

        writer.writerow([
            idx,
            fio,
            job_title,
            division_title,
            mat_type,
            mat_title,
            first_date,
            last_date,
            item.views_count,
            item.last_ip or "—",
        ])

    return output.getvalue()
