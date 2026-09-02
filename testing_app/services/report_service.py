"""Сервисы выгрузки итоговых протоколов проверки знаний в форматы Microsoft Excel (.xlsx) и CSV."""

import io
import csv
from typing import Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from testing_app.models import Testing, TestingAssignment, TestingAttempt


def generate_testing_protocol_excel(testing: Testing) -> openpyxl.Workbook:
    """Генерирует официальный итоговый протокол проверки знаний в формате Excel (.xlsx).

    Включает шапку с реквизитами приказа авиакомпании, реестр всех аттестуемых работников
    со снимками должностей, набранными баллами, номерами сертификатов, условным
    форматированием статусов сдачи и подписями аттестационной комиссии.

    Args:
        testing (Testing): Мероприятие проверки знаний.

    Returns:
        openpyxl.Workbook: Стилизованная рабочая книга Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Протокол проверки знаний"
    ws.views.sheetView[0].showGridLines = True

    # Стили оформления
    title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="475569")
    bold_font = Font(name="Calibri", size=9.5, bold=True, color="000000")
    reg_font = Font(name="Calibri", size=9.5, color="1E293B")

    hdr_font = Font(name="Calibri", size=9.5, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    passed_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    passed_font = Font(name="Calibri", size=9.5, bold=True, color="166534")

    failed_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    failed_font = Font(name="Calibri", size=9.5, bold=True, color="991B1B")

    pending_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    pending_font = Font(name="Calibri", size=9.5, color="64748B")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # 1. Шапка документа
    ws.merge_cells("A1:J1")
    ws["A1"] = "ООО «АВИАКОМПАНИЯ «БАРКОЛ» • ИНЖЕНЕРНО-АВИАЦИОННАЯ СЛУЖБА"
    ws["A1"].font = subtitle_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = "ИТОГОВЫЙ ПРОТОКОЛ ЗАСЕДАНИЯ КОМИССИИ ПО ПРОВЕРКЕ ЗНАНИЙ ПЕРСОНАЛА"
    ws["A2"].font = title_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:J3")
    order_str = f"Приказ №{testing.order_number} от {testing.order_date.strftime('%d.%m.%Y')} «{testing.order_name}»"
    ws["A3"] = f"Мероприятие: «{testing.title}» | {order_str}"
    ws["A3"].font = bold_font
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:J4")
    dates_str = f"Период проведения: с {testing.start_datetime.strftime('%d.%m.%Y %H:%M')} по {testing.end_datetime.strftime('%d.%m.%Y %H:%M')}"
    params_str = f"Параметры теста: {testing.questions_count} вопр., проходной балл {testing.passing_score_percentage}%, таймер {testing.attempt_duration_minutes} мин"
    ws["A4"] = f"{dates_str} | {params_str}"
    ws["A4"].font = subtitle_font
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 26
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 10  # пустая строка-разделитель

    # 2. Заголовки таблицы
    headers = [
        "№ п/п",
        "ФИО работника",
        "Должность на момент приказа",
        "Подразделение",
        "Группа аттестации",
        "Попыток",
        "Итоговый балл (%)",
        "Решение комиссии",
        "Номер сертификата",
        "Дата сдачи"
    ]

    header_row = 6
    ws.row_dimensions[header_row].height = 28
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=text)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 3. Данные по назначениям
    assignments = testing.assignments.select_related(
        "employee", "group"
    ).prefetch_related(
        "attempts"
    ).order_by("group__code", "assigned_division_title", "employee__last_name")

    start_data_row = 7
    current_row = start_data_row

    total_count = assignments.count()
    passed_count = 0
    failed_count = 0
    pending_count = 0

    for idx, assign in enumerate(assignments, 1):
        # Поиск успешной попытки или последней завершенной
        best_attempt = assign.attempts.filter(is_passed=True).order_by("-score_percentage").first()
        if not best_attempt:
            best_attempt = assign.attempts.filter(status=TestingAttempt.Status.COMPLETED).order_by("-id").first()

        score_val = f"{best_attempt.score_percentage}%" if best_attempt else "—"
        cert_num = best_attempt.result_number if (best_attempt and best_attempt.is_passed) else "—"
        date_passed = best_attempt.finished_at.strftime("%d.%m.%Y %H:%M") if best_attempt else "—"

        if assign.status == TestingAssignment.Status.PASSED:
            decision_text = "СДАНО"
            row_fill = passed_fill
            row_font = passed_font
            passed_count += 1
        elif assign.status in [TestingAssignment.Status.FAILED, TestingAssignment.Status.ON_CONTROL]:
            decision_text = "НЕ СДАНО"
            row_fill = failed_fill
            row_font = failed_font
            failed_count += 1
        else:
            decision_text = "В ПРОЦЕССЕ" if assign.attempts_used > 0 else "НЕ ПРИСТУПАЛ"
            row_fill = pending_fill
            row_font = pending_font
            pending_count += 1

        row_values = [
            idx,
            assign.employee.get_full_name(),
            assign.assigned_job_title,
            assign.assigned_division_title or "—",
            assign.group.name,
            f"{assign.attempts_used} / {testing.max_attempts}",
            score_val,
            decision_text,
            cert_num,
            date_passed
        ]

        ws.row_dimensions[current_row].height = 24
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in [1, 6, 7, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = row_fill
                cell.font = row_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_idx != 8:
                cell.font = reg_font

        current_row += 1

    # 4. Итоговая сводка
    summary_row = current_row + 1
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=10)
    pass_pct = round((passed_count / float(total_count)) * 100, 1) if total_count > 0 else 0.0
    ws.cell(
        row=summary_row,
        column=1,
        value=f"ИТОГО: Всего подлежало аттестации: {total_count} чел. | Сдали проверку: {passed_count} ({pass_pct}%) | Не сдали: {failed_count} | Не завершили/не приступали: {pending_count}"
    ).font = bold_font

    # 5. Блок подписей комиссии
    sig_row = summary_row + 3
    ws.cell(row=sig_row, column=2, value="Председатель аттестационной комиссии:").font = bold_font
    ws.cell(row=sig_row, column=5, value="____________________ / ____________________ /").font = subtitle_font

    ws.cell(row=sig_row + 2, column=2, value="Члены аттестационной комиссии:").font = bold_font
    ws.cell(row=sig_row + 2, column=5, value="____________________ / ____________________ /").font = subtitle_font
    ws.cell(row=sig_row + 3, column=5, value="____________________ / ____________________ /").font = subtitle_font

    # Автоширина колонок
    col_widths = [8, 30, 28, 22, 32, 14, 18, 18, 24, 20]
    for idx, w in enumerate(col_widths, 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A7"
    return wb


def generate_testing_protocol_csv(testing: Testing) -> str:
    """Формирует итоговый протокол тестирования в текстовом формате CSV.

    Использует точку с запятой ';' в качестве разделителя и кодировку UTF-8 с BOM
    для корректного отображения русских символов в Excel.

    Args:
        testing (Testing): Мероприятие проверки знаний.

    Returns:
        str: Содержимое CSV файла.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    # Шапка приказа
    writer.writerow(["ООО «АВИАКОМПАНИЯ «БАРКОЛ»", "ПРОТОКОЛ ПРОВЕРКИ ЗНАНИЙ ПЕРСОНАЛА"])
    writer.writerow(["Мероприятие", testing.title])
    writer.writerow(["Основание", f"Приказ №{testing.order_number} от {testing.order_date.strftime('%d.%m.%Y')} «{testing.order_name}»"])
    writer.writerow([])

    # Колонки
    writer.writerow([
        "№ п/п",
        "ФИО работника",
        "Должность",
        "Подразделение",
        "Группа",
        "Попыток использовано",
        "Итоговый балл (%)",
        "Решение",
        "Номер сертификата",
        "Дата сдачи"
    ])

    assignments = testing.assignments.select_related(
        "employee", "group"
    ).prefetch_related(
        "attempts"
    ).order_by("employee__last_name")

    for idx, assign in enumerate(assignments, 1):
        best_attempt = assign.attempts.filter(is_passed=True).first()
        if not best_attempt:
            best_attempt = assign.attempts.filter(status=TestingAttempt.Status.COMPLETED).order_by("-id").first()

        score_val = f"{best_attempt.score_percentage}%" if best_attempt else "—"
        cert_num = best_attempt.result_number if (best_attempt and best_attempt.is_passed) else "—"
        date_passed = best_attempt.finished_at.strftime("%d.%m.%Y %H:%M") if best_attempt else "—"
        status_label = assign.get_status_display()

        writer.writerow([
            idx,
            assign.employee.get_full_name(),
            assign.assigned_job_title,
            assign.assigned_division_title or "—",
            assign.group.name,
            assign.attempts_used,
            score_val,
            status_label,
            cert_num,
            date_passed
        ])

    return output.getvalue()
