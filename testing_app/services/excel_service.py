"""Сервисы работы с файлами Excel для модуля тестирования: генерация шаблонов и импорт вопросов."""

import io
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.core.files.uploadedfile import UploadedFile

from testing_app.models import Question, AnswerOption, QuestionCategory, TestingAuditLog


def generate_question_import_template() -> openpyxl.Workbook:
    """Генерирует эталонную книгу Excel (шаблон) для последующего импорта вопросов в базу данных.

    Шаблон содержит стилизованную шапку, понятные наименования колонок,
    подсказки по заполнению и несколько демонстрационных примеров вопросов.

    Returns:
        openpyxl.Workbook: Оформленная рабочая книга с шаблоном импорта.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Вопросы для импорта"
    ws.views.sheetView[0].showGridLines = True

    # Цветовая гамма и шрифты
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    desc_font = Font(name="Calibri", size=9, italic=True, color="475569")
    desc_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    sample_font = Font(name="Calibri", size=9.5, color="1E293B")
    correct_font = Font(name="Calibri", size=9.5, bold=True, color="166534")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    headers = [
        "Категория вопроса",
        "Текст вопроса",
        "Вариант 1",
        "Вариант 2",
        "Вариант 3",
        "Вариант 4",
        "Номер правильного ответа (1-4)",
        "Сложность (Низкая / Средняя / Повышенная / Высокая)",
        "Пояснение к ответу (необязательно)"
    ]

    # 1. Заголовки колонок (Строка 1)
    ws.row_dimensions[1].height = 30
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 2. Описание / подсказка (Строка 2)
    ws.row_dimensions[2].height = 24
    descriptions = [
        "Существующая или новая категория",
        "Полный текст формулировки вопроса",
        "Первый вариант ответа",
        "Второй вариант ответа",
        "Третий вариант ответа",
        "Четвертый вариант ответа",
        "Цифра от 1 до 4",
        "По умолчанию: Средняя",
        "Комментарий или ссылка на регламент"
    ]
    for col_idx, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.fill = desc_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 3. Примеры заполнения (Строки 3-5)
    samples = [
        [
            "Конструкция и системы ВС",
            "Каково минимально допустимое давление в гидросистеме перед вылетом ВС согласно РЛЭ?",
            "150 кг/см²",
            "210 кг/см²",
            "280 кг/см²",
            "320 кг/см²",
            2,
            "Средняя",
            "Согласно РЛЭ разд. 4 нормальное рабочее давление составляет 210 кг/см²."
        ],
        [
            "Авиационная безопасность",
            "Какой документ дает право на выполнение технического обслуживания воздушного судна?",
            "Свидетельство специалиста с соответствующей квалификационной отметкой",
            "Удостоверение сотрудника авиакомпании",
            "Внутренний пропуск аэропорта",
            "Водительское удостоверение",
            1,
            "Низкая",
            "ФАП-147: только действующее свидетельство с отметками дает допуск к ТО ВС."
        ],
        [
            "Охрана труда и техника безопасности",
            "Какова периодичность проверки знаний по охране труда для персонала, выполняющего ТО ВС?",
            "1 раз в 3 года",
            "1 раз в 6 месяцев",
            "1 раз в 12 месяцев",
            "Только при приеме на работу",
            2,
            "Средняя",
            "В соответствии с положением компании периодическая проверка проводится каждые 6 месяцев."
        ]
    ]

    for row_idx, sample_row in enumerate(samples, 3):
        ws.row_dimensions[row_idx].height = 26
        for col_idx, val in enumerate(sample_row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = correct_font if col_idx == 7 else sample_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [7, 8] else "left", vertical="center", wrap_text=True)
            cell.border = thin_border

    # 4. Настройка ширины колонок
    col_widths = [26, 45, 25, 25, 25, 25, 20, 22, 35]
    for idx, width in enumerate(col_widths, 1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A3"
    return wb


def import_questions_from_excel(file_obj: Any, user=None) -> Dict[str, Any]:
    """Выполняет импорт вопросов и вариантов ответов из переданного файла Excel.

    Args:
        file_obj (Any): Файловый объект (.xlsx), переданный из формы.
        user (Optional[User]): Пользователь, загрузивший файл (для привязки автора и аудита).

    Returns:
        Dict[str, Any]: Сводный результат импорта с полями:
            - 'success' (bool): Успешность операции;
            - 'total_rows' (int): Общее число обработанных строк;
            - 'created_questions' (int): Количество успешно созданных вопросов;
            - 'created_categories' (int): Количество созданных новых категорий;
            - 'skipped_questions' (int): Количество пропущенных вопросов (дубликаты/пустые);
            - 'errors' (List[str]): Список сообщений об ошибках по строкам;
            - 'created_items' (List[Dict[str, str]]): Краткая информация о созданных вопросах.
    """
    result = {
        "success": False,
        "total_rows": 0,
        "created_questions": 0,
        "created_categories": 0,
        "skipped_questions": 0,
        "errors": [],
        "created_items": [],
    }

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as e:
        result["errors"].append(f"Ошибка чтения файла Excel: {str(e)}. Убедитесь, что загружаемый файл в формате .xlsx.")
        return result

    ws = wb.active
    if not ws:
        result["errors"].append("Файл не содержит активных листов с данными.")
        return result

    # Определяем карту соответствия текстовых сложностей
    difficulty_mapping = {
        "низкая": Question.Difficulty.EASY,
        "easy": Question.Difficulty.EASY,
        "1": Question.Difficulty.EASY,
        "средняя": Question.Difficulty.MEDIUM,
        "medium": Question.Difficulty.MEDIUM,
        "2": Question.Difficulty.MEDIUM,
        "повышенная": Question.Difficulty.HARD,
        "hard": Question.Difficulty.HARD,
        "3": Question.Difficulty.HARD,
        "высокая": Question.Difficulty.VERY_HARD,
        "very_hard": Question.Difficulty.VERY_HARD,
        "4": Question.Difficulty.VERY_HARD,
    }

    categories_cache = {cat.name.strip().lower(): cat for cat in QuestionCategory.objects.all()}
    rows_to_process = []

    # Чтение данных: пропускаем заголовок (строка 1) и возможную строку-подсказку (строка 2)
    start_row = 2
    header_cell = ws.cell(row=1, column=1).value
    sub_header_cell = ws.cell(row=2, column=1).value

    if sub_header_cell and "категория" in str(sub_header_cell).lower() and "существующая" in str(sub_header_cell).lower():
        start_row = 3

    for row_idx in range(start_row, ws.max_row + 1):
        cat_val = ws.cell(row=row_idx, column=1).value
        q_text_val = ws.cell(row=row_idx, column=2).value

        # Пропускаем пустые строки
        if not cat_val and not q_text_val:
            continue

        cat_str = str(cat_val).strip() if cat_val else ""
        q_text_str = str(q_text_val).strip() if q_text_val else ""

        opt1 = str(ws.cell(row=row_idx, column=3).value or "").strip()
        opt2 = str(ws.cell(row=row_idx, column=4).value or "").strip()
        opt3 = str(ws.cell(row=row_idx, column=5).value or "").strip()
        opt4 = str(ws.cell(row=row_idx, column=6).value or "").strip()

        correct_raw = ws.cell(row=row_idx, column=7).value
        diff_raw = str(ws.cell(row=row_idx, column=8).value or "").strip().lower()
        explanation = str(ws.cell(row=row_idx, column=9).value or "").strip()

        rows_to_process.append({
            "row_idx": row_idx,
            "category": cat_str,
            "text": q_text_str,
            "options": [opt1, opt2, opt3, opt4],
            "correct_raw": correct_raw,
            "difficulty": difficulty_mapping.get(diff_raw, Question.Difficulty.MEDIUM),
            "explanation": explanation,
        })

    result["total_rows"] = len(rows_to_process)
    if not rows_to_process:
        result["errors"].append("В загруженном файле не найдено строк с данными для импорта.")
        return result

    # Транзакционный импорт
    created_count = 0
    categories_created_count = 0

    try:
        with transaction.atomic():
            for item in rows_to_process:
                r_num = item["row_idx"]

                # Валидация категории
                if not item["category"]:
                    result["errors"].append(f"Строка {r_num}: не указано название категории.")
                    result["skipped_questions"] += 1
                    continue

                # Валидация вопроса
                if not item["text"]:
                    result["errors"].append(f"Строка {r_num}: отсутствует текст вопроса.")
                    result["skipped_questions"] += 1
                    continue

                # Валидация 4 вариантов
                options = item["options"]
                if any(not opt for opt in options):
                    result["errors"].append(
                        f"Строка {r_num}: заполнены не все 4 варианта ответов ({item['text'][:40]}...)."
                    )
                    result["skipped_questions"] += 1
                    continue

                # Валидация правильного ответа
                try:
                    correct_idx = int(float(item["correct_raw"]))
                    if correct_idx not in [1, 2, 3, 4]:
                        raise ValueError()
                except (ValueError, TypeError):
                    result["errors"].append(
                        f"Строка {r_num}: некорректный номер правильного ответа '{item['correct_raw']}' (должно быть число 1, 2, 3 или 4)."
                    )
                    result["skipped_questions"] += 1
                    continue

                # Поиск или создание категории
                cat_key = item["category"].lower()
                category = categories_cache.get(cat_key)
                if not category:
                    category, created = QuestionCategory.objects.get_or_create(
                        name=item["category"],
                        defaults={"description": f"Автоматически создана при импорте из файла"}
                    )
                    categories_cache[cat_key] = category
                    if created:
                        categories_created_count += 1

                # Проверка на дубликат вопроса в рамках категории
                existing_q = Question.objects.filter(
                    category=category,
                    text__iexact=item["text"]
                ).first()

                if existing_q:
                    # Вопрос уже существует — пропускаем во избежание дублирования банка
                    result["skipped_questions"] += 1
                    continue

                # Создаем вопрос
                question = Question.objects.create(
                    category=category,
                    text=item["text"],
                    explanation=item["explanation"],
                    status=Question.Status.ACTIVE,
                    difficulty=item["difficulty"],
                    author=user
                )

                # Создаем 4 варианта ответа
                answer_options = []
                for idx, opt_text in enumerate(options, 1):
                    answer_options.append(
                        AnswerOption(
                            question=question,
                            text=opt_text,
                            order_num=idx,
                            is_correct=(idx == correct_idx)
                        )
                    )
                AnswerOption.objects.bulk_create(answer_options)

                created_count += 1
                result["created_items"].append({
                    "category": category.name,
                    "text": question.text[:70] + ("..." if len(question.text) > 70 else ""),
                    "correct_option": correct_idx,
                })

        # Фиксация в журнале аудита
        if created_count > 0:
            TestingAuditLog.objects.create(
                user=user,
                action="questions_import",
                object_repr=f"Импорт вопросов: создано {created_count}, категорий {categories_created_count}",
                details={
                    "total_rows": len(rows_to_process),
                    "created": created_count,
                    "skipped": result["skipped_questions"],
                    "categories_created": categories_created_count,
                    "errors_count": len(result["errors"]),
                }
            )

        result["success"] = True
        result["created_questions"] = created_count
        result["created_categories"] = categories_created_count

    except Exception as exc:
        result["success"] = False
        result["errors"].append(f"Критическая ошибка транзакции базы данных при импорте: {str(exc)}")

    return result
